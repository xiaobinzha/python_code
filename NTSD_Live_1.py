 
import psycopg2
 
import datetime
from datetime import date, timedelta
#import xlsxwriter
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.cell import Cell  
import pandas as pd
import numpy as np
import os
import sys
import logging
import locale
from openpyxl.styles import Alignment
import dateutil.relativedelta
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Color, Fill, Font, Border,Side, PatternFill
from openpyxl.styles import colors
import openpyxl 
from openpyxl import Workbook
from datetime import date, timedelta
from win32com import client
import win32api
import pathlib
from functools import reduce
from xlutils.copy import copy #work xls
from xlrd import open_workbook
import xlwt
 

dirpath = os.getcwd()
print (dirpath)
MAXCOL=68
TITLE_ROW = 7
 
last_day_month = date.today().replace(day=1) - timedelta(days=1)
start_day_month = date.today().replace(day=1) - timedelta(days=last_day_month.day)

start_day__qt =start_day_month - dateutil.relativedelta.relativedelta(months=2)
 
file_temp =  dirpath +  '\\data.xls'
  
filename1 =   dirpath +  '\\NTSD_' + str(last_day_month) + '.xlsx'

def add_header () -> dict:
    #wb = load_workbook(filename1)
    data_dict = pd.read_excel(file_temp, header=None) #reading file
   # data = pd.concat(data_dict.values(), axis=0)
    new_header= (data_dict.head(TITLE_ROW))
     
    
    return (new_header)

def move_book():
    book=open_workbook(file_temp, formatting_info=True)
    book1=copy(book)
    sheet=book.sheet_by_name('All Applicant Data')
    sheet1=book1.get_sheet('All Applicant Data')
    totalrows=4
    print (totalrows)
    j,k = 0, 0
    while k < totalrows:
        for i in range(0,totalrows):
            row=sheet.cell_value(i,1)
            sheet1.write(j,1,row)
        j+=totalrows
        k += 1
    book1.save(filename1)

def write_book(header, df_new):
    wb = load_workbook(filename1)
    ws = wb.get_sheet_by_name('All Applicant Data')

    redFill =PatternFill(fill_type='solid',start_color='5B9BD5',end_color='5B9BD5') 
    side  = Side(border_style='thin',  color="FF000000")
    border  = Border(left=side, right=side, top=side, bottom=side)
    side_thin  = Side(border_style='thin',  color="FF000000")
    thin_border = Border(
    left=Side(border_style=None, color='FFFFFF'),
    right=Side(border_style=None, color='FFFFFF'),
    top=Side(border_style=None, color='FFFFFF'),
    bottom=Side(border_style=None, color='FFFFFF'))

    thick_border = Border(
    left=Side(border_style=None, color='FFFFFF'),
    right=Side(border_style=None, color='FFFFFF'),
    top=Side(border_style='thick', color='00000000'),
    bottom=Side(border_style=None, color='FFFFFF'))


    ws.cell(1, 1, header[0][0])
    ws.cell(1, 1).font = Font(bold=True, color="FFFFFF", name="Verdana", size=8)     
    ws.merge_cells('A' + str(1) + ':BO' + str(1))
    ws['A' + str(1)].alignment = Alignment(horizontal='center', vertical = 'center')
    #ws['A' + str(1)].border =border
    ws['A' + str(1)].fill = redFill
         
    for x in range(2,TITLE_ROW+1): 
        j=x-1
        ws.cell(x, 1, header[0][j])
        ws.merge_cells('A' + str(x) + ':BO' + str(x))
        ws['A' + str(x)].alignment = Alignment(horizontal='center', vertical = 'center')
        ws['A' + str(x)].fill = redFill       
        ws.cell(x, 1).font = Font(color="FFFFFF", name="Verdana", size=8)   
    
    for x in range(1,68):
        ws.cell(8, x).alignment = Alignment(wrapText=True)
     
     
    o_companycode=''
    df_new.reset_index(inplace = True,drop = True)

    for index, row in df_new.iterrows():
        if (o_companycode != str(row[7])+str(row[9])+str(row[10]) + str(row[11]) ):
            for col_num  in range(1, MAXCOL):
                 ws.cell(index+TITLE_ROW+2, col_num).border = thick_border
        else:
                for col_num  in range(1, MAXCOL):
                 ws.cell(index+TITLE_ROW+2, col_num).border = None #thin_border 
        o_companycode = str(row[7])+str(row[9])+str(row[10]) + str(row[11])
  
    ws.auto_filter.ref = 'A8:BO8'
    ws.freeze_panes ='A9'

    wb.save(filename1)

def Emailer(message, subject, recipient):
    import win32com.client as win32   

    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = recipient
    mail.Subject = subject
    mail.GetInspector 

    index = mail.HTMLbody.find('>', mail.HTMLbody.find('<body')) 
    mail.HTMLbody = mail.HTMLbody[:index + 1] + message + mail.HTMLbody[index + 1:]
    mail.Attachments.Add(attachment)
    mail.Display(True)
    #mail.Send()

if __name__ == '__main__':

    data = pd.read_excel(file_temp, header=7) #reading file
    df=data[data.duplicated(subset=['First Name', 'Last Name', 'Property ID', 'Credit Run'],keep=False )]
    #df_new=df.drop_duplicates(subset=['First Name', 'Last Name', 'Property ID', 'Original Score','Final Score','Credit Run'], keep = False )
    df_new = df
 
    df_new=df_new.sort_values(by=['First Name', 'Last Name', 'Property ID', 'Applicant ID','Credit Run'])
    print(df_new['Positive Employment'])
    df_new.reset_index(drop = True)
    df_new.dropna()
    header = add_header()
    
    #header.append (df_new, ignore_index=True, sort=False)
    #df_new=header
    #print (df_new)
    df_new['Positive Employment'] = '=CONCATENATE("' + df_new['Positive Employment'] + '")'
    df_new['Positive Housing'] = '=CONCATENATE("' + df_new['Positive Housing'] + '")'
    
    df_new.to_excel (filename1, sheet_name='All Applicant Data', index=False, startrow=TITLE_ROW)
    #move_book()
    write_book(header, df_new)
    attachment  =  filename1  

    filelink = """
    <p>  <a href=%s>%s</a>  </p>
    """%(attachment, filename1)

#print (filelink)
#Emailer(filelink, "New pinac for jen.C@Yardi.Com@Yardi.Comis attached", "xiaobin.zhang@yardi.com") 
#Emailer(filelink, "Case 10091404 Monthly Dup Report-- 	Carl.Enberg@Yardi.com", "xiaobin.zhang@yardi.com") 


    
sys.exit(0)
quit()
