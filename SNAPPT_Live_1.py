 
from tkinter import FIRST
import psycopg2
 
import datetime
from datetime import date, timedelta,datetime
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.cell import Cell  
import pandas as pd
import numpy as np
import os
import sys
import logging
import locale
from openpyxl.styles import Alignment
import dateutil.relativedelta
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Color, Fill, Font, Border,Side, PatternFill
from openpyxl.styles import colors
import openpyxl 
from openpyxl import Workbook
 
from win32com import client
 
import pathlib
from functools import reduce
import win32api
from shutil import copyfile
from config import config
from config2 import config2
import dateutil.relativedelta 
 
from dateutil.relativedelta import relativedelta
from win32com.client import Dispatch
import win32com.client as win32
import codecs
from pyexcelerate import Workbook, Color, Style, Fill
import pyexcelerate
import psutil

import easygui
import warnings
warnings.filterwarnings("ignore")
import re 
from tkinter import filedialog
from tkinter import * 
import tkinter as tk

dirpath = os.getcwd()
print (dirpath)
d =  datetime.today()  
end_date = d.strftime("%x")
 
last_day_month = date.today().replace(day=1) - timedelta(days=1)
start_day_month = date.today().replace(day=1) - timedelta(days=last_day_month.day)

start_day__qt =start_day_month - dateutil.relativedelta.relativedelta(months=2)



today =  date.today()
idx =  (today.weekday() + 1) % 7
w_start_date = sys.argv[1] if len(sys.argv) > 1 else today -  timedelta(7+idx)
w_end_date =  sys.argv[2] if len(sys.argv) > 2 else w_start_date + timedelta(days=6)

w_start_date = w_start_date.strftime("%m/%d/%Y")  
w_end_date=w_end_date.strftime("%m/%d/%Y")  
 


filename =    d.strftime("%m/%d/%Y")  
filename1 =  "\\ysifwfs07\\Vol2\ANALYTICS\\Reports\\SNAPPT\\SNAPPT " + filename.replace("/", ".") + ".xlsx"
filename1 = "O:\\ANALYTICS\\Reports\\SNAPPT\\SNAPPT "  +   filename.replace("/", ".") + ".xlsx"
 

redFill =PatternFill(fill_type='solid',start_color='5B9BD5',end_color='5B9BD5') 
greenFill =PatternFill(fill_type='solid',start_color='b2f5d9',end_color='b2f5d9') 
yellowFill =PatternFill(fill_type='solid',start_color='Ffff00',end_color='Ffff00') 
side  = Side(border_style='thin',  color="FF000000")
border  = Border(left=side, right=side, top=side, bottom=side)
side_thin  = Side(border_style='thin',  color="FF000000")



#prior_file = "O:\\ANALYTICS\\Reports\\SNAPPT\\SNAPPT 6.17.24"  + ".xlsx" 
root = tk.Tk()
root.attributes('-alpha', 0.0)
# Always have it on top
root.attributes('-topmost', True) 
prior_file = tk.filedialog.askopenfilename(parent=root,initialdir = "O:\ANALYTICS\Reports\SNAPPT" + "/",title = "Select Prior file       ",filetypes = (("excel files","*.xlsx"),("all files","*.*"))) 
root.destroy()

if  not os.path.exists(prior_file):
        output = easygui.msgbox("There is no last week file! ", "Error")
        sys.exit(0)
        quit(0)    


thin_border = Border(
    left=Side(border_style='thin', color='00000000'),
    right=Side(border_style='thin', color='00000000'),
    top=Side(border_style='thin', color='00000000'),
    bottom=Side(border_style='thin', color='00000000'))

thick_border = Border(
    top=Side(border_style='thick', color='00000000'))


def generate_terminate_code ( ):
    conn = None
    params = config2()
    
    conn = psycopg2.connect(**params )
 
    cur = conn.cursor()

    SQL = ("""   select company_code  from "Test_Code_Master"  ; """   )
        
    cur.execute(SQL) 
    #col = cur.description
    col = [i[0] for i in cur.description]
    a = np.array(col)
    rows =   [item[0] for item in cur.fetchall()]
            
    #df_codes = pd.DataFrame(rows, columns = a)
    cur.close()
    conn.close()
    return rows

#write exlude test codes; replace with abreviation
def dispatch(app_name:str):
    try:
        from win32com import client
        app = client.gencache.EnsureDispatch(app_name)
    except AttributeError:
        # Corner case dependencies.
        import os
        import re
        import sys
        import shutil
        # Remove cache and try again.
        MODULE_LIST = [m.__name__ for m in sys.modules.values()]
        for module in MODULE_LIST:
            if re.match(r'win32com\.gen_py\..+', module):
                del sys.modules[module]
        shutil.rmtree(os.path.join(os.environ.get('LOCALAPPDATA'), 'Temp', 'gen_py'))
        from win32com import client
        app = client.gencache.EnsureDispatch(app_name)
    return app         
 

def save_workbook (s_name):
    excel = dispatch('Excel.Application')
    excel.Interactive = False
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = excel.Workbooks.Open(s_name)
     
   # getNumSheet = wb.Worksheets.count+1

    #for i in range  (1, getNumSheet ):
      #  excel.Worksheets(i).Activate() 

    wb.Worksheets('SNAPPT').Select()    
    excel.ActiveSheet.Columns.AutoFit()

 
        
#Save changes in a new file
    excel.DisplayAlerts = False    
    wb.SaveAs (s_name)
    wb.Close()
 
 
def connect_trans ():
    SQL =    ( """  
             
   select node_code,min(dtcreated)::date first_invite_date,  max(dtcreated)::date last_invite_date  from voi_snappit_invite_applicant  
group by node_code  

"""  )
    list_term_code =  generate_terminate_code () 

    string_replace = "  ('" + "','".join(name.upper().strip().replace("'", r"\'") for name in list_term_code) + "'"  + ')'
    SQL = SQL.replace(   "('APP')",   string_replace     )

    conn = None
 
    params = config()
    conn = psycopg2.connect(**params)       
    cur = conn.cursor()
    cur.execute(SQL) 
    col = [i[0] for i in cur.description]
    conn.set_client_encoding('ISO-8859-1') 
    rows = cur.fetchall()
    cur.close()
    trans_data = pd.DataFrame(rows, columns = col)

   # Row_list =[] 
  
# Iterate over each row 
   # for rows in data.itertuples(): 
    # Create list for the current row 
   #     my_list =[rows.dd, rows.dd_1, rows.dd_2] 
      
    # append the list to the final list 
   # Row_list.append(my_list) 
    return trans_data


##### weekly

def connect_all_prop ():
    SQL =    ( """  
             
  select  co_code, company_name, pr_code, market_rate_units,  (select vendor_name from services where voi = service_id) voi, 
(select vendor_name from services where voi_ondmd = service_id) voi_on_demand, 
node_name, node_city, node_state, auxiliary_node.email_address as contact_email , null as first_show_on_report  , null   first_show, --,  to_char(current_date,'MM/DD/YYYY') as first_show_on_report,
            null   first_invite_date, null  last_invite_date
               

 from node_services inner join  node_tbl on  pr_code = node_tbl.node_code inner join  auxiliary_node on node_tbl.node_code=auxiliary_node.node_code inner join 
	 company_tbl  on company_tbl.company_code=node_tbl.company_code  
	 where canceled = false 
	 and    (voi=29 or voi_ondmd=29) 
 and pr_code = node_tbl.node_code  and auxiliary_node.node_code=node_tbl.node_code 
 and node_tbl.company_code = company_tbl.company_code
 and company_tbl.company_code not in ('APP') 

"""  )
    list_term_code =  generate_terminate_code () 

    string_replace = "  ('" + "','".join(name.upper().strip().replace("'", r"\'") for name in list_term_code) + "'"  + ')'
    SQL = SQL.replace(   "('APP')",   string_replace     )

    conn = None
 
    params = config()
    conn = psycopg2.connect(**params)       
    cur = conn.cursor()
    cur.execute(SQL) 
    col = [i[0] for i in cur.description]
    conn.set_client_encoding('ISO-8859-1') 
    rows = cur.fetchall()
    cur.close()
    data = pd.DataFrame(rows, columns = col)

   # Row_list =[] 
  
# Iterate over each row 
   # for rows in data.itertuples(): 
    # Create list for the current row 
   #     my_list =[rows.dd, rows.dd_1, rows.dd_2] 
      
    # append the list to the final list 
   # Row_list.append(my_list) 
    return data
      
   
def connect_cancelled_prop ():
    SQL =    ( """  
             
  select  node_code from  node_tbl   where canceled = true 
 
"""  )
    

    conn = None
 
    params = config()
    conn = psycopg2.connect(**params)       
    cur = conn.cursor()
    cur.execute(SQL) 
    col = [i[0] for i in cur.description]
    conn.set_client_encoding('ISO-8859-1') 
    rows = cur.fetchall()
    cur.close()
    data = pd.DataFrame(rows, columns = col)

   # Row_list =[] 
  
# Iterate over each row 
   # for rows in data.itertuples(): 
    # Create list for the current row 
   #     my_list =[rows.dd, rows.dd_1, rows.dd_2] 
      
    # append the list to the final list 
   # Row_list.append(my_list) 
    return data

def connect_old_book():
     
    old_data = pd.read_excel(prior_file, header=0  ) # before the title name) #reading file
   

    old_data = old_data.replace('\n', ' ').replace('\r', ' ') 
    old_data = old_data.replace('\\n', ' ').replace('\\r', ' ') 
    old_data.iloc[0].replace(to_replace=[r"\\t|\\n|\\r", "\t|\n|\r"], value=["",""], regex=True, inplace=True ) 
    old_data.columns = [x.replace("\n", " ") for x in old_data.columns.tolist()]
    #data_dict=data_dict.drop(['Property ID'], inplace=True, axis=1 ) 
    old_data.rename(columns=lambda x: x.strip(),   inplace=True)
    old_data = old_data.sort_values (by= ['Disabled',   'first_show_on_report', 'co_code','ID']  )


    old_data['Node Code']=old_data['Node Code'].str.strip() 
    old_data['Node Code']=old_data['Node Code'].str.upper()
    
    old_data= old_data.drop_duplicates('Node Code', keep=FIRST)

    all_data = connect_all_prop()
 
    disabled_node = old_data[~(old_data['Node Code'].isin(all_data['pr_code'].values.tolist()))] 

    cancelled_node = connect_cancelled_prop()
    
    disabled_node = disabled_node[~disabled_node['Node Code'].isin(cancelled_node['node_code'])]

    disabled_node ['Disabled'] = 'Yes'
    disabled_node ['voi'] = ''
    disabled_node ['voi_on_demand'] = ''
    disabled_node ['first_show'] = 'No'
    

    disabled_node['first_show_on_report'] = pd.to_datetime(disabled_node['first_show_on_report']).dt.date 
    
    disabled_node = disabled_node.sort_values (by= ['Disabled',   'first_show_on_report', 'co_code', 'node_name' ]  )
  
    disabled_node.insert(1, 'NEW_ID', range(1, 1 + len(disabled_node)))
    #disabled_node ["ID"] = disabled_node['NEW_ID']

    disabled_node = disabled_node.drop (['ID' ] , axis = 1)

    disabled_node.rename(columns={'Node Code': 'pr_code' }, inplace=True)

    new_data_old=all_data[(all_data['pr_code'].isin(old_data['Node Code'].values.tolist()))]  ## all active nodes
    new_data_old ['first_show_on_report'] =   new_data_old['pr_code'].map(old_data.set_index('Node Code')['first_show_on_report']).fillna(new_data_old['first_show_on_report'])
   

    new_data_old.insert(0, "NEW_ID", 0)
    new_data_old ['NEW_ID'] =   new_data_old['pr_code'].map(old_data.set_index('Node Code')['ID']).fillna(new_data_old['NEW_ID'])
    
    new_data_old['first_show_on_report']=   pd.to_datetime(new_data_old['first_show_on_report'] ).dt.date

    new_data_old = new_data_old.sort_values (by= ['first_show_on_report', 'co_code', 'node_name'  ] )
    new_data_old.reset_index(inplace=True, drop=True)

    l_ID = []
    for index, row in new_data_old.iterrows():
        l_ID.append(index+1)

    new_data_old['NEW_ID'] =l_ID
   # new_data_old.drop(['ID'],inplace=True, axis=1)
   # new_data_old.reset_index()

          #| (data['dd_1'].isin(df_weekly_data['dd_1'].values.tolist())) | (data['dd_2'].isin(df_weekly_data['dd_2'].values.tolist())) ] 

    new_data_old ['first_show'] = 'No'
  
    new_data_new = all_data[(~all_data['pr_code'].isin(old_data['Node Code'].values.tolist()))]

    new_data_new ['first_show'] = 'Yes'
    new_data_new['first_show_on_report']= pd.to_datetime('today').date().strftime('%m/%d/%Y')

    #new_data_new.drop(['NEW_ID'],inplace=True, axis=1)
    new_data_new =        new_data_new.sort_values (by= ['first_show_on_report', 'co_code' , 'node_name' ] )             
    new_data_new.insert(0, 'NEW_ID', range(len(new_data_old)+1, len(new_data_new) + len(new_data_old)+1))
 
   
    #df = pd.concat([df_dd_2_new, df_dd_1_new, df_dd_new  ]).drop_duplicates('node_code', keep=FIRST)
    new_data_old ['Disabled'] = 'No'
    new_data_new ['Disabled'] = 'No'

    df = pd.concat([ new_data_old, new_data_new, disabled_node   ])#.drop_duplicates('node_code', keep=FIRST)
 
    df['first_show_on_report'] = pd.to_datetime(df['first_show_on_report']).dt.date 
   # df['first_show_on_report'] = df['first_show_on_report'].dt.strftime('%m/%d/%Y')
   # df ['max_code']=   df.groupby('type')['company_code'].transform('max')
    #df ['min_code']=   df.groupby('type')['company_code'].transform('min')

    list_trans = connect_trans()
    df['first_invite_date']=  df['pr_code'].map(list_trans.set_index('node_code')['first_invite_date']).fillna(df['first_invite_date'])
    df['last_invite_date']=  df['pr_code'].map(list_trans.set_index('node_code')['last_invite_date']).fillna(df['last_invite_date'])
    df.replace(np.nan, '',   regex=True, inplace=True)
  
    #df.insert(0, 'New_ID', range(1, 1 + len(df)))
    
    #df = df.drop (['ID', 'Node Code' ] , axis = 1)
    
    
    df = df.sort_values (by= ['Disabled', 'NEW_ID'  ] )
    df.reset_index(inplace=True, drop=True)  
    
    return df


  

def write_dup_book(header, df_new,  ws):
    MAXCOL=len(df_new.columns)  + 2

    for x in range(1,len(df_new.columns)+1):
         
        ws.cell(1, x).fill = redFill     
        ws.cell(1, x).font = Font(color="FFFFFF", name="Verdana", size=12)   
        ws.cell(1, x).alignment = Alignment(horizontal='left', vertical = 'center')
     
    o_companycode=n_code=''
    df_new.reset_index(inplace = True,drop = True)
    first_time = True 
    for index, row in df_new.iterrows():
      
            if (  o_companycode  !=   str(row[11]  ) and  row[15]=='No' )   :    #11 and n_code  !=  ( str(row[0])  ) first_show_on_report
              for col_num  in range(1, MAXCOL-1):
                   ws.cell(index+2, col_num).border =  thick_border
          
                 

            if (str(row[12]) == 'Yes') :
              for col_num  in range(1, MAXCOL-1):
                       ws.cell(index+2, col_num).fill = greenFill   
            if (str(row[15]) == 'Yes') :
              for col_num  in range(1, MAXCOL-1):
                       ws.cell(index+2, col_num).fill = yellowFill   
              if(first_time==True):
                   for col_num  in range(1, MAXCOL-1):
                       ws.cell(index+2, col_num).border =  thick_border
                       ws.cell(index+2, col_num).border =  thick_border
                   first_time=False
                      
            ws.cell(index+2, col_num).alignment = Alignment(horizontal='center', vertical = 'center')         
            o_companycode = (str(row[11])   ) #[0:6] #11
          
  
    for col_num  in range(1, MAXCOL-1):
                   ws.cell(index+3, col_num).border = thick_border
                   ws.cell(index+3, col_num).alignment = Alignment(horizontal='center', vertical = 'center')
    
    ws.auto_filter.ref = 'A1:P1'
    ws.freeze_panes ='A2'

   

def send_to_finance(filename1):


    if (not os.path.exists(filename1)): 
         return
    
    #signature_name="signature (Xiaobin.Zhang@Yardi.com)"
    #sig_files_path = 'AppData\Roaming\Microsoft\Signatures\\' + signature_name + '_files\\'
    #sig_html_path = 'AppData\\Roaming\\Microsoft\\Signatures\\' + signature_name + '.htm'
    signature_file = "C:\App\Data\Sig\sig.htm"
    #signature_path = os.path.join((os.environ['USERPROFILE']), sig_files_path) # Finds the path to Outlook signature files with signature name "Work"
    #html_doc = os.path.join((os.environ['USERPROFILE']),sig_html_path)     #Specifies the name of the HTML version of the stored signature
    #html_doc = html_doc.replace('\\\\', '\\')


    html_file = codecs.open(signature_file, 'r', 'utf-8', errors='ignore') #Opens HTML file and converts to UTF-8, ignoring errors
    signature_code = html_file.read()               #Writes contents of HTML signature file to a string

    #signature_code = signature_code.replace((signature_name + '_files/'), signature_path)      #Replaces local directory with full directory path
    html_file.close()

    
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = "xiaobin.zhang@yardi.com"
    mail.Subject = "Weekly New SNAPPT List Report - david.carr@Yardi.Com "  
    filename2 =  filename1.replace('O:\\ANALYTICS\\Reports\\', '\\\ysifwfs07\\Vol2\ANALYTICS\\Reports\\') # + filename.replace("/", ".") + ".xlsx"
        #path  = "\"\\\\windows_Server\\golobal_directory\\the folder\\file yyymm.xlsx\""
    path = '"' + filename2 + '"'
    string = """<a href=""" +  path + ' style=text-decoration: none>' + filename1 +  '<' +  r'\a'  + '>'

      #  string.replace('\\a>', '\a>')
    #mail.body = string
     
    mail.HTMLbody =   string + " <BR><BR><BR> "  +signature_code + " <BR><BR><BR> "
    mail.Attachments.Add(Source=filename1)
    mail.send
    

def send_book( filename1, recipient, filename2=''):

    outlook = win32.Dispatch('outlook.application')
    if (os.path.exists(filename1) or os.path.exists(filename2)):
        send_account = None
        From =None

        for myEmailAddress in outlook.Session.Accounts:
            if "Xiaobin.Zhang@yardi.com" in str(myEmailAddress):
                From = myEmailAddress
                break

        
        mail = outlook.CreateItem(0)
        mail.To = recipient
        #mail.SentOnBehalfOfName = 'RS_Analytics@yardi.com'
        mail.SentOnBehalfOfName = 'xiaobin.zhang@yardi.com'
        
        if From != None:
            mail._oleobj_.Invoke(*(64209, 0, 8, 0, From))


        #mail.Cc=' YRScushwake@yardi.com; Judy.Tutt@cushwake.com; Natalie.Gatjanis@cushwake.com'
    #mail.Subject = subject
        mail.GetInspector 
        mail.Subject = "Weekly SNAPPT Report - David.Carr@Yardi.Com ; "  
       
        index = mail.HTMLbody.find('>', mail.HTMLbody.find('<body')) 
   # mail.HTMLbody = mail.HTMLbody[:index + 1] + message + mail.HTMLbody[index + 1:]
        if (os.path.exists(filename1)):
            #mail.Attachments.Add(filename1)
            pass
           #pass
        if (os.path.exists(filename2)):
            pass 
            #mail.Attachments.Add(filename2)    

        
        filename2 =  filename1.replace('O:\\ANALYTICS\\New Property Lists\\', '\\\ysifwfs07\\Vol2\ANALYTICS\\New Property Lists\\') # + filename.replace("/", ".") + ".xlsx"
        #path  = "\"\\\\windows_Server\\golobal_directory\\the folder\\file yyymm.xlsx\""
        path = '"' + filename2 + '"'
        string = """<a href=""" +  path + ' style=text-decoration: none>' + filename1 +  '<' +  r'\a'  + '>'

      #  string.replace('\\a>', '\a>')

       
        mail.body = filename1 
        mail.To = 'xiaobin.zhang@yardi.com'
        mail.Save()
        #mail.From = 'Xiaobin.zhang@yardi.com'
        mail.send

def kill_excel():
    
    for proc in psutil.process_iter():
        if proc.name() == "EXCEL.EXE":
            proc.kill()


if __name__ == '__main__':
    kill_excel()

    if os.path.exists(filename1 ):
         os.remove(filename1) # send_to_finance(filename1) 
 
          
    f = open("weekly_snappt.txt", "w") 
    error = 0
    
    try:
 
        df_new_prop = connect_old_book()



        #data.to_excel (filename1, sheet_name='duplidate property', index=False) #, startrow=0

        writer2 = pd.ExcelWriter(filename1)
        df_new_prop.reset_index(  inplace=True, drop=True)


        df_new_prop = df_new_prop 
        df_new_prop.rename(columns={'NEW_ID': 'ID', 'pr_code': 'Node Code'}, inplace=True)
        df_new_prop.to_excel (writer2,  sheet_name= 'SNAPPT', index=False, startrow=0)

        writer2.close() 



        wb = load_workbook(filename1)
        ws = wb['SNAPPT']

        # todolist
            
        write_dup_book(filename1, df_new_prop,   ws)



        wb.save(filename1)
        save_workbook (filename1) 


    
        #send_book (filename1, 'xiaobin.zhang@yardi.com' )

    except Exception as Argument:

        f.write(str(Argument))
        error =1 
    finally: 
        send_to_finance(filename1)     
        f.close() 
    
        if (error ==0 and os.path.exists(f.name)): 
            os.remove(f.name)

sys.exit(0)
quit()
