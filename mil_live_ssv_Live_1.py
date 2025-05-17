 
import psycopg2
from config import config
from datetime import date, timedelta
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, Fill, Font, Border,Side 
from openpyxl.cell import Cell  
import pandas as pd
import numpy as np
import os
import sys
import logging
import dateutil.relativedelta 
from datetime import datetime,  date, timedelta
from dateutil.relativedelta import relativedelta

import psycopg2
 
import datetime
from datetime import date, timedelta
#import xlsxwriter
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.cell import Cell  
import pandas as pd
import numpy as np
import locale
from openpyxl.styles import Alignment
import dateutil.relativedelta
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Color, Fill, Font, Border,Side, PatternFill
from openpyxl.styles import colors
import openpyxl 
from openpyxl import Workbook
from datetime import date, timedelta
from win32com import client
import win32api
import pathlib
from functools import reduce
from xlutils.copy import copy #work xls
from xlrd import open_workbook
import xlwt

dirpath = os.getcwd()


def connect_ssv_prop ( ) :
    conn = None
    params = config()
    
    conn = psycopg2.connect(**params )
    
    

    cur = conn.cursor()
        
    SQL = (""" select nc.node_code, node_name as "Property", market_rate_units as "Units" ,    'Manager' as manager2
 
from twnssv_emp_request ssv right join  node_tbl nc on ssv.node_code =  nc.node_code right join auxiliary_node an
on nc.node_code= an.node_code 
     where company_code in ('MIL')    and nc.node_code in ('U7597','W9765','BA661','T1694','W7746','Z0285','AJ840','T1686','AX705','AX715' )
 --and nc.canceled = false  -- and market_rate_units>0
--and nc.node_code= an.node_code 
group by nc.node_code, node_name, yardi_prop_code, manager2  order by 2, 1

 """)

    cur.execute(SQL)

    col = cur.description 
            
    rows = cur.fetchall()
    col = [i[0] for i in cur.description]
    a = np.array(col)
 
      
    df_nodes = pd.DataFrame(rows, columns = a)
 
    
    return df_nodes

def connect_twn_prop ( ) :
    conn = None
    params = config()
    
    conn = psycopg2.connect(**params )
    
    

    cur = conn.cursor()
        
    SQL = (""" select nc.node_code, node_name as "Property", market_rate_units as "Units" ,    'Manager' as manager2
 
from twn_request twn right join  node_tbl nc on twn.node_code =  nc.node_code right join auxiliary_node an
on nc.node_code= an.node_code 
     where company_code in ('MIL')    and nc.node_code in ( 'BA661' ) and   twn.node_code =  nc.node_code      
 --and nc.canceled = false
group by nc.node_code, node_name, yardi_prop_code, manager2  order by 2, 1

 """)

    cur.execute(SQL)

    col = cur.description 
            
    rows = cur.fetchall()
    col = [i[0] for i in cur.description]
    a = np.array(col)
 
      
    df_nodes = pd.DataFrame(rows, columns = a)
 
    
    return df_nodes

#detail
def connect_ssv_apps (start_day_month, last_day_month ) :
    conn = None
    params = config()
    
    conn = psycopg2.connect(**params )
   # print (start_day_month, last_day_month) 
 

    cur = conn.cursor()
    
      
    SQL = (""" select ssv.node_code as "Property Code",  node_name as "Property Name", 
           ssv.applicant_uuid,trim(coalesce(first_name,'')) || ' ' || trim(coalesce(last_name)) "APPLICANT NAME",  
	case  when COALESCE(pdf_path,'')<>''  and ( select_status = '0' or select_status='Request is successful') then 'Found' else 'Not found'  end "Status", ssv.create_stamp::date as "Date Processed"
	 
	from   twnssv_emp_request ssv , node_tbl nc, applicants  where company_code in ('MIL') and ssv.applicant_uuid = applicants.applicant_uuid  
and ssv.node_code =  nc.node_code  
and  id in (select max(id) from twnssv_emp_request group by node_code, date_trunc('month',twnssv_emp_request.create_stamp)::date, applicant_uuid)
and canceled=false  and nc.node_code in ('U7597','W9765','BA661','T1694','W7746','Z0285','AJ840','T1686','AX705','AX715' )
and ssv.create_stamp::date  >= '%s' and   ssv.create_stamp::date <= '%s'
 
order by ssv.create_stamp,1,2  
 """  % (start_day_month, last_day_month))

    cur.execute(SQL)

    col = cur.description 
            
    rows = cur.fetchall()
    col = [i[0] for i in cur.description]
    a = np.array(col)
 
      
    df_app = pd.DataFrame(rows, columns = a)

    df_app['week_start'] = pd.to_datetime(df_app['Date Processed'], errors='coerce')

    #df_app["week_start"] = df_app["week_start"].dt.to_period("W").dt.to_timestamp()

    df_app["week_start"] =  (df_app['week_start']+timedelta(days=1)).dt.to_period('W').apply(lambda r: r.start_time)-timedelta(days=1)

    df_app["week_start"] =  df_app["week_start"].dt.strftime('%m/%d/%y')  #+ chr(13) + chr(10) 
    #df_app= df_app.drop_duplicates ('applicant_uuid', keep='last')
   
    cur.close()
    return df_app

def connect_twn_apps (start_day_month, last_day_month ) :
    conn = None
    params = config()
    
    conn = psycopg2.connect(**params )
   # print (start_day_month, last_day_month) 
 

    cur = conn.cursor()
    
      
    SQL = (""" select twn.node_code as "Property Code",  node_name as "Property Name", 
           twn.applicant_uuid,trim(coalesce(first_name,'')) || ' ' || trim(coalesce(last_name)) "APPLICANT NAME",  
	case  when COALESCE(pdf_path,'')<>''  and ( select_status = '0' or select_status='Request is successful') then 'Found' else 'Not found'  end "Status", twn.create_stamp::date as "Date Processed"
	 
	from   twn_request twn , node_tbl nc, applicants  where company_code in ('MIL') and twn.applicant_uuid = applicants.applicant_uuid  
and twn.node_code =  nc.node_code  
and  id in (select max(id) from twn_request group by node_code, date_trunc('month',twn_request.create_stamp)::date, applicant_uuid)
and canceled=false  and nc.node_code in ( 'BA661'  )
and twn.create_stamp::date  >= '%s' and   twn.create_stamp::date <= '%s'
 
order by twn.create_stamp,1,2  
 """  % (start_day_month, last_day_month))

    cur.execute(SQL)

    col = cur.description 
            
    rows = cur.fetchall()
    col = [i[0] for i in cur.description]
    a = np.array(col)
 
      
    df_app = pd.DataFrame(rows, columns = a)

    df_app['week_start'] = pd.to_datetime(df_app['Date Processed'], errors='coerce')

    #df_app["week_start"] = df_app["week_start"].dt.to_period("W").dt.to_timestamp()

    df_app["week_start"] =  (df_app['week_start']+timedelta(days=1)).dt.to_period('W').apply(lambda r: r.start_time)-timedelta(days=1)

    df_app["week_start"] =  df_app["week_start"].dt.strftime('%m/%d/%y')  #+ chr(13) + chr(10) 
    #df_app= df_app.drop_duplicates ('applicant_uuid', keep='last')
   
    cur.close()
    return df_app

def connect_ssv_App (start_day_month, last_day_month ) :
    conn = None
    params = config()
    
    conn = psycopg2.connect(**params )
   # print (start_day_month, last_day_month) 
 

    cur = conn.cursor()
    
      
    SQL = (""" select ssv.id, select_status, 1 as processing, coalesce(pdf_path, '') pdf_path, case when ( select_status = '0' or select_status='Request is successful')  and coalesce(pdf_path, '') <>'' then 1 else 0 end found, applicant_uuid, node_name, ssv.create_stamp::date as "Date Processed", nc.node_code  
    from twnssv_emp_request ssv, node_tbl nc  where company_code in ('MIL')  and nc.canceled = false and nc.node_code in ('U7597','W9765','BA661','T1694','W7746','Z0285','AJ840','T1686','AX705','AX715')
and ssv.node_code =  nc.node_code     and ssv.create_stamp  >= '%s' and  ssv.create_stamp <= '%s' order by nc.node_code, applicant_uuid """  % (start_day_month, last_day_month))

    cur.execute(SQL)

    col = cur.description 
            
    rows = cur.fetchall()
    col = [i[0] for i in cur.description]
    a = np.array(col)
 
      
    df_app = pd.DataFrame(rows, columns = a)
    df_app= df_app.drop_duplicates ('applicant_uuid', keep='last')
    df_app['week_start'] = pd.to_datetime(df_app['Date Processed'], errors='coerce')
    df_app["week_start"] =  (df_app['week_start']+timedelta(days=1)).dt.to_period('W').apply(lambda r: r.start_time)-timedelta(days=1)
    df_app["week_start"] =  df_app["week_start"].dt.strftime('%y-%m/%d')  #+ chr(13) + chr(10) 
    
    #df_app_grp = df_app.groupby('node_code').agg({'Processing': lambda x: len(x), 'found': lambda x: len(x)})
   # df_app_grp = df_app.groupby('node_code')[['select_status', 'pdf_path']].agg({'applicant_uuid': lambda x: len(x), 
    #'applicant_uuid': lambda x:   ( 1 if (x['select_status'] == 0 and x['pdf_path'] != '') else  0 ).sum()})
    df_app_grp = df_app.groupby([ 'node_code', 'week_start']  , as_index=False )[["processing", "found"]].sum()#.reset_index(drop=True)
 
    df_app_grp ['perc']= ( df_app_grp['found']/df_app_grp['processing']*100).astype(int)
    df_app_grp ['perc']=  df_app_grp ['perc'].astype(str) + '%'
    
    df_app_grp=df_app_grp.pivot(index='node_code', columns='week_start', values=["processing", "found", "perc"])


    df_app_grp.columns = df_app_grp.columns.map(lambda index: f'{index[1]}_{index[0]}')
    df_app_grp.columns = df_app_grp.columns.str.replace('processing', '1processing').str.replace('found', '2found').str.replace('perc', '3perc')

    cols = list(df_app_grp.columns.sort_values())
 
    df_app_grp = df_app_grp[cols]

    df_app_grp.columns = df_app_grp.columns.str.replace('1processing', 'processing').str.replace('2found', 'found').str.replace('3perc', 'perc')

    df_app_grp.columns = df_app_grp.columns.str.replace('23-', '').str.replace('24-', '').str.replace('25-', '').str.replace('_', chr(13) + chr(10))
     
    ##new_month = start_day_month.strftime("%m-%Y")   + chr(13) + chr(10) 
    df_app_grp = df_app_grp.reset_index()
   # df_app_grp.columns=[new_month   + 'Processing', new_month      +'Found', new_month   +'Percent']
 
     
     
    #df_nodes.to_excel(filename1, index = True, header=True, float_format='%.00f' )
    cur.close()
    return df_app_grp


def connect_twn_App (start_day_month, last_day_month ) :
    conn = None
    params = config()
    
    conn = psycopg2.connect(**params )
   # print (start_day_month, last_day_month) 
 

    cur = conn.cursor()
    
      
    SQL = (""" select twn.id, select_status, 1 as processing, coalesce(pdf_path, '') pdf_path, case when ( select_status = '0' or select_status='Request is successful')  and coalesce(pdf_path, '') <>'' then 1 else 0 end found, applicant_uuid, node_name, twn.create_stamp::date as "Date Processed", nc.node_code  
    from twn_request twn, node_tbl nc  where company_code in ('MIL')  and nc.canceled = false and nc.node_code in ( 'BA661' )
and twn.node_code =  nc.node_code     and twn.create_stamp  >= '%s' and  twn.create_stamp <= '%s' order by nc.node_code, applicant_uuid """  % (start_day_month, last_day_month))

    cur.execute(SQL)

    col = cur.description 
            
    rows = cur.fetchall()
    col = [i[0] for i in cur.description]
    a = np.array(col)
 
      
    df_app = pd.DataFrame(rows, columns = a)
    df_app= df_app.drop_duplicates ('applicant_uuid', keep='last')
    df_app['week_start'] = pd.to_datetime(df_app['Date Processed'], errors='coerce')
    df_app["week_start"] =  (df_app['week_start']+timedelta(days=1)).dt.to_period('W').apply(lambda r: r.start_time)-timedelta(days=1)
    df_app["week_start"] =  df_app["week_start"].dt.strftime('%y-%m/%d')  #+ chr(13) + chr(10) 
    
    #df_app_grp = df_app.groupby('node_code').agg({'Processing': lambda x: len(x), 'found': lambda x: len(x)})
   # df_app_grp = df_app.groupby('node_code')[['select_status', 'pdf_path']].agg({'applicant_uuid': lambda x: len(x), 
    #'applicant_uuid': lambda x:   ( 1 if (x['select_status'] == 0 and x['pdf_path'] != '') else  0 ).sum()})
    df_app_grp = df_app.groupby([ 'node_code', 'week_start']  , as_index=False )[["processing", "found"]].sum()#.reset_index(drop=True)
 
    df_app_grp ['perc']= ( df_app_grp['found']/df_app_grp['processing']*100).astype(int)
    df_app_grp ['perc']=  df_app_grp ['perc'].astype(str) + '%'
    
    df_app_grp=df_app_grp.pivot(index='node_code', columns='week_start', values=["processing", "found", "perc"])


    df_app_grp.columns = df_app_grp.columns.map(lambda index: f'{index[1]}_{index[0]}')
    df_app_grp.columns = df_app_grp.columns.str.replace('processing', '1processing').str.replace('found', '2found').str.replace('perc', '3perc')

    cols = list(df_app_grp.columns.sort_values())
 
    df_app_grp = df_app_grp[cols]

    df_app_grp.columns = df_app_grp.columns.str.replace('1processing', 'processing').str.replace('2found', 'found').str.replace('3perc', 'perc')

    df_app_grp.columns = df_app_grp.columns.str.replace('23-', '').str.replace('24-', '').str.replace('25-', '').str.replace('_', chr(13) + chr(10))
     
    ##new_month = start_day_month.strftime("%m-%Y")   + chr(13) + chr(10) 
    df_app_grp = df_app_grp.reset_index()
   # df_app_grp.columns=[new_month   + 'Processing', new_month      +'Found', new_month   +'Percent']
 
     
     
    #df_nodes.to_excel(filename1, index = True, header=True, float_format='%.00f' )
    cur.close()
    return df_app_grp

#last_day_month = date.today().replace(day=1) - timedelta(days=1)
#start_day_month = date.today().replace(day=1) - timedelta(days=last_day_month.day)
 
def column_num_to_string(n):
    n, rem = divmod(n - 1, 26)
    next_char = chr(65 + rem)
    if n:
        return column_num_to_string(n) + next_char
    else:
        return next_char 
    
def write_book_2( df_detail,   wb , sheettype = 'ssv'):
   
    if (sheettype =='ssv'):
        ws = wb.get_sheet_by_name('Sheet2')
        ws.title =  'Affordable Applicants'
    else :
       ws = wb.get_sheet_by_name('Sheet4')
       ws.title =  'TWN Applicants'

    greyFill = PatternFill(fill_type='solid',start_color='A9A9A9',end_color='A9A9A9') 
    side  = Side(border_style='thin',  color="FF000000")
    border  = Border(left=side, right=side, top=side, bottom=side)
   # side_thin  = Side(border_style='thin',  color="FF000000")

    for x in range(1,len(df_detail.columns)+1 ):  #format title
            ws.cell(1, x).fill =  greyFill
           # ws.cell(1, len(df_detail)+2).alignment = Alignment( horizontal='center', vertical='center')  

    for index, row in df_detail.iterrows() :  #align
                 
        for col_num  in range(1, len(df_detail.columns)+1):
            ws.cell(row=index+2, column=col_num).alignment = Alignment( horizontal='center', vertical='center')   
            ws.cell(row=index+2, column=col_num).border = border

 
    ws.auto_filter.ref = 'A1:' + column_num_to_string (len(df_detail.columns )) +'1'
    ws.freeze_panes ='A2'


def write_book_1( df_new,  wb, sheettype = 'ssv'):
   
    if (sheettype =='ssv'):
        ws = wb.get_sheet_by_name('Sheet1')
        ws.title =  '2025 SSV MIL'
        redFill =PatternFill(fill_type='solid',start_color='FF8C00',end_color='FF8C00') 
        greenFill =PatternFill(fill_type='solid',start_color='9ACD32',end_color='9ACD32') 
        greyFill = PatternFill(fill_type='solid',start_color='A9A9A9',end_color='A9A9A9') 

    else:
        ws = wb.get_sheet_by_name('Sheet3')
        ws.title =  '2025 TWN'
        redFill =PatternFill(fill_type='solid',start_color='EFC3CA',end_color='EFC3CA') 
        greenFill =PatternFill(fill_type='solid',start_color='5DE2E7',end_color='5DE2E7') 
        greyFill = PatternFill(fill_type='solid',start_color='A9A9A9',end_color='A9A9A9') 
   

    side  = Side(border_style='thin',  color="FF000000")
    border  = Border(left=side, right=side, top=side, bottom=side)
    side_thin  = Side(border_style='thin',  color="FF000000")
    
    top_bottom_border = Border(
    top=Side(border_style='thick', color='00000000'),
    bottom=Side(border_style='thick', color='00000000'))

    thick_right_border = Border(
    left=Side(border_style='none', color='00000000'),
    right=Side(border_style='thick', color='00000000'),
    top=Side(border_style='none', color='00000000'),
    bottom=Side(border_style='none', color='00000000'))

    pert_top_right_border = Border(
    left=Side(border_style='none', color='00000000'),
    right=Side(border_style='thick', color='00000000'),
    top=Side(border_style='thick', color='00000000'),
    bottom=Side(border_style='thick', color='00000000'))

 

    color = greyFill 

    for x in range(1,len(df_new.columns)+4):  #format title
        
        ws.cell(1,x).fill = color  #second line fill
        ws.cell(len(df_new)+1,x).fill = color #last laine fill
        y = x 
        if (y%3==0  ):
            ws.cell(1, y).border = thick_right_border

        if (x < len(df_new.columns) and x >=3):
            if (x%3==0  ): 
                if (color== redFill):
                    color = greenFill
                    
                else:
                    color=redFill
        else:
                color = greyFill

        ws.cell(1, x).alignment = Alignment(wrapText=True, horizontal='center', vertical='center')   
        #ws.cell(1,column=x).alignment = Alignment(horizontal='center', vertical='center')  
     

 
    df_new.reset_index(inplace = True,drop = True)
     
        
    for index, row in df_new.iterrows() :  #total horizontal
        sum_range = ""
        sum_range_2=""
        sum_range_grand=""
        sum_range_2_grand=""
        
        for col_num  in range(4, len(df_new.columns),3):
               sum_range = sum_range + column_num_to_string(col_num ) + str(index+2)  + "+"
               sum_range_2 = sum_range_2 + column_num_to_string(col_num+1 ) + str(index+2)  + "+"
               if (col_num ==4):
                ws.cell(index+2,  3).border = thick_right_border
                 

        ws.cell(row=index+2, column=len(df_new.columns)+1).value = '=sum(' +  sum_range[0:-1] + ")"
        ws.cell(row=index+2, column=len(df_new.columns)+2).value = '=sum(' +  sum_range_2[0:-1] + ")"
        ws.cell(row=index+2, column=len(df_new.columns)+3).value = '=IFERROR(' +  column_num_to_string(len(df_new.columns) +2 ) + str(index+2) + '/' + column_num_to_string(len(df_new.columns) +1  )  + str(index+2) + ',"-")'
          
        ws.cell(index+2, len(df_new.columns)+3).border = thick_right_border 
        #ws.cell(index+3, len(df_new.columns)+3).border = thick_right_border 

 
    ws.cell(1,len(df_new.columns)+1).value = 'Total' + chr(13) + chr(10) + 'Processing' 
    ws.cell(1,len(df_new.columns)+2).value = 'Total' + chr(13) + chr(10) + 'Found' 
    ws.cell(1,len(df_new.columns)+3).value = 'Total' + chr(13) + chr(10) + 'Percent' 
    ws.cell(1,len(df_new.columns)+1).font = Font(bold=True )  
    ws.cell(1,len(df_new.columns)+2).font = Font(bold=True )
    ws.cell(1,len(df_new.columns)+3).font = Font(bold=True )   
    ws.cell(1,len(df_new.columns)+3).border = thick_right_border 

    for col_num  in range(1, len(df_new.columns)+4): #column total vertical direction
         
        if (col_num >=3):
            ws.cell(row=len(df_new)+1,column=col_num).value = '= sum(' + column_num_to_string(col_num ) + '2:' + column_num_to_string(col_num )  + str(len(df_new))  + ')'
           #ws.cell(row=len(df_new)+1,column=col_num).alignment = Alignment(horizontal='center', vertical= 'center')  
        ws.cell(len(df_new)+1,  col_num ).border  = top_bottom_border
        ws.cell(len(df_new)+1,  col_num ).font = Font(bold=True )  
 

    for index, row in df_new.iterrows(): #column total vertical direction format percentage
        
         for col_num  in range(3, len(df_new.columns)+4,3):
           # ws.cell(index+2, col_num).border = thick_right_border    
            if (col_num>3) :
                ws.cell(row=index+2, column=col_num).value = '=IFERROR(' +  column_num_to_string(col_num -1 ) + str(index+2) + '/' + column_num_to_string(col_num-2 )  + str(index+2) + ',"-")'
                ws.cell(row=index+2, column=col_num).number_format = '0%'  #normal column
                ws.cell(index+2, col_num).border = thick_right_border
            if (col_num>=3):
                ws.cell(len(df_new)+1,  col_num ).border  = pert_top_right_border
            

    ws.cell(row=len(df_new)+1,column=2).value = 'Total' 
    ws.cell(row=1,column=1).value = 'Property Code' 
    
    for index, row in df_new.iterrows() :  #align
                 
        for col_num  in range(4, len(df_new.columns)+4):
            ws.cell(row=index+2, column=col_num).alignment = Alignment( horizontal='center', vertical='center')   


    ws.auto_filter.ref = 'A1:' + column_num_to_string (len(df_new.columns)) +'1'
    ws.freeze_panes ='A2'
 

def get_ssv_report ( writer2) :
    df_prop = connect_ssv_prop()
    df_row = pd.DataFrame()
    Manual = 0
    for value in df_prop['manager2'].dropna().unique():
        value = str(value)
        df_row =df_prop[df_prop['manager2'].str.contains(value, na=False)]  
        if (len(value)<=2):
            continue
        df_row=df_row.drop(['manager2'], axis = 1)
        if (Manual ==0):
            #start_date = date(2024, 1,1)
            start_date = date(2025, 1,1)
            end_date = date.today()
    #end_date = date(2020, 9, 1)
            # = timedelta(days=1)
        else:
            last_day_month = date.today().replace(day=1) - timedelta(days=1)
            start_date = date.today().replace(day=1) - timedelta(days=last_day_month.day)
            end_date =  (start_date.replace(day=1) + timedelta(days=31)).replace(day=1) + timedelta(days=0)
            r_date = relativedelta(months=12) 
            start_date = end_date - r_date
            end_date = last_day_month
            delta = timedelta(days=1)
    
    
    
     
    #end_date = datetime.strptime(end_date.get(), '%m/%d/%Y').date()
        #df_detail = connect_apps(start_date, date.today().replace(day=1) - timedelta(days=0)) 
        df_detail = connect_ssv_apps(start_date, end_date) 
      

        while start_date < end_date:
    # print (start_date.strftime("%Y-%m-%d"))
            e_date = (start_date.replace(day=1) + timedelta(days=31)).replace(day=1) + timedelta(days=-1)
 
            e_date = (start_date.replace(day=1) + timedelta(days=31)).replace(day=1) + timedelta(days=0)

            e_date = date.today()
   
            df_app=connect_ssv_App(start_date, e_date )
            df_row = pd.merge(df_row, df_app ,on = ["node_code"  ], how='left')
            #start_date =(start_date.replace(day=1) + timedelta(days=31)).replace(day=1)
            #start_date = start_date + timedelta(days=8) 
            break

        df_row = (df_row._append(pd.Series(), ignore_index = True))
       
        df_row.to_excel(writer2, index = False, header=True,   sheet_name = 'Sheet1' ) #float_format='%.00f',header=True,
       
       
        df_detail.to_excel(writer2, index = False,  header=True,   sheet_name = 'Sheet2' ) #float_format='%.00f',header=True,
       
        worksheet= writer2.sheets['Sheet2'] 
        for i, col in enumerate(df_detail.columns):
            column_len = max(df_detail[col].astype(str).str.len().max(), len(col) + 2) + 10
            worksheet.set_column(i, i, column_len)
         
        
        return df_row, df_detail 
    
      
    
def get_twn_report (  writer2) :
    df_prop = connect_twn_prop()
    df_row = pd.DataFrame()
    Manual = 0
    for value in df_prop['manager2'].dropna().unique():
        value = str(value)
        df_row =df_prop[df_prop['manager2'].str.contains(value, na=False)]  
        if (len(value)<=2):
            continue
        df_row=df_row.drop(['manager2'], axis = 1)
        if (Manual ==0):
            #start_date = date(2024, 1,1)
            start_date = date(2025, 1,1)
            end_date = date.today()
    #end_date = date(2020, 9, 1)
            # = timedelta(days=1)
        else:
            last_day_month = date.today().replace(day=1) - timedelta(days=1)
            start_date = date.today().replace(day=1) - timedelta(days=last_day_month.day)
            end_date =  (start_date.replace(day=1) + timedelta(days=31)).replace(day=1) + timedelta(days=0)
            r_date = relativedelta(months=12) 
            start_date = end_date - r_date
            end_date = last_day_month
            delta = timedelta(days=1)
    
    
    
     
    #end_date = datetime.strptime(end_date.get(), '%m/%d/%Y').date()
        #df_detail = connect_apps(start_date, date.today().replace(day=1) - timedelta(days=0)) 
        df_detail = connect_twn_apps(start_date, end_date) 
      

        while start_date < end_date:
    # print (start_date.strftime("%Y-%m-%d"))
            e_date = (start_date.replace(day=1) + timedelta(days=31)).replace(day=1) + timedelta(days=-1)
 
            e_date = (start_date.replace(day=1) + timedelta(days=31)).replace(day=1) + timedelta(days=0)

            e_date = date.today()
   
            df_app=connect_twn_App(start_date, e_date )
            df_row = pd.merge(df_row, df_app ,on = ["node_code"  ], how='left')
            #start_date =(start_date.replace(day=1) + timedelta(days=31)).replace(day=1)
            #start_date = start_date + timedelta(days=8) 
            break

        df_row = (df_row._append(pd.Series(), ignore_index = True))
        
        df_row.to_excel(writer2, index = False, header=True,   sheet_name = 'Sheet3' ) #float_format='%.00f',header=True,
       
       
        df_detail.to_excel(writer2, index = False,  header=True,   sheet_name = 'Sheet4' ) #float_format='%.00f',header=True,
       
        worksheet= writer2.sheets['Sheet4'] 
        for i, col in enumerate(df_detail.columns):
            column_len = max(df_detail[col].astype(str).str.len().max(), len(col) + 2) + 10
            worksheet.set_column(i, i, column_len)
        
        return df_row, df_detail 
   
        

if __name__ == '__main__':
     
       
        filename1 =  "O:\\ANALYTICS\\Client_Analysis\\MIL\\SSV"  +  "\\" + 'Highmark Affordable_TWN  - ' + date.today().strftime("%m-%d-%Y")  + '.xlsx'
        filename1 = os.getcwd() +  "\\" + 'Highmark Affordable_TWN  - ' + date.today().strftime("%m-%d-%Y")  + '.xlsx'
        writer2 = pd.ExcelWriter(filename1 )
        df_row_ssv, df_detail_ssv  = get_ssv_report ( writer2  )
   
        df_row_twn, df_detail_twn  = get_twn_report ( writer2  )

        writer2.close() 

        wb = load_workbook(filename1) 
        
   
        write_book_1( df_row_ssv,   wb, 'ssv')
        write_book_2( df_detail_ssv,  wb, 'ssv')
        
        write_book_1( df_row_twn,   wb, 'twn')
        write_book_2( df_detail_twn,   wb, 'twn')
         

        wb.save(filename1)
        

sys.exit(0)
quit()
