 
from tkinter import FIRST
import psycopg2
 
import datetime
from datetime import date, timedelta,datetime
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.cell import Cell  
import pandas as pd
import numpy as np
import os
import sys
import logging
import locale
from openpyxl.styles import Alignment
import dateutil.relativedelta
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Color, Fill, Font, Border,Side, PatternFill
from openpyxl.styles import colors
import openpyxl 
from openpyxl import Workbook
 
from win32com import client
 
import pathlib
from functools import reduce
import win32api
from shutil import copyfile
from config import config
from config2 import config2
import dateutil.relativedelta 
 
from dateutil.relativedelta import relativedelta
from win32com.client import Dispatch
import win32com.client as win32
import codecs
from pyexcelerate import Workbook, Color, Style, Fill
import pyexcelerate
from openpyxl.styles import numbers
from openpyxl.styles import NamedStyle
#file_template = dirpath() +    '\\setup_audit_template.xlsx' 
import re 

dirpath = os.getcwd()
print (dirpath)
d =  datetime.today()  
end_date = d.strftime("%x")
 
last_day_month = date.today().replace(day=1) - timedelta(days=1)
start_day_month = date.today().replace(day=1) - timedelta(days=last_day_month.day)

start_day__qt =start_day_month - dateutil.relativedelta.relativedelta(months=2)



today =  date.today()
idx =  (today.weekday() + 1) % 7
w_start_date = sys.argv[1] if len(sys.argv) > 1 else today -  timedelta(7+idx)
w_end_date =  sys.argv[2] if len(sys.argv) > 2 else w_start_date + timedelta(days=6)

w_start_date = w_start_date.strftime("%m/%d/%Y")  
w_end_date=w_end_date.strftime("%m/%d/%Y")  

#filename1 =   dirpath +  'CAPIL Screening - ' + str(start_date) + '-' + str(end_date) + '.xlsx'
#file_temp =  dirpath +  '\\template_week.xls'


filename = w_start_date + "-" + w_end_date
filename1 =  "\\ysifwfs07\\Vol2\ANALYTICS\\New Property Lists\\new_property_list " + filename.replace("/", ".") + ".xlsx"
filename1 = "O:\\ANALYTICS\\New Property Lists\\new_property_list "  +   filename.replace("/", ".") + ".xlsx"
#MAXCOL=18

redFill =PatternFill(fill_type='solid',start_color='5B9BD5',end_color='5B9BD5') 
side  = Side(border_style='thin',  color="FF000000")
border  = Border(left=side, right=side, top=side, bottom=side)
side_thin  = Side(border_style='thin',  color="FF000000")
    
    
thin_border = Border(
    left=Side(border_style='thin', color='00000000'),
    right=Side(border_style='thin', color='00000000'),
    top=Side(border_style='thin', color='00000000'),
    bottom=Side(border_style='thin', color='00000000'))

thick_border = Border(
    top=Side(border_style='thick', color='00000000'))


def generate_terminate_code ( ):
    conn = None
    params = config2()
    
    conn = psycopg2.connect(**params )
 
    cur = conn.cursor()

    SQL = ("""   select company_code  from "Test_Code_Master"  ; """   )
        
    cur.execute(SQL) 
    #col = cur.description
    col = [i[0] for i in cur.description]
    a = np.array(col)
    rows =   [item[0] for item in cur.fetchall()]
            
    #df_codes = pd.DataFrame(rows, columns = a)
    cur.close()
    conn.close()
    return rows

list_term_code =  generate_terminate_code () 


def dispatch(app_name:str):
    try:
        from win32com import client
        app = client.gencache.EnsureDispatch(app_name)
    except AttributeError:
        # Corner case dependencies.
        import os
        import re
        import sys
        import shutil
        # Remove cache and try again.
        MODULE_LIST = [m.__name__ for m in sys.modules.values()]
        for module in MODULE_LIST:
            if re.match(r'win32com\.gen_py\..+', module):
                del sys.modules[module]
        shutil.rmtree(os.path.join(os.environ.get('LOCALAPPDATA'), 'Temp', 'gen_py'))
        from win32com import client
        app = client.gencache.EnsureDispatch(app_name)
    return app         
 


#write exlude test codes; replace with abreviation
def save_workbook (s_name):
    excel = dispatch('Excel.Application')
    excel.Interactive = False
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = excel.Workbooks.Open(s_name)
     
   # getNumSheet = wb.Worksheets.count+1

    #for i in range  (1, getNumSheet ):
      #  excel.Worksheets(i).Activate() 

    wb.Worksheets('New Property').Select()    
    excel.ActiveSheet.Columns.AutoFit()


    wb.Worksheets('Duplicate Property').Select()    
    excel.ActiveSheet.Columns.AutoFit()

    wb.Worksheets('Breeze Pricing').Select()    
    excel.ActiveSheet.Columns.AutoFit()
        
#Save changes in a new file
    excel.DisplayAlerts = False    
    wb.SaveAs (s_name)
    wb.Close()
 

def get_streettype():
    onn = None

    params = config()
    conn = psycopg2.connect(**params)       
    cur = conn.cursor()
    SQL = (""" select typeabbr, typename  from streettypes """ ) 
    cur.execute(SQL) 
    col = [i[0] for i in cur.description]
    conn.set_client_encoding('ISO-8859-1') 
    rows = cur.fetchall()
    cur.close()
    street_data = pd.DataFrame(rows, columns = col)
    street_data = street_data.map(lambda x: x.lower() if pd.notnull(x) else x)
    conn.close()
    return street_data

def connect_acct_manager():
    params = config2()
    conn = psycopg2.connect(**params)       
    cur = conn.cursor() 
    SQL = ("""  
     select company_code,  max(rs_account_manager)  account_manager from   ycrm_company where coalesce(rs_account_manager,'') <> '' and coalesce(company_code,'')<>''
group by company_code     
   
""" ) 
    cur.execute(SQL) 
    col = [i[0] for i in cur.description]
    conn.set_client_encoding('ISO-8859-1') 
    rows = cur.fetchall()
    acct_am = pd.DataFrame(rows, columns = col)
    cur.close()
    return acct_am

##### weekly

def street_sql ():
    SQL =    ( """ SET client_encoding= 'SQL_ASCII'; 
       select trim(NOD.company_code) company_code, company_name, node_code, node_name, NOD.create_stamp::date create_date, '                       ' as acct_manager,  
           trim(coalesce(node_street_1,'')) node_street_1, trim(coalesce(NOD.node_street_2,'')) node_street_2, 
       trim(coalesce(NOD.node_city,'')) node_city, 
trim(NOD.node_state) node_state, trim(NOD.node_zip) node_zip ,  market_rate_units, last_applicant,  first_applicant,
           
           --street 1
concat(  trim(coalesce(node_street_1,'')),   trim(coalesce(NOD.node_city,'')), trim(coalesce(NOD.node_state,'')), left(coalesce(trim(NOD.node_zip),''),4) ) dd, 

           --- street2
concat(
           case when 
left(node_street_1,2) ~ '^(([-+]?[0-9]+(\.[0-9]+)?)|([-+]?\.[0-9]+))$'  then  node_street_1 
when 
left(node_street_2,2) ~ '^(([-+]?[0-9]+(\.[0-9]+)?)|([-+]?\.[0-9]+))$'  then  node_street_2


else trim(coalesce(node_street_1,'') || ' ' || coalesce(replace(node_street_2, node_street_1,''), '')) end ,
           
           left(trim(coalesce(NOD.node_city,'')), 3), left(coalesce(trim(NOD.node_state),''),2)
           
           ) dd_1,


           --- partial
           
concat(  left(trim(replace(coalesce(node_street_1,''), ' ',''))  , 5), 
              left(trim(replace(coalesce( coalesce(replace(node_street_2, node_street_1,''), ''),'') , ' ','') )  , 3), 
           left(trim(coalesce(NOD.node_city,'')), 3),
 left(trim(coalesce(node_zip,'')),3),  (market_rate_units )) dd_2
              
     --         concat(  left(trim(replace(coalesce(node_street_1,''), ' ',''))  , 5), 
          --    left(trim(replace(coalesce( coalesce(replace(node_street_2, node_street_1,''), ''),'') , ' ','') )  , 3), 
       --    left(trim(coalesce(NOD.node_city,'')), 3),
 --left(trim(coalesce(node_zip,'')),3)) dd_2
 
from node_tbl NOD inner join company_tbl on  company_tbl.company_code = NOD.company_code 
              left join breeze_url on breeze_url.company_code = NOD.company_code 
where  market_rate_units>0 and canceled = false and NOD.company_code not in ('APP')
  
and node_code not  in  (select node_code from node_suspension_details where suspend_id in (select max(suspend_id) from node_suspension_details group by node_code )
   
 and suspended = true )
 and   ( coalesce(breeze_package,'') = '' or  coalesce(breeze_package,'') = 'BreezePremier') 
and coalesce(first_applicant::date,'1/1/1900') > '1/1/1900'             
  

"""  )
    return SQL


def connect_rentgrow_data_frame(cus_sql):
     
    conn = None
 
    params = config()
    conn = psycopg2.connect(**params)       
    cur = conn.cursor()
   
    
    string_replace = "  ('" + "','".join(name.upper().strip().replace("'", r"\'") for name in list_term_code) + "'"  + ')'
   # SQL_1 =  (o_SQL).replace( regexp=True, to = "request.company_code not in ('APP') ", value = string_replace    )
    cus_sql =  (cus_sql).replace(   "('APP')",   string_replace     ) #live
    
    SQL =  cus_sql
    cur.execute(SQL) 
    col = [i[0] for i in cur.description]
    conn.set_client_encoding('ISO-8859-1') 
    rows = cur.fetchall()
    cur.close()
    data = pd.DataFrame(rows, columns = col)
   
    return data


def connect_breeze_premier_model(w_start_date, w_end_date):

    conn = None
 
    params = config()
    conn = psycopg2.connect(**params)       
    cur = conn.cursor()
     
    SQL = (""" SELECT CT.company_code, CT.company_name,   NT.node_code, node_name,
           case AXC.pricing_model
	WHEN '0' THEN 'NOT SET'
	WHEN '1' THEN 'Transactional - Standard'
	WHEN '2' THEN 'Transactional - Bundled'
	WHEN '3' THEN 'Unit-Monthly'
	WHEN '4' THEN 'Unit-Quarterly'
	WHEN '5' THEN 'Unit_Annual'
	WHEN '6' THEN 'Manual'
	WHEN '7' THEN 'Mixed'
	ELSE NULL
	END AS company_pricing_model,  

	case AXN.pricing_model
	WHEN '0' THEN 'NOT SET'
	WHEN '1' THEN 'Transactional - Standard'
	WHEN '2' THEN 'Transactional - Bundled'
	WHEN '3' THEN 'Unit-Monthly'
	WHEN '4' THEN 'Unit-Quarterly'
	WHEN '5' THEN 'Unit_Annual'
	WHEN '6' THEN 'Manual'
	WHEN '7' THEN 'Mixed'
	ELSE NULL
	END AS node_pricing_model ,

case when breeze_package is null then 'Standard' else  breeze_package end as "Company Type"  
FROM node_tbl AS NT
FULL JOIN company_tbl AS CT
ON NT.company_code = CT.company_code
 
FULL JOIN auxiliary_node AS AXN
ON NT.node_code = AXN.node_code
FULL JOIN auxiliary_company AS AXC
ON CT.company_code = AXC.company_code
  left join breeze_url on breeze_url.company_code = NT.company_code
 
where NT.canceled = 'f' and (   coalesce(breeze_package,'') = 'BreezePremier') and NT.company_code not in ('APP')
 and  NT.create_stamp >= '%s' and  NT.create_stamp<'%s'
 ORDER BY NT.company_code, NT.node_code, 1;

; """ % (w_start_date,  w_end_date))
    
    return SQL
     

def connect_breeze_premier_price(w_start_date, w_end_date):

    conn = None
 
    params = config()
    conn = psycopg2.connect(**params)       
    cur = conn.cursor()
     
    SQL = (""" SELECT NT.node_code,  index_stamp, transaction_code  , price
 
           
FROM node_tbl AS NT
inner JOIN current_price_plans AS CT
ON NT.node_code = CT.node_code inner join breeze_url on breeze_url.company_code = NT.company_code
  
 
where NT.canceled = 'f' and (   coalesce(breeze_package,'') = 'BreezePremier') and NT.company_code not in ('APP')
           and start_date <'%s'  and end_date > '%s'  
 and  NT.create_stamp >= '%s' and  NT.create_stamp<'%s'
 ;

; """ % (w_start_date,  w_end_date,w_start_date,  w_end_date))
    
    return SQL

def run_transaction_part( df_pricing, df_no_charge_final):
      
 
    #df_trans = connect_rentgrow_data_frame (generate_trans_sql(start_date, end_date))
    #df_invoice = connect_rentgrow_data_frame (generate_invoice_zero_sql(start_date, end_date))
    #df_no_charge =   pd.merge(df_trans, df_invoice ,on = [  "node_code"  ], how='inner')

    if (len(df_no_charge_final) < 1):
        return df_no_charge_final
 
    df_no_charge_final = df_no_charge_final.sort_values (by=[ 'node_code', 'transaction_code','index_stamp'])
    df_no_charge_final= df_no_charge_final.drop_duplicates(subset=['node_code', 'transaction_code' ], keep='last' )
    
 
    aggFunc = { 
        #   'transaction_uuid' : np.count_nonzero,
            'price' :  max
  
           }

    #'Property Name','company_code',, 'total' margins=True, margins_name='Grand Total'
    df_no_charge_final = pd.pivot_table(df_no_charge_final,index=[ "node_code" ] ,
                                     aggfunc=aggFunc, columns=[  'transaction_code'],
                                    values=["price"] ,  dropna=False, fill_value=0, ).reset_index()
    

  
     
    #df_no_charge_final.columns = df_no_charge_final.columns.astype(str).str.replace('count_tran', '').str.replace('(', '').replace(')', '')
    df_no_charge_final.columns = ['_'.join(col) for col in df_no_charge_final.columns.values]
   
    df_no_charge_final.rename(columns=lambda x: x.replace('price', 'Price'), inplace=True)
    df_no_charge_final.rename(columns=lambda x: x.replace("('',", '').replace("')", '').replace(')', ''), inplace=True)
    df_no_charge_final.rename(columns=lambda x: x.strip(), inplace=True)

 
    df_no_charge_final.replace( np.nan, '',inplace = True)
 
    
    df_no_charge_final = pd.DataFrame(df_no_charge_final.to_records())
    df_no_charge_final.rename(columns=lambda x: x.replace('node_code_', 'node_code'),  inplace=True)
    df_no_charge_final = pd.merge( df_no_charge_final,df_pricing  ,on = ['node_code'], how='left')

    df_no_charge_final = df_no_charge_final.sort_values (by=[ 'company_code', 'node_name'])
    df_no_charge_final = df_no_charge_final.reset_index(drop=True)
    df_no_charge_final = df_no_charge_final.drop('index', axis=1)

    for i, col in enumerate(df_no_charge_final.columns):
           if re.search('price', col, re.IGNORECASE):
               df_no_charge_final[col] = df_no_charge_final[col].apply(lambda x: float(x))

    col = df_no_charge_final.pop ('company_code')
    df_no_charge_final.insert(0, col.name, col)

    col = df_no_charge_final.pop ('company_name')
    df_no_charge_final.insert(1, col.name, col)
    #cols_at_end = ['company_code','company_name' ,'node_code',  'node_name', 'company_pricing_model','node_pricing_model','Price_DOCVERIFY','Price_NOVAINCOME','Price_RENTHISTORY','DOCVERIFY','NOVAINCOME','RENTHISTORY','Monthly Invoice$','canceled','amt']
    #df_no_charge_final = df_no_charge_final[ cols_at_end + [ col for col in df_no_charge_final.columns if col not in cols_at_end ]]
    col = df_no_charge_final.pop ('node_code')
    df_no_charge_final.insert(2, col.name, col)
   
    col = df_no_charge_final.pop ('node_name')
    df_no_charge_final.insert(3, col.name, col)

    col = df_no_charge_final.pop ('company_pricing_model')
    df_no_charge_final.insert(4, col.name, col)

    col = df_no_charge_final.pop ('node_pricing_model')
    df_no_charge_final.insert(5, col.name, col)
    return df_no_charge_final 

def connect_new_prop(w_start_date, w_end_date):
    
    street_type = get_streettype()


    conn = None
 
    params = config()
    conn = psycopg2.connect(**params)       
    cur = conn.cursor()
        
  

    SQL = (""" SELECT NT.create_stamp::TIMESTAMP WITHOUT TIME ZONE AS node_create_stamp, CT.create_stamp::TIMESTAMP WITHOUT TIME ZONE   AS company_create_stamp, CT.company_name, NT.company_code, NT.node_code, 
NT.node_name,  NT.node_street_1, NT.node_street_2, NT.node_city, NT.node_state, NT.node_zip, 

--ACC.billto_street_1, ACC.billto_street_2, 
--ACC.billto_city, ACC.billto_state, ACC.billto_zip, 

AXN.enable_einvoices, AXN.email_for_invoices, AXN.email_for_bulk_invoices,
--CASE WHEN NT.node_code = PL.node_code THEN 'Yes' ELSE 'No' END     AS node_pull_list,
--CASE WHEN lower(ACC.billto_street_1) LIKE '%%leasing%%' THEN 'Yes' ELSE 'No' END AS leasing, 
--CASE WHEN lower(AXC.remarks) LIKE '%%e-invoicing%%'  then 'Yes' when lower(AXC.remarks) LIKE '%%einvoicing%%' then 'Yes' ELSE 'No' END AS company_einvoicing,
--CASE WHEN lower(AXC.remarks) LIKE '%%pull list%%' THEN 'Yes' ELSE 'No' END AS company_pull_list,
 CASE WHEN lower(AXC.remarks) LIKE '%%invoice summary%%'  then 'Yes' else 'No'  END AS Invoice_Summary,
 CASE WHEN lower(AXC.remarks) LIKE '%%party billing%%' and  ( lower(AXC.remarks) LIKE '%%third%%' or lower(AXC.remarks) LIKE '%%3rd%%' ) THEN 'Yes' 
      WHEN trim(lower(AXN.email_for_invoices)) like '%%@rcash.com%%' then 'Yes'
      WHEN trim(lower(AXN.email_for_invoices)) like '%%@avidbill.com%%' then 'Yes'
      WHEN trim(lower(AXN.email_for_invoices)) like '%%@yardifs.com%%' then 'Yes'

      WHEN trim(lower(AXN.email_for_bulk_invoices)) like '%%@rcash.com%%' then 'Yes'
      WHEN trim(lower(AXN.email_for_bulk_invoices)) like '%%@avidbill.com%%' then 'Yes'
      WHEN trim(lower(AXN.email_for_bulk_invoices)) like '%%@yardifs.com%%' then 'Yes'
      WHEN trim(lower(AXN.email_for_bulk_invoices)) like '%%@@nexussystems.com%%' then 'Yes'
      WHEN trim(lower(AXN.email_for_bulk_invoices)) like '%%nexusservices%%@%%' then 'Yes'
ELSE 'No' END AS Third_Party_Billing,

CASE WHEN replace(lower(AXC.remarks),' ','') LIKE '%%vendorca%%'   THEN 'Yes' ELSE 'No' END AS "Vendor Café",
CASE WHEN replace(lower(AXC.remarks),' ','') LIKE '%%paysca%%'  THEN 'Yes' ELSE 'No' END AS "PayScan",
case when breeze_package is null then 'Standard' else  breeze_package end as "Company Type", username 
FROM node_tbl AS NT
FULL JOIN company_tbl AS CT
ON NT.company_code = CT.company_code
FULL JOIN account_tbl as ACC
ON NT.node_code = ACC.account_code
FULL JOIN auxiliary_node AS AXN
ON NT.node_code = AXN.node_code
FULL JOIN auxiliary_company AS AXC
ON CT.company_code = AXC.company_code
FULL JOIN pull_lst_tbl as PL
ON NT.node_code = PL.node_code left join breeze_url on breeze_url.company_code = NT.company_code
  left join (select   max(username) username, who  from audit_tbl   where    trim(type) = 'Add' and coalesce(username,'')<>'' group by  who ) audit 
              on who = NT.node_code 
where NT.canceled = 'f' and ( coalesce(breeze_package,'') = '' or  coalesce(breeze_package,'') = 'BreezePremier') and NT.company_code not in ('APP')
 and  NT.create_stamp >= '%s' and  NT.create_stamp<'%s'
 ORDER BY NT.company_code, NT.node_code, 1;

; """ % (w_start_date,  w_end_date))
    
 
    list_term_code =  generate_terminate_code () 

    string_replace = "  ('" + "','".join(name.upper().strip().replace("'", r"\'") for name in list_term_code) + "'"  + ')'
    SQL = SQL.replace(   "('APP')",   string_replace     )


    cur.execute(SQL) 
    col = [i[0] for i in cur.description]
    conn.set_client_encoding('ISO-8859-1') 
    rows = cur.fetchall()
    cur.close()
    data = pd.DataFrame(rows, columns = col)

   # Row_list =[] 
  
# Iterate over each row 
   # for rows in data.itertuples(): 
    # Create list for the current row 
   #     my_list =[rows.dd, rows.dd_1, rows.dd_2] 
      
    # append the list to the final list 
   # Row_list.append(my_list) 
    return data
# Print the list 


def connect_weekly_dup(w_start_date, w_end_date):
    
    street_type = get_streettype()


    conn = None
 
    params = config()
    conn = psycopg2.connect(**params)       
    cur = conn.cursor()
        
    #SET client_encoding= 'SQL_ASCII'; 

    SQL = street_sql () +  ( """   
 and NOD.create_stamp >= '%s' and NOD.create_stamp<'%s'
 and  NOD.company_code not in ('APP')
; """ % (w_start_date,  w_end_date))
    
    #SQL = street_sql()

    list_term_code =  generate_terminate_code () 

    string_replace = "  ('" + "','".join(name.upper().strip().replace("'", r"\'") for name in list_term_code) + "'"  + ')'
    SQL = SQL.replace(   "('APP')",   string_replace     )


    cur.execute(SQL) 
    col = [i[0] for i in cur.description]
    conn.set_client_encoding('ISO-8859-1') 
    rows = cur.fetchall()
    cur.close()
    data = pd.DataFrame(rows, columns = col)

   # Row_list =[] 
  
# Iterate over each row 
   # for rows in data.itertuples(): 
    # Create list for the current row 
   #     my_list =[rows.dd, rows.dd_1, rows.dd_2] 
      
    # append the list to the final list 
   # Row_list.append(my_list) 
    return data
# Print the list 
 

 

def connect_a(df_weekly_data):
     
    street_type = get_streettype()
 
    conn = None
 
    params = config()
    conn = psycopg2.connect(**params)       
    cur = conn.cursor()
        
  
    #SET client_encoding= 'SQL_ASCII'; 
    SQL =   street_sql()
    
 
    list_term_code =  generate_terminate_code () 

    string_replace = "  ('" + "','".join(name.upper().strip().replace("'", r"\'") for name in list_term_code) + "'"  + ')'
    SQL = SQL.replace(   "('APP')",   string_replace     )


    cur.execute(SQL) 
    col = [i[0] for i in cur.description]
    conn.set_client_encoding('ISO-8859-1') 
    rows = cur.fetchall()
    cur.close()
    data = pd.DataFrame(rows, columns = col)
 
    data=data[(data['dd'].isin(df_weekly_data['dd'].values.tolist()))
          | (data['dd_1'].isin(df_weekly_data['dd_1'].values.tolist())) | (data['dd_2'].isin(df_weekly_data['dd_2'].values.tolist())) ] 
  
    
                        
    data['acct_manager']=data['acct_manager'].str.strip()

    list_am = connect_acct_manager()
    data['acct_manager']=  data['company_code'].map(list_am.set_index('company_code')['account_manager']).fillna(data['acct_manager'])



    data['dd']=data['dd'].str.lower() #.replace(' ','',  regex=True, inplace=True)
    data['dd_1']=data['dd_1'].str.lower()#.replace(' ','', regex=True, inplace=True)
    data['dd_2']=data['dd_2'].str.lower()#.replace(' ','', regex=True, inplace=True)
    
    data['dd']=data['dd'].replace({'.': '', '-':'', '_':'',' ':'',',':''})
    data['dd_1']=data['dd_1'].replace({'.': '', '-':'', '_':'',' ':'',',':''})
    data['dd_2']=data['dd_2'].replace({'.': '', '-':'', '_':'',' ':'',',':''})
 
    data['node_street_1_1']=data['node_street_1'].replace({'.': '', '-':'', '_':'',' ':'',',':''}).str.lower()
    data['node_street_2_1']=data['node_street_2'].replace({'.': '', '-':'', '_':'',' ':'',',':''}).str.lower() 
     
    for index, row in street_type.iterrows():

        data['dd'].replace( str(row[1]), str(row[0]), regex=True, inplace=True )
        data['dd_1'].replace(  str(row[1]),  str(row[0]), regex=True, inplace=True )
        data['dd_2'].replace( str(row[1]),  str(row[0]), regex=True, inplace=True )

        data ['node_street_1_1'].replace( str(row[1]),  str(row[0]), regex=True, inplace=True )
        data ['node_street_2_1'].replace( str(row[1]),  str(row[0]), regex=True, inplace=True )
    
    data['dd'] = [str(x).strip().replace(' ', '').replace('.', '')  for x in data.dd]
    data['dd']= [str(x).strip().replace('xxx-', '').replace('.', '')  for x in data.dd]
    data['dd']= [str(x).strip().replace('xxyy-', '').replace('.', '')  for x in data.dd]

    data['dd_1'] = [str(x).strip().replace(' ', '').replace('.', '')  for x in data.dd_1]   
    data['dd_1']= [str(x).strip().replace('xxx-', '').replace('.', '')  for x in data.dd_1]
    data['dd_1']= [str(x).strip().replace('xxyy-', '').replace('.', '')  for x in data.dd_1]


    data['dd_2'] = [str(x).strip().replace(' ', '').replace('.', '')  for x in data.dd_2]   
    data['dd_2']= [str(x).strip().replace('xxx-', '').replace('.', '')  for x in data.dd_2]
    data['dd_2']= [str(x).strip().replace('xxyy-', '').replace('.', '')  for x in data.dd_2]


    data['node_street_1_1'] = [str(x).strip().replace(' ', '').replace('.', '')  for x in data.node_street_1_1]  
    data['node_street_1_1']= [str(x).strip().replace('xxx-', '').replace('.', '')  for x in data.node_street_1_1]
    data['node_street_1_1']= [str(x).strip().replace('xxyy-', '').replace('.', '')  for x in data.node_street_1_1]


    data['node_street_2_1'] = [str(x).strip().replace(' ', '').replace('.', '')  for x in data.node_street_2_1]  
    data['node_street_2_1']= [str(x).strip().replace('xxx-', '').replace('.', '')  for x in data.node_street_2_1]
    data['node_street_2_1']= [str(x).strip().replace('xxyy-', '').replace('.', '')  for x in data.node_street_2_1]
 


    df_dd_new=data[data.duplicated(subset=['dd' ],keep=False )]
    #df_dd_new = df_dd.drop_duplicates (subset=['company_code', 'dd', ], keep = False)
 
    
    #df_dd_new =df_dd[df_dd.duplicated(subset=['dd',],keep=False )]
    df_dd_new ['type'] = '1-str-zp-' + df_dd_new ['dd']


    df_dd_1_new=data[data.duplicated(subset=['dd_1' ],keep=False )]
 
    df_dd_1_new ['type'] = '2-str,cy-' + df_dd_1_new ['dd_1']
 

    df_dd_2_new=data[data.duplicated(subset=['dd_2' ],keep=False )]
 
    df_dd_2_new ['type'] = '3-partial-' +  df_dd_2_new ['dd_2']
    
     
   
    #df = pd.concat([df_dd_2_new, df_dd_1_new, df_dd_new  ]).drop_duplicates('node_code', keep=FIRST)

    df = pd.concat([ df_dd_new ,df_dd_1_new, df_dd_2_new ]).drop_duplicates('node_code', keep=FIRST)

    df ['max_code']=   df.groupby('type')['company_code'].transform('max')
    df ['min_code']=   df.groupby('type')['company_code'].transform('min')


    df = df[df['max_code']!=df['min_code']]

   
    df.replace(np.nan, '',   regex=True, inplace=True)
  
 
    
    df = df.drop (['dd', 'dd_1', 'dd_2', 'node_street_1_1', 'node_street_2_1', 'max_code', 'min_code'] , axis = 1)
 
    df = df.sort_values (by= ['type', 'company_code', 'market_rate_units'] )
 

    df = df.drop_duplicates('node_code', keep=FIRST)
 
    

    conn.close()

    #new_df_aging =  connect_aging()
    linked_nodes = connect_linked_nodes()
     
    #new_df_aging.rename(columns={'Remote Acct #':'node_code'},   inplace = True)

    
    
    #df_merge_fee = df.merge(new_df_aging ,on = ["node_code"], how='left').merge(linked_nodes, on = ["node_code"], how='left') 
    df_merge_fee =df.merge(linked_nodes, on = ["node_code"], how='left')   
     
    
    return df_merge_fee


def connect_linked_nodes( ):
    conn = None
    params = config()
    
    conn = psycopg2.connect(**params )
   
    cur = conn.cursor()
 
    SQL1 = ("""    select primary_node, array_agg(link_node)::text link_node from b2b_node  where active<>'N' and primary_node <> link_node  group by primary_node
  """     )
   # print ('exe 2')
    cur.execute(SQL1) 
    col = cur.description 
            
    rows = cur.fetchall()
    df_monthly_fee_unit = pd.DataFrame()
    df_monthly_fee_unit = pd.DataFrame(rows, columns=["node_code", "link_node" ] )
 

    cur.close()
    return df_monthly_fee_unit 



def write_dup_book(header, df_new,  ws):

    if (len(df_new)==0):
         ws.cell(2,1,  'No Data')  
         return 
          
    for x in range(1,len(df_new.columns)+1):
         
        ws.cell(1, x).fill = redFill     
        ws.cell(1, x).font = Font(color="FFFFFF", name="Verdana", size=12)   
        ws.cell(1, x).alignment = Alignment(horizontal='left', vertical = 'center')
     
    o_companycode=n_code=''
    df_new.reset_index(inplace = True,drop = True)

    for index, row in df_new.iterrows():
      
            if (o_companycode  !=  ( str(row[14]  )  )):    #11 and n_code  !=  ( str(row[0])  )
              for col_num  in range(1, len(df_new.columns)+1):
                   ws.cell(index+2, col_num).border =  thick_border
            #else:
                   #for col_num  in range(1, MAXCOL-1):
                     #ws.cell(index+2, col_num).border = thin_border
                     #pass
            ws.cell(index+2, col_num).alignment = Alignment(horizontal='left', vertical = 'center')         
            o_companycode = (str(row[14])   ) #[0:6] #11
          
  
    for col_num  in range(1, len(df_new.columns)+1):
                   ws.cell(index+3, col_num).border = thick_border
                   ws.cell(index+3, col_num).alignment = Alignment(horizontal='left', vertical = 'center')
    
    ws.auto_filter.ref = 'A1:X1'
    ws.freeze_panes ='A2'




def write_new_book(header, df_new,   ws):
   
 
    for x in range(1,len(df_new.columns)+1):
         
        ws.cell(1, x).fill = redFill     
        ws.cell(1, x).font = Font(color="FFFFFF", name="Verdana", size=12)   
        ws.cell(1, x).alignment = Alignment(horizontal='left', vertical = 'center')
  

    for index, row in df_new.iterrows():
          for col_num  in range(1, len(df_new.columns)+1):
                   ws.cell(index+2, col_num).border = thin_border
                   ws.cell(index+2, col_num).alignment = Alignment(horizontal='left', vertical = 'center')

  #  ws.cell(index+3, col_num).border = thin_border
   # ws.cell(index+3, col_num).alignment = Alignment(horizontal='left', vertical = 'center')
    
    ws.auto_filter.ref = 'A1:T1'
    ws.freeze_panes ='A2'


def columnToLetter(column):
    letter = ''
    while column > 0:
        temp = (column - 1) % 26
        letter = chr(temp + 65) + letter
        column = (column - temp - 1) // 26
    return letter
 
def letterToColumn(letter):
    column = 0
    length = len(letter)
    for i in range(length):
        column += (ord(letter[i].upper()) - 64) * 26**(length - i - 1)
    return column

def  format_pricing_book( df_no_charge_final,  ws ):



    if (len(df_no_charge_final)==0):
         ws.cell(2,1,  'No Data')  
         return 
    
    dollar_style = NamedStyle(name="dollar_style", number_format='$#,##0.00')
    for i, col in enumerate(df_no_charge_final.columns):
           column_letter = columnToLetter(i+1)
           if re.search('price', col, re.IGNORECASE):
               for cell in ws[column_letter]:  # Change 'B' to your target column
                  cell.style = dollar_style
          # column_len = max(df_no_charge_final[col].astype(str).str.len().max(), len(col))  
         #   ws.set_column(i, i, column_len) 
           #ws.column_dimensions[column_letter].width = column_len
    
    for x in range(1,len(df_no_charge_final.columns)+1):
         
        ws.cell(1, x).fill = redFill     
        ws.cell(1, x).font = Font(color="FFFFFF", name="Verdana", size=12)   
        ws.cell(1, x).alignment = Alignment(horizontal='left', vertical = 'center')

    thick_border = Border(
  
    top=Side(border_style='thick', color='00000000') )
   
      
    o_companycode=''
    df_no_charge_final.reset_index(inplace = True,drop = True)

    for index, row in df_no_charge_final.iterrows():
        if (o_companycode != str(row[0])  ):
            for col_num  in range(1, len(df_no_charge_final.columns)+1):
                 ws.cell(index+2, col_num).border = thick_border           
      
        ws.cell(index+2, col_num).alignment = Alignment(horizontal='left', vertical = 'center')
        
        o_companycode = str(row[0])  

    for col_num  in range(1, len(df_no_charge_final.columns)+1):
        ws.cell( ((index+3)), col_num).border = thick_border 
        ws.cell(index+3, col_num).alignment = Alignment(horizontal='left', vertical = 'center')

    ws.auto_filter.ref = 'A1:O1'
    ws.freeze_panes ='E2'
 

 
    #ColumnDimension(ws, bestFit=True)


def send_to_finance(filename1):


    if (not os.path.exists(filename1)): 
         return
    
    #signature_name="signature (Xiaobin.Zhang@Yardi.com)"
    #sig_files_path = 'AppData\Roaming\Microsoft\Signatures\\' + signature_name + '_files\\'
    #sig_html_path = 'AppData\\Roaming\\Microsoft\\Signatures\\' + signature_name + '.htm'
    signature_file = "C:\App\Data\Sig\sig.htm"
    #signature_path = os.path.join((os.environ['USERPROFILE']), sig_files_path) # Finds the path to Outlook signature files with signature name "Work"
    #html_doc = os.path.join((os.environ['USERPROFILE']),sig_html_path)     #Specifies the name of the HTML version of the stored signature
    #html_doc = html_doc.replace('\\\\', '\\')


    html_file = codecs.open(signature_file, 'r', 'utf-8', errors='ignore') #Opens HTML file and converts to UTF-8, ignoring errors
    signature_code = html_file.read()               #Writes contents of HTML signature file to a string

    #signature_code = signature_code.replace((signature_name + '_files/'), signature_path)      #Replaces local directory with full directory path
    html_file.close()

    
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = "xiaobin.zhang@yardi.com"
    mail.Subject = "Weekly New Property List Report - Susan.Sheppard@Yardi.Com ;  Jacqui.Adler@Yardi.Com ;  Neha.Bansal@Yardi.Com "  
    filename2 =  filename1.replace('O:\\ANALYTICS\\New Property Lists\\', '\\\ysifwfs07\\Vol2\ANALYTICS\\New Property Lists\\') # + filename.replace("/", ".") + ".xlsx"
        #path  = "\"\\\\windows_Server\\golobal_directory\\the folder\\file yyymm.xlsx\""
    path = '"' + filename2 + '"'
    string = """<a href=""" +  path + ' style=text-decoration: none>' + filename1 +  '<' +  r'\a'  + '>'

      #  string.replace('\\a>', '\a>')
    #mail.body = string
     
    mail.HTMLbody =   string + " <BR><BR><BR> "  +signature_code + " <BR><BR><BR> "
    
    mail.send

def send_book( filename1, recipient, filename2=''):

    outlook = win32.Dispatch('outlook.application')
    if (os.path.exists(filename1) or os.path.exists(filename2)):
        send_account = None
        From =None

        for myEmailAddress in outlook.Session.Accounts:
            if "Xiaobin.Zhang@yardi.com" in str(myEmailAddress):
                From = myEmailAddress
                break

        
        mail = outlook.CreateItem(0)
        mail.To = recipient
        #mail.SentOnBehalfOfName = 'RS_Analytics@yardi.com'
        mail.SentOnBehalfOfName = 'xiaobin.zhang@yardi.com'
        
        if From != None:
            mail._oleobj_.Invoke(*(64209, 0, 8, 0, From))


        #mail.Cc=' YRScushwake@yardi.com; Judy.Tutt@cushwake.com; Natalie.Gatjanis@cushwake.com'
    #mail.Subject = subject
        mail.GetInspector 
        mail.Subject = "Weekly New Property List Report - Susan.Sheppard@Yardi.Com ;  Jacqui.Adler@Yardi.Com ;  Neha.Bansal@Yardi.Com ;  Odilia.Walker@Yardi.Com"  
       
        index = mail.HTMLbody.find('>', mail.HTMLbody.find('<body')) 
   # mail.HTMLbody = mail.HTMLbody[:index + 1] + message + mail.HTMLbody[index + 1:]
        if (os.path.exists(filename1)):
            #mail.Attachments.Add(filename1)
            pass
           #pass
        if (os.path.exists(filename2)):
            pass 
            #mail.Attachments.Add(filename2)    

        
        filename2 =  filename1.replace('O:\\ANALYTICS\\New Property Lists\\', '\\\ysifwfs07\\Vol2\ANALYTICS\\New Property Lists\\') # + filename.replace("/", ".") + ".xlsx"
        #path  = "\"\\\\windows_Server\\golobal_directory\\the folder\\file yyymm.xlsx\""
        path = '"' + filename2 + '"'
        string = """<a href=""" +  path + ' style=text-decoration: none>' + filename1 +  '<' +  r'\a'  + '>'

      #  string.replace('\\a>', '\a>')

       
        mail.body = filename1 
        mail.To = 'xiaobin.zhang@yardi.com'
        mail.Save()
        #mail.From = 'Xiaobin.zhang@yardi.com'
        mail.send

if __name__ == '__main__':
        

        if os.path.exists(filename1 ):
           # send_to_finance(filename1) 
           os.remove(filename1)

        if os.path.exists(filename1 ):
           # send_to_finance(filename1) 
           os.remove(filename1)
        else:
         

            df_new_prop = connect_new_prop (w_start_date, w_end_date)

            df_weekly_data = connect_weekly_dup(w_start_date, w_end_date)

            data = connect_a(df_weekly_data)

            df_prem_model  = connect_rentgrow_data_frame (connect_breeze_premier_model(w_start_date, w_end_date )) 
            df_prem_price   = connect_rentgrow_data_frame (connect_breeze_premier_price(w_start_date, w_end_date )) 
            df_price_final    = run_transaction_part( df_prem_model ,df_prem_price )
            
            writer2 = pd.ExcelWriter(filename1)
            
            df_new_prop.to_excel (writer2,  sheet_name= 'New Property', index=False, startrow=0)
            data.to_excel (writer2,  sheet_name= 'Duplicate Property', index=False, startrow=0)
            
            df_price_final.to_excel(writer2, index = False, header=True,   sheet_name = 'Breeze Pricing' ) #float_format='%.00f',header=True,

            writer2.close() 



            wb = load_workbook(filename1)

            ws = wb['Breeze Pricing'] 
            format_pricing_book (df_price_final, ws)

     
            ws = wb['Duplicate Property']
 
            write_dup_book(filename1, data,   ws)

            ws = wb['New Property']
            write_new_book(filename1, df_new_prop,  ws)
 
            wb.save(filename1)

 
            save_workbook (filename1)
        

            f = open("weekly_property_error.txt", "w") 
            error = 0
            try:
                 send_to_finance(filename1) 
                 #os.remove(filename1)
                 
        # send_book (filename1, 'xiaobin.zhang@yardi.com' )

            except Exception as Argument:

                f.write(str(Argument))
                error =1   
    
            finally:
                f.close() 

                if (error ==0 and os.path.exists(f.name)): 
                    os.remove(f.name)
     
    
sys.exit(0)
quit()
