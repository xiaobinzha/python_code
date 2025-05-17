 
#from tkinter import FIRST Spire.XLS
import psycopg2
 
import datetime
from datetime import date, timedelta,datetime
#import xlsxwriter
#from openpyxl import load_workbook
#from openpyxl.workbook import Workbook
#from openpyxl.cell import Cell  
import pandas as pd
import numpy as np
import os
import sys
import logging
import locale
 
from config import config
from config2 import config2
import dateutil.relativedelta 
 
from dateutil.relativedelta import relativedelta
#from win32com.client import Dispatch
#import win32com.client as win32
import codecs
import shutil
#from pyexcelerate import Workbook
import psutil
import win32com.client as win32
from pyexcelerate import Workbook, Color, Style, Fill
import pyexcelerate
from openpyxl.styles import Alignment
import dateutil.relativedelta
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Color, Fill, Font, Border,Side, PatternFill
from openpyxl.styles import colors
import openpyxl 
from openpyxl import Workbook
from datetime import date, timedelta
from win32com import client
import win32api
import pathlib
from functools import reduce
from xlutils.copy import copy #work xls
from xlrd import open_workbook
import xlwt
from openpyxl import load_workbook
from openpyxl.worksheet.dimensions import ColumnDimension 
from openpyxl.styles import numbers
from openpyxl.styles import NamedStyle
#file_template = dirpath() +    '\\setup_audit_template.xlsx' 
import re 
 
   
def generate_terminate_code ( ):
    conn = None
    params = config2()
    
    conn = psycopg2.connect(**params )
 
    cur = conn.cursor()

    SQL = ("""   select trim(company_code)  from "Test_Code_Master"  ; """   )
        
    cur.execute(SQL) 
    #col = cur.description
    col = [i[0] for i in cur.description]
    a = np.array(col)
    rows =   [item[0] for item in cur.fetchall()]
            
    #df_codes = pd.DataFrame(rows, columns = a)
    cur.close()
    conn.close()
    return rows

list_term_code =  generate_terminate_code () 



# no pricing and has transaction
def generate_current_price_plans_one (start_date, end_date):
    SQL =    ( """  select   
     invoice_detail.node_code,   invoice_detail.transaction_code, price, invoice_detail.amount  ,invoice_detail.transaction_uuid , total, 'match_plan_id' as match_type 
 from invoice_detail, invoice_tbl, trans_tbl, current_price_plans, node_tbl    where 

              invoice_tbl.invoice_id=invoice_detail.invoice_id and trans_tbl.transaction_uuid=invoice_detail.transaction_uuid
              and invoice_tbl.begin_service_date >= '%s' and begin_service_date < '%s'  
               and coalesce(exception_code,'')<>'dup'  and current_price_plans.index_stamp = invoice_detail.price_plan
                and current_price_plans.transaction_code = invoice_detail.transaction_code
               and total =0 and price <>0 and invoice_detail.node_code =  node_tbl.node_code and start_date <='%s' and end_date >= '%s'
              and trans_tbl.transaction_code in (
  '1STBUREAU','CRPN','CRPC','CRPS','EVICTNATPREM','RENTHISTORY','WORKNUM','NOVAINCOME','WORKNUMSSV','WORKSSV_LOOK',
'EVICTNATFILT','DOCVERIFY','VS_TWNSSVEMP','VS_ASSET_VER','VSMONTHUNIT', 'CRIMPREM', 'CRIMNATPREM', 'MONTHUNIT','VSMONTHUNIT') 
               and  node_tbl.company_code  not in ('APP')    union 
 
 select     invoice_detail.node_code,  invoice_detail.transaction_code,  price, invoice_detail.amount  ,invoice_detail.transaction_uuid , total , 'match_plan_name' as match_type 
 from trans_tbl,invoice_detail, invoice_tbl,   current_price_plans, node_tbl  
 where  invoice_detail.transaction_uuid =trans_tbl.transaction_uuid 
 and invoice_tbl.invoice_id=invoice_detail.invoice_id 
 and invoice_tbl.begin_service_date >='%s' and begin_service_date <  '%s'
  and coalesce(exception_code,'')<>'dup' and node_tbl.node_code = invoice_tbl.node_code
   and current_price_plans.transaction_code = invoice_detail.transaction_code
              and current_price_plans.node_code = invoice_detail.node_code 
  and  invoice_detail.price_plan not in  (select index_stamp from current_price_plans where 
               current_price_plans.transaction_code = invoice_detail.transaction_code
              and current_price_plans.node_code = invoice_detail.node_code and   start_date <='%s' and end_date >= '%s')
                  and total =0  and  node_tbl.company_code  not in ('APP')  and price <>0
                
                   and trans_tbl.transaction_code in (
  '1STBUREAU','CRPN','CRPC','CRPS','EVICTNATPREM','RENTHISTORY','WORKNUM','NOVAINCOME','WORKNUMSSV','WORKSSV_LOOK',
'EVICTNATFILT','DOCVERIFY','VS_TWNSSVEMP','VS_ASSET_VER','VSMONTHUNIT', 'CRIMPREM', 'CRIMNATPREM', 'MONTHUNIT','VSMONTHUNIT') 
        

; """  % (start_date, end_date,start_date, end_date,start_date, end_date, start_date, end_date))   
    return SQL

 

def generate_no_current_price_two (start_date, end_date):
    SQL =    ( """    select    trans_tbl.node_code,   transaction_code, 0 price, 0 amount  , transaction_uuid , 0 total, 'not match price' as match_type 
    --company_code, trans_tbl.node_code, trans_tbl.transaction_code, 0 price, coalesce(trans_tbl.amount,0) as amount   ,trans_tbl.transaction_uuid, 0 as total
 from  trans_tbl, node_tbl  where trans_tbl.node_code = node_tbl.node_code 
 and trans_tbl.node_code  not in  (select node_code from current_price_plans where transaction_code in (
  '1STBUREAU','CRPN','CRPC','CRPS','EVICTNATPREM','RENTHISTORY','WORKNUM','NOVAINCOME','WORKNUMSSV','WORKSSV_LOOK',
'EVICTNATFILT','DOCVERIFY','VS_TWNSSVEMP','VS_ASSET_VER','VSMONTHUNIT', 'CRIMPREM', 'CRIMNATPREM', 'MONTHUNIT','VSMONTHUNIT') 
              and   start_date <='%s' and end_date >= '%s' group by node_code  )        
 and           
                trans_tbl.transaction_date >= '%s' and transaction_date < '%s'
               and coalesce(exception_code,'')<>'dup'  
               
               and transaction_code in (
  '1STBUREAU','CRPN','CRPC','CRPS','EVICTNATPREM','RENTHISTORY','WORKNUM','NOVAINCOME','WORKNUMSSV','WORKSSV_LOOK',
'EVICTNATFILT','DOCVERIFY','VS_TWNSSVEMP','VS_ASSET_VER','VSMONTHUNIT', 'CRIMPREM', 'CRIMNATPREM', 'MONTHUNIT','VSMONTHUNIT' 
 ) and  company_code  not in ('APP')  union 
  
 select  
    trans_tbl.node_code,   transaction_code, 0 price, 0 amount  , transaction_uuid , 0 total, 'not match invoice' as match_type 
 from  trans_tbl, node_tbl  where trans_tbl.node_code = node_tbl.node_code 
     
  and           
                trans_tbl.transaction_date >=  '%s' and transaction_date < '%s'
               and coalesce(exception_code,'')<>'dup'  
               
               and transaction_code in (
  '1STBUREAU','CRPN','CRPC','CRPS','EVICTNATPREM','RENTHISTORY','WORKNUM','NOVAINCOME','WORKNUMSSV','WORKSSV_LOOK',
'EVICTNATFILT','DOCVERIFY','VS_TWNSSVEMP','VS_ASSET_VER','VSMONTHUNIT', 'CRIMPREM', 'CRIMNATPREM', 'MONTHUNIT','VSMONTHUNIT' 
 ) and  company_code  not in ('APP')
 and trans_tbl.node_code not in (  select node_code from invoice_tbl 
  where begin_service_date >='%s' and begin_service_date< '%s'  group by  node_code)   
        
 ; """  % (start_date, end_date,start_date, end_date, start_date, end_date,start_date, end_date)) 
    
    return SQL
def generate_pricing ( start_date,  end_date) :
    '''  
     start_date = date(2025, 1, 1)
   -- end_date = date(2025, 1, 31) 
    '''
    SQL =    ( """    select node_tbl.node_code, node_name, node_tbl.company_code, company_name, canceled, --amt,  
	case ac.pricing_model
	WHEN '0' THEN 'NOT SET'
	WHEN '1' THEN 'Transactional - Standard'
	WHEN '2' THEN 'Transactional - Bundled'
	WHEN '3' THEN 'Unit-Monthly'
	WHEN '4' THEN 'Unit-Quarterly'
	WHEN '5' THEN 'Unit_Annual'
	WHEN '6' THEN 'Manual'
	WHEN '7' THEN 'Mixed'
	ELSE NULL
	END AS company_pricing_model,  

	case an.pricing_model
	WHEN '0' THEN 'NOT SET'
	WHEN '1' THEN 'Transactional - Standard'
	WHEN '2' THEN 'Transactional - Bundled'
	WHEN '3' THEN 'Unit-Monthly'
	WHEN '4' THEN 'Unit-Quarterly'
	WHEN '5' THEN 'Unit_Annual'
	WHEN '6' THEN 'Manual'
	WHEN '7' THEN 'Mixed'
	ELSE NULL
	END AS node_pricing_model , first_applicant, last_applicant

 from node_tbl inner join (select node_code, sum(invoice_tbl.total) amt from invoice_tbl
 where begin_service_date::date >=  '%s' and begin_service_date::date < '%s' group by node_code ) aa  

 on aa.node_code = node_tbl.node_code
inner join company_tbl on  node_tbl.company_code = company_tbl.company_code LEFT JOIN auxiliary_company AS ac
ON node_tbl.company_code = ac.company_code left JOIN auxiliary_node AS an 
ON node_tbl.node_code = an.node_code
        where    node_tbl.company_code  not in ('APP') and aa.node_code = node_tbl.node_code and node_tbl.company_code = company_tbl.company_code
  ; """  % (start_date, end_date))   
  
 
    return SQL

 

def connect_rentgrow_data_frame(cus_sql):
     
    conn = None
 
    params = config()
    conn = psycopg2.connect(**params)       
    cur = conn.cursor()
   
    
    string_replace = "  ('" + "','".join(name.upper().strip().replace("'", r"\'") for name in list_term_code) + "'"  + ')'
   # SQL_1 =  (o_SQL).replace( regexp=True, to = "request.company_code not in ('APP') ", value = string_replace    )
    cus_sql =  (cus_sql).replace(   "('APP')",   string_replace     ) #live
    
    SQL =  cus_sql
    cur.execute(SQL) 
    col = [i[0] for i in cur.description]
    conn.set_client_encoding('ISO-8859-1') 
    rows = cur.fetchall()
    cur.close()
    data = pd.DataFrame(rows, columns = col)
   
    return data
# Print the list 
   
def send_to_finance(filename1="", filename2=""):
    
    signature_file = "C:\App\Data\Sig\sig.htm"
    if os.path.exists(signature_file):
        
        #signature_path = os.path.join((os.environ['USERPROFILE']), sig_files_path) # Finds the path to Outlook signature files with signature name "Work"
        #html_doc = os.path.join((os.environ['USERPROFILE']),sig_html_path)     #Specifies the name of the HTML version of the stored signature
        #html_doc = html_doc.replace('\\\\', '\\')


        html_file = codecs.open(signature_file, 'r', 'utf-8', errors='ignore') #Opens HTML file and converts to UTF-8, ignoring errors
        signature_code = html_file.read()               #Writes contents of HTML signature file to a string

        #signature_code = signature_code.replace((signature_name + '_files/'), signature_path)      #Replaces local directory with full directory path
        html_file.close()

    
        outlook = win32.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.To = "xiaobin.zhang@yardi.com"

        
        mail.Subject = "Weekly Setup Audit - RS_AM@Yardi.Com ;   "  
        
        #filename2 =  filename1.replace('O:\\ANALYTICS\\New Property Lists\\', '\\\ysifwfs07\\Vol2\ANALYTICS\\New Property Lists\\') # + filename.replace("/", ".") + ".xlsx"
            #path  = "\"\\\\windows_Server\\golobal_directory\\the folder\\file yyymm.xlsx\""
        path = '"' + filename1 + '"'
        string  = """<a href=""" +  path + ' style=text-decoration: none>' + filename1 +  '<' +  r'\a'  + '>'
        string =  string.replace('O:\\ANALYTICS\\Setup_Audit\\', '\\\ysifwfs07\\Vol2\ANALYTICS\\Setup_Audit\\')

        path_2 = '"' + filename2 + '"'
        string_2  = """<a href=""" +  path_2 + ' style=text-decoration: none>' + filename2 +  '<' +  r'\a'  + '>'
        string_2 =  string_2.replace('O:\\ANALYTICS\\Setup_Audit\\', '\\\ysifwfs07\\Vol2\ANALYTICS\\Setup_Audit\\')
        
        #  string.replace('\\a>', '\a>')
        #mail.body = string
        
        mail.HTMLbody =   string + " <BR> " + string_2 +" <BR><BR><BR> "  +signature_code + " <BR><BR><BR> "
    
    
        mail.send


def columnToLetter(column):
    letter = ''
    while column > 0:
        temp = (column - 1) % 26
        letter = chr(temp + 65) + letter
        column = (column - temp - 1) // 26
    return letter
 
def letterToColumn(letter):
    column = 0
    length = len(letter)
    for i in range(length):
        column += (ord(letter[i].upper()) - 64) * 26**(length - i - 1)
    return column

  
def kill_excel():
    
    for proc in psutil.process_iter():
        if proc.name() == "EXCEL.EXE":
            proc.kill()
   
def dispatch(app_name:str):
    try:
        from win32com import client
       # app = client.gencache.EnsureDispatch(app_name)
        app = win32.dynamic.Dispatch("Excel.Application")
    except AttributeError:
        # Corner case dependencies.
        import os
        import re
        import sys
        import shutil
        # Remove cache and try again.
        MODULE_LIST = [m.__name__ for m in sys.modules.values()]
        for module in MODULE_LIST:
            if re.match(r'win32com\.gen_py\..+', module):
                del sys.modules[module]
        shutil.rmtree(os.path.join(os.environ.get('LOCALAPPDATA'), 'Temp', 'gen_py'))
        from win32com import client
        app = client.gencache.EnsureDispatch(app_name)
    return app         
 
def  format_book( df_no_charge_final,  wb, ws_name  ):

    ws = wb[ws_name] 

    if (len(df_no_charge_final)==0):
         ws.cell(2,1,  'No Data')  
         return 
    
    redFill =PatternFill(fill_type='solid',start_color='5B9BD5',end_color='5B9BD5') 
    

    ''' 
    side  = Side(border_style='thin',  color="FF000000")
    border  = Border(left=side, right=side, top=side, bottom=side)
    side_thin  = Side(border_style='thin',  color="FF000000")

    thin_border = Border(
    left=Side(border_style='thin', color='00000000'),
    right=Side(border_style='thin', color='00000000'),
    top=Side(border_style='thin', color='00000000'),
    bottom=Side(border_style='thin', color='00000000'))
    '''
    thick_border = Border(
    #left=Side(border_style='thin', color='00000000'),
    #right=Side(border_style='thin', color='00000000'),
    top=Side(border_style='thick', color='00000000') )#,
   # bottom=Side(border_style='thin', color='00000000') )
    
    dollar_style = NamedStyle(name="dollar_style", number_format='$#,##0.00')
    for i, col in enumerate(df_no_charge_final.columns):
           column_letter = columnToLetter(i+1)
           if re.search('price', col, re.IGNORECASE):
               for cell in ws[column_letter]:  # Change 'B' to your target column
                  cell.style = dollar_style
           column_len = max(df_no_charge_final[col].astype(str).str.len().max(), len(col))  
         #   ws.set_column(i, i, column_len) 
           ws.column_dimensions[column_letter].width = column_len


    for x in range(1,len(df_no_charge_final.columns)+1):
         
        ws.cell(1, x).fill = redFill     
        ws.cell(1, x).font = Font(color="FFFFFF", name="Verdana", size=12)   
        ws.cell(1, x).alignment = Alignment(horizontal='left', vertical = 'center')

    o_companycode=''
    df_no_charge_final.reset_index(inplace = True,drop = True)

    for index, row in df_no_charge_final.iterrows():
        if (o_companycode != str(row[0])  ):
            for col_num  in range(1, len(df_no_charge_final.columns)+1):
                 ws.cell(index+2, col_num).border = thick_border           
      
        ws.cell(index+2, col_num).alignment = Alignment(horizontal='left', vertical = 'center')
        
        o_companycode = str(row[0])  

    for col_num  in range(1, len(df_no_charge_final.columns)+1):
        ws.cell( ((index+3)), col_num).border = thick_border 
        ws.cell(index+3, col_num).alignment = Alignment(horizontal='left', vertical = 'center')

    ws.auto_filter.ref = 'A1:O1'
    ws.freeze_panes ='A2'
    

 
    #ColumnDimension(ws, bestFit=True)

 
        
def run_transaction_part( df_pricing, df_no_charge_final):
      
 
    #df_trans = connect_rentgrow_data_frame (generate_trans_sql(start_date, end_date))
    #df_invoice = connect_rentgrow_data_frame (generate_invoice_zero_sql(start_date, end_date))
    #df_no_charge =   pd.merge(df_trans, df_invoice ,on = [  "node_code"  ], how='inner')

    if (len(df_no_charge_final) < 1):
        return df_no_charge_final
    
    df_node = df_no_charge_final [['node_code', 'match_type']]
    df_node= df_node.drop_duplicates(subset=['node_code' ], keep='first' )
    
 
    aggFunc = { 
           'transaction_uuid' : np.count_nonzero,
            'price' :  'first'
  
           }

    #'Property Name','company_code',, 'total' margins=True, margins_name='Grand Total'
    df_no_charge_final = pd.pivot_table(df_no_charge_final,index=[ "node_code" ] ,
                                     aggfunc=aggFunc, columns=[  'transaction_code'],
                                    values=["transaction_uuid", "price"] ,  dropna=False, fill_value=0, ).reset_index()
    
     
    #df_no_charge_final ['match_type'] =  df_node['node_code'].map(df_node.set_index('node_code')['match_type']).fillna(df_node['match_type'])
  
  
     
    #df_no_charge_final.columns = df_no_charge_final.columns.astype(str).str.replace('count_tran', '').str.replace('(', '').replace(')', '')
    df_no_charge_final.columns = ['_'.join(col) for col in df_no_charge_final.columns.values]
    df_no_charge_final.rename(columns=lambda x: x.replace('transaction_uuid_', '').replace('(', '').replace(')', ''), inplace=True)
    df_no_charge_final.rename(columns=lambda x: x.replace('price', 'Price'), inplace=True)
    df_no_charge_final.rename(columns=lambda x: x.replace("('',", '').replace("')", '').replace(')', ''), inplace=True)
    df_no_charge_final.rename(columns=lambda x: x.strip(), inplace=True)
    df_no_charge_final ["Monthly Invoice$"] = 0
   # list_am = connect_acct_manager()
    #df_new_company['Account Manager']=  df_new_company['Company_Code'].map(list_am.set_index('company_code')['account_manager']).fillna(df_new_company['Account Manager'])
    #df_new_company.to_excel (writer2,  sheet_name= 'Company', index=False, startrow=0,engine='xlsxwriter')
    #df_new_company.columns = [col.replace('_', ' ') for col in df_new_company.columns]
    df_no_charge_final.replace( np.nan, '',inplace = True)

 
               
 
    
    df_no_charge_final = pd.DataFrame(df_no_charge_final.to_records())
    df_no_charge_final.rename(columns=lambda x: x.replace('node_code_', 'node_code'),  inplace=True)
    df_no_charge_final = pd.merge( df_no_charge_final,df_pricing  ,on = ['node_code'], how='left')
    df_no_charge_final=df_no_charge_final.merge( df_node ,on = ['node_code'], how='left') 

    df_no_charge_final.rename(columns=lambda x: x.replace('match_type_', 'match_type'), inplace=True)
    df_no_charge_final = df_no_charge_final.sort_values (by=[ 'company_code', 'node_name'])
    df_no_charge_final = df_no_charge_final.reset_index(drop=True)
    df_no_charge_final = df_no_charge_final.drop('index', axis=1)

    for i, col in enumerate(df_no_charge_final.columns):
           if re.search('price', col, re.IGNORECASE):
               df_no_charge_final[col] = df_no_charge_final[col].apply(lambda x: float(x))

    col = df_no_charge_final.pop ('company_code')
    df_no_charge_final.insert(0, col.name, col)

    col = df_no_charge_final.pop ('company_name')
    df_no_charge_final.insert(1, col.name, col)
    #cols_at_end = ['company_code','company_name' ,'node_code',  'node_name', 'company_pricing_model','node_pricing_model','Price_DOCVERIFY','Price_NOVAINCOME','Price_RENTHISTORY','DOCVERIFY','NOVAINCOME','RENTHISTORY','Monthly Invoice$','canceled','amt']
    #df_no_charge_final = df_no_charge_final[ cols_at_end + [ col for col in df_no_charge_final.columns if col not in cols_at_end ]]
    col = df_no_charge_final.pop ('node_code')
    df_no_charge_final.insert(2, col.name, col)
   
    col = df_no_charge_final.pop ('node_name')
    df_no_charge_final.insert(3, col.name, col)

    col = df_no_charge_final.pop ('company_pricing_model')
    df_no_charge_final.insert(4, col.name, col)

    col = df_no_charge_final.pop ('node_pricing_model')
    df_no_charge_final.insert(5, col.name, col)
    return df_no_charge_final 

    


if __name__ == '__main__':
    
    Manual_Run=0
    if (Manual_Run==1):
       start_date = date(2025, 1, 1)  #date(2021, 9, 1)
       end_date = date(2025, 2, 1)
       delta = timedelta(days=1)
       last_day_month = (start_date.replace(day=1) + timedelta(days=31)).replace(day=1) + timedelta(days=-1)

      
    #print ("last day of month")
    else:
        last_day_month = date.today().replace(day=1) - timedelta(days=1)
        start_date = date.today().replace(day=1) - timedelta(days=last_day_month.day)
        end_date =  (start_date.replace(day=1) + timedelta(days=31)).replace(day=1) + timedelta(days=0)
      
   
    start_date_format = start_date.strftime("%m/%d/%Y")
    last_day_month_format = last_day_month.strftime("%m/%d/%Y")
    end_date_format = end_date.strftime("%m/%d/%Y")   

    current_month_abre = last_day_month.strftime('%b') + ' ' +  last_day_month.strftime('%Y')
    filename1 = os.getcwd() + "\\"  + 'Finance Pricing QA - '  + current_month_abre + ".xlsx" 
 
     
    filename1 = filename1.replace('/','.',10)
    
    if  os.path.exists(filename1):
        os.remove(filename1)

        
    if  os.path.exists(filename1):
        pass
    else:
         writer2 = pd.ExcelWriter(filename1 )
         df_pricing = connect_rentgrow_data_frame(generate_pricing(start_date, end_date))

         #######part one
         df_no_charge_final  = connect_rentgrow_data_frame (generate_current_price_plans_one(start_date, end_date )) 
         df_no_charge_final  = run_transaction_part(df_pricing, df_no_charge_final  )
         df_no_charge_final.to_excel(writer2, index = False, header=True,   sheet_name = 'Pricing Plan' ) #float_format='%.00f',header=True,
 
         
         #######part two
     
         df_no_pricing_final  = connect_rentgrow_data_frame (generate_no_current_price_two(start_date, end_date )) 
         df_no_pricing_final = run_transaction_part(df_pricing,df_no_pricing_final )
         df_no_pricing_final.to_excel(writer2, index = False, header=True,   sheet_name = 'No Pricing Plan' ) #flo 

 
         writer2.close() 

         wb = load_workbook(filename1) 
 
         format_book( df_no_charge_final,  wb , 'Pricing Plan') 
         format_book( df_no_pricing_final,  wb , 'No Pricing Plan') 
    

         wb.save(filename1)

 #   send_to_finance(filename1_2, filename2_2 ) 

    

    
sys.exit(0)
quit()
