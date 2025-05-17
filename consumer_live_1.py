 
import psycopg2
 
from config_reporting import config_reporting
from config_rentgrow import config_rentgrow
import datetime
from datetime import date, timedelta
from datetime import datetime, time
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.cell import Cell  
import pandas as pd
import numpy as np
import os
import sys
import logging
import locale
from openpyxl.styles import Alignment
import dateutil.relativedelta
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Color, Fill, Font, Border,Side, PatternFill
from openpyxl.styles import colors
import openpyxl 
from openpyxl import Workbook
from datetime import date, timedelta
from win32com import client
import win32api
import pathlib
from functools import reduce
import tkinter as tk
from tkinter import simpledialog
from time import sleep
import easygui
import warnings
warnings.filterwarnings("ignore")
import re 
from tkinter import filedialog
from tkinter import * 
global year_start  

dirpath = os.getcwd()
print (dirpath)
#file_template =  dirpath +  '\\Template_2022.xlsx'
file_template =  dirpath +  '\\Template_2025.xlsx'
file_template = file_template.replace('/','.',10)

year_start = date(2025, 1, 1) 

global file_ycrm 
file_ycrm=  "O:\ANALYTICS\Consumer Relations" +  '\\ycrm_case.xlsx'

exception_code = ('APP','APP2','APPCK','AUTO','BETAA','BZ001','BZ002','BZ655','BZ689','BZ805','BZ890','BZ903','BZ907','BZ965','BZB22','BZB27','BZB92','BZC78','BZC79','BZF80','BZG46','DEMO','ESLES','FAKE','GEN2','GREYT','INTF2','INTF3','INTF4','INTFC','PRCRD','RCCRM','RGSAL','RGSW','RGTST','RICHC','RICHD','RICHE','RICHK','SALES','STIJL','SWDMO','SWRKS','TEST9','TESTR','TRAIN','TSEXE','TSTER','VOYA','XTEST','YASC','YRK', 'SYLJ', 'RICHT', '2SYLJ')

#filename1 =   dirpath +  'CAPIL Screening - ' + str(start_date) + '-' + str(end_date) + '.xlsx'

START_ROW = 3


try:
    #mutex = win32event.CreateMutex(None, False, 'name')
    #last_error = win32api.GetLastError(easygui)

    if  not os.path.exists(file_template):
        output = easygui.msgbox("There is no template  file! ", "Error")
        sys.exit(0)
        quit(0)
    
    
		 
except:
    sys.exit(0)



def generate_terminate_code ( ):
    conn = None
    params = config_reporting()
    
    conn = psycopg2.connect(**params )
 
    cur = conn.cursor()

    SQL = ("""   select trim(company_code)  from "Test_Code_Master"  ; """   )
        
    cur.execute(SQL) 
    #col = cur.description
    col = [i[0] for i in cur.description]
    a = np.array(col)
    rows =   [item[0] for item in cur.fetchall()]
            
    #df_codes = pd.DataFrame(rows, columns = a)
    cur.close()
    conn.close()
    return rows

list_term_code =  generate_terminate_code () 

#ist_node = list([x.strip() for x in list_term_code ])

def generate_criminal_1(start_date, end_date,   wb, data_dict ):
    
    
    ws = wb.get_sheet_by_name('Criminal')
 
    o_SQL = (""" 
    
      select  num_1,   num_1 -(( num_2_1 + num_2_2) - num_2_3)  as num_2, num_3_1   as num_3 , num_4_1 as num_4 
 
 from 
 
 (select count(*) num_1   from request  where  
    requesttypeid = 19 and request.created >=   '%s' and created < '%s' and 
 upper(trim(company_code)) not in ('APP') )  aa ,
 
(select  count(distinct request.requestid) num_2_1  from request,requesttoaction   where  request.requestid  = requesttoaction.requestid and  
   requesttypeid = 19 and stamp>= '%s' and stamp < '%s' and 
upper(trim(request.company_code)) not in ('APP') and  actionid= 5 and coalesce(requesttoaction.employeeid,0)=0  ) bb_1,
 
(select count(*) num_2_2  from request where 
   requesttypeid = 19 and request.created >= '%s' and created< '%s' and 
upper(trim(request.company_code)) not in ('APP') and employeeid in ( 649,177,-1,-2,-6) ) bb_2 , 
  
 ( select count(*) num_2_3   from request where 
   requesttypeid = 19 and request.created  >= '%s' and created<'%s' and 
upper(trim(request.company_code)) not in ('APP') and employeeid in ( 496) and request.resultcode = '2' ) bb_3 , 
 
 ( select count(*) num_3_1    from request where 
   requesttypeid = 19 and request.created  >= '%s' and created< '%s'
 and upper(trim(request.company_code)) not in ('APP') and  (request.resultcode = '2' or sexoffenderhits='2') and  employeeid not in ( 649,177,-1,-2,-6) ) cc_1 ,
  
 ( select count(*) num_4_1    from request where 
   requesttypeid = 19 and request.created  >= '%s' and created< '%s'  
 and upper(trim(request.company_code)) not in ('APP') and   (request.resultcode = '2' or sexoffenderhits='2') and  employeeid not in ( 649,177,-1,-2,-6) and meets_requirements='2') dd_1  
 
 
   ; """  % (start_date, end_date,start_date, end_date,start_date, end_date,start_date, end_date,start_date, end_date,start_date, end_date ))
    
     
    
    #o_SQL = ("""  select count(*) from node_tbl where  company_code not in ( 'APP' )  ; """   )
    
   # string_replace = "   request.company_code not in ('" + "','".join(name.upper().strip().replace("'", r"\'") for name in list_term_code) + "'"  + ')'
    string_replace = "  ('" + "','".join(name.upper().strip().replace("'", r"\'") for name in list_term_code) + "'"  + ')'
   # SQL_1 =  (o_SQL).replace( regexp=True, to = "request.company_code not in ('APP') ", value = string_replace    )
    SQL_1 =  (o_SQL).replace(   "('APP')",   string_replace     ) #live
   # SQL_1 = re.sub("('APP')", string_replace, o_SQL, flags=re.IGNORECASE)
     
    col_num = {1: 2,
            2: 3,
            3: 5,
            4: 7}
    rows_1=get_rows (SQL_1)
    for r_idx, row in enumerate(rows_1, 1):         
            for c_idx, value_1 in enumerate(row, 1):
                   ws.cell(row=START_ROW, column=col_num[c_idx] , value=value_1)
                   ws.cell(row=START_ROW,column=col_num[c_idx]).number_format = '#,##0'
                  
     
    ycrm_data_crim  =  data_dict[(data_dict["Status"] == 'Closed') & (data_dict["Issue Type"].str.contains('criminal',  regex=False, case=False))]
              

    ycrm_multiple_service_crim  =   data_dict[ (data_dict["Status"] == 'Closed') & (data_dict["Issue Type"] == 'RS Applicant | Multiple Services') &  (data_dict['Brief Description'].str.contains('crim',  regex=False, case=False))]    
    
    ws.cell(row=START_ROW, column=9, value=len(ycrm_data_crim ) + len (ycrm_multiple_service_crim) ) #b7
    ws.cell(row=START_ROW, column=9  ).alignment = Alignment(horizontal='center', vertical = 'center')


    ycrm_data_crim_succ = ycrm_data_crim[ (ycrm_data_crim["Secondary Status"]=='RS Dispute Complete - Result Changed') |  (ycrm_data_crim["Secondary Status"]=='RS Dispute Complete - Result Not Changed')  ]
    ycrm_multiple_service_crim_succ= ycrm_multiple_service_crim[ (ycrm_multiple_service_crim["Secondary Status"]=='RS Dispute Complete - Result Changed') | (ycrm_multiple_service_crim["Secondary Status"]=='RS Dispute Complete - Result Not Changed') ]
    ws.cell(row=START_ROW, column=11, value=len(ycrm_data_crim_succ ) + len (ycrm_multiple_service_crim_succ) ) #b8
    ws.cell(row=START_ROW, column=11  ).alignment = Alignment(horizontal='center', vertical = 'center')


    ycrm_data_crim_changed = ycrm_data_crim[ (ycrm_data_crim["Secondary Status"]=='RS Dispute Complete - Result Changed')   ]
    ycrm_multiple_service_crim_changed= ycrm_multiple_service_crim[ (ycrm_multiple_service_crim["Secondary Status"]=='RS Dispute Complete - Result Changed')  ]
    ws.cell(row=START_ROW, column=15, value=len(ycrm_data_crim_changed ) + len (ycrm_multiple_service_crim_changed) ) #b9
    ws.cell(row=START_ROW, column=15  ).alignment = Alignment(horizontal='center', vertical = 'center')

    ws.cell(row=START_ROW,column= 9 ).number_format = '#,##0'
    ws.cell(row=START_ROW,column= 11 ).number_format = '#,##0'
    ws.cell(row=START_ROW,column= 15 ).number_format = '#,##0'

    return True

def generate_civil_1(start_date, end_date,  wb, data_dict ):
    
    ws = wb.get_sheet_by_name('Civil Courts')
 
    o_SQL = (""" 
    
  select  num_1,   num_3_1 + num_2_2 +  num_2_3 as num_2, num_3_1   as num_3 , num_4_1 as num_4 
 from 
 
 (select count(*) num_1   from request where requesttypeid in (27,28) and request.created >=   '%s' and created < '%s' and 
upper(trim(request.company_code)) not in ('APP') )  aa ,
  
(select count(*) num_2_2  from request where 
   requesttypeid in (27,28) and request.created >= '%s' and created< '%s' and 
upper(trim(request.company_code)) not in ('APP') and version = 20) bb_2 , 
  
 ( select count(distinct r.requestid) num_2_3   FROM request r, applicant a, civilcourt_record ccr, civilcourt_scoring ccs where 
   requesttypeid in (27,28) and r.created  >= '%s' and r.created<'%s' and r.applicantid = a.applicantid AND r.requestid = ccr.requestid 
   AND ccr.id = ccs.recordid AND ccs.stype IN ('Dismissal','Vacate') AND r.version = 40 and r.resultcode = 1 
   and upper(trim(r.company_code)) not in ('APP') 
 ) bb_3 , 
 
 ( select count(*) num_3_1    from request where  ----------------------------------------------------------------------G5
   requesttypeid in (27,28) and request.created  >= '%s' and created< '%s'
 and upper(trim(request.company_code)) not in ('APP') and  request.resultcode = '2'   ) cc_1 ,
  
 ( select count(*) num_4_1    from request where  -------------------------------------------------g6
   requesttypeid in (27,28 ) and request.created  >= '%s' and created< '%s'  
 and upper(trim(request.company_code)) not in ('APP') and  request.resultcode = '2' and  meets_requirements='2') dd_1  
 
   ; """  % (start_date, end_date,start_date, end_date,start_date, end_date,start_date, end_date,start_date, end_date ))
    
    #o_SQL = ("""  select count(*) from node_tbl where  company_code not in ( 'APP' )  ; """   )
    
   # string_replace = "   request.company_code not in ('" + "','".join(name.upper().strip().replace("'", r"\'") for name in list_term_code) + "'"  + ')'
    string_replace = "  ('" + "','".join(name.upper().strip().replace("'", r"\'") for name in list_term_code) + "'"  + ')'
   # SQL_1 =  (o_SQL).replace( regexp=True, to = "request.company_code not in ('APP') ", value = string_replace    )
    SQL_1 =  (o_SQL).replace(   "('APP')",   string_replace     )
   # SQL_1 = re.sub("('APP')", string_replace, o_SQL, flags=re.IGNORECASE)
    
    col_num = {1: 2,
            2: 3,
            3: 5,
            4: 7}
   
    rows_1=get_rows (SQL_1)
    for r_idx, row in enumerate(rows_1, 1):         
            for c_idx, value_1 in enumerate(row, 1):
                ws.cell(row=START_ROW, column=col_num[c_idx] , value=value_1)
                ws.cell(row=START_ROW, column=col_num[c_idx]  ).alignment = Alignment(horizontal='center', vertical = 'center')
                ws.cell(row=START_ROW,column=col_num[c_idx]).number_format = '#,##0'

    ycrm_data_civil  =  data_dict[(data_dict["Status"] == 'Closed') & (data_dict["Issue Type"].str.contains('civil',  regex=False, case=False))]
    
    ycrm_multiple_service_civil  =    data_dict[ (data_dict["Status"] == 'Closed') & (data_dict["Issue Type"] == 'RS Applicant | Multiple Services') &  (data_dict['Brief Description'].str.contains('civil',  regex=False, case=False))]    
    
    ws.cell(row=START_ROW, column=9, value=len(ycrm_data_civil ) + len (ycrm_multiple_service_civil) ) #b7
    ws.cell(row=START_ROW, column=9  ).alignment = Alignment(horizontal='center', vertical = 'center')



    ycrm_data_civil_succ = ycrm_data_civil[ (ycrm_data_civil["Secondary Status"]=='RS Dispute Complete - Result Changed') |  (ycrm_data_civil["Secondary Status"]=='RS Dispute Complete - Result Not Changed')  ]
    ycrm_multiple_service_civil_succ= ycrm_multiple_service_civil[ (ycrm_multiple_service_civil["Secondary Status"]=='RS Dispute Complete - Result Changed') | (ycrm_multiple_service_civil["Secondary Status"]=='RS Dispute Complete - Result Not Changed') ]
    ws.cell(row=START_ROW, column=11, value=len(ycrm_data_civil_succ ) + len (ycrm_multiple_service_civil_succ) ) #b8
    ws.cell(row=START_ROW, column=11  ).alignment = Alignment(horizontal='center', vertical = 'center')


    ycrm_data_civil_changed = ycrm_data_civil[ (ycrm_data_civil["Secondary Status"]=='RS Dispute Complete - Result Changed')   ]
    ycrm_multiple_service_civil_changed= ycrm_multiple_service_civil[ (ycrm_multiple_service_civil["Secondary Status"]=='RS Dispute Complete - Result Changed')  ]
    ws.cell(row=START_ROW, column=15, value=len(ycrm_data_civil_changed ) + len (ycrm_multiple_service_civil_changed) ) #b9
    ws.cell(row=START_ROW, column=15  ).alignment = Alignment(horizontal='center', vertical = 'center')

    ws.cell(row=START_ROW,column= 9).number_format = '#,##0'
    ws.cell(row=START_ROW,column= 11).number_format = '#,##0'
    ws.cell(row=START_ROW,column= 15).number_format = '#,##0'

    return True

def generate_credit_1(start_date, end_date,  wb , data_dict):
    
    ws = wb.get_sheet_by_name('Credit')
 

    o_SQL = ("""  select  count(*) 
 from trans_tbl, node_tbl  where   upper(trim(company_code))  not in ('APP')  and transaction_date::date  >='%s' and transaction_date::date < '%s' 
 and transaction_code='1STBUREAU' and trans_tbl.node_code = node_tbl.node_code and coalesce(exception_code,'')<>'dup'
  ; """  % (start_date, end_date  ))

    string_replace = "  ('" + "','".join(name.upper().strip().replace("'", r"\'") for name in list_term_code) + "'"  + ')'
   # SQL_1 =  (o_SQL).replace( regexp=True, to = "request.company_code not in ('APP') ", value = string_replace    )
    SQL_1 =  (o_SQL).replace(   "('APP')",   string_replace     )
   # SQL_1 = re.sub("('APP')", string_replace, o_SQL, flags=re.IGNORECASE)
 
   
    rows_1=get_rows (SQL_1)
    for r_idx, row in enumerate(rows_1, 1):         
            for c_idx, value_1 in enumerate(row, 1):
                if ( c_idx ==3 ):
                   ws.cell(row=START_ROW, column=c_idx+2 , value=value_1)
                   ws.cell(row=START_ROW, column=c_idx+2 ).alignment = Alignment(horizontal='center', vertical = 'center')
                   ws.cell(row=START_ROW,column= c_idx+2).number_format = '#,##0'
                elif ( c_idx ==4 ):
                   ws.cell(row=START_ROW, column=c_idx+3 , value=value_1)
                   ws.cell(row=START_ROW, column=c_idx+3).alignment = Alignment(horizontal='center', vertical = 'center')
                   ws.cell(row=START_ROW,column= c_idx+3).number_format = '#,##0'
                else :
                   ws.cell(row=START_ROW, column=c_idx+1 , value=value_1)
                   ws.cell(row=START_ROW, column=c_idx+1 ).alignment = Alignment(horizontal='center', vertical = 'center')
                   ws.cell(row=START_ROW,column= c_idx+1).number_format = '#,##0'
    

    
    ycrm_data_credit  =  data_dict[(data_dict["Status"] == 'Closed') & 
                ((data_dict["Issue Type"] == 'RS Applicant | Credit Dispute') )    
                   ]
    ycrm_multiple_service_credit  =  ( data_dict[(data_dict["Status"] == 'Closed') & (data_dict["Issue Type"] == 'RS Applicant | Multiple Services') &  (data_dict['Brief Description'].str.contains('credit',  regex=False, case=False))]  )  
    
    ws.cell(row=START_ROW, column=3, value=len(ycrm_data_credit ) + len (ycrm_multiple_service_credit) ) #b7
    
    ycrm_data_credit_succ = ycrm_data_credit[ (ycrm_data_credit["Secondary Status"]=='RS Dispute Complete - Result Changed') |  (ycrm_data_credit["Secondary Status"]=='RS Dispute Complete - Result Not Changed')  ]
    ycrm_multiple_service_credit_succ= ycrm_multiple_service_credit[ (ycrm_multiple_service_credit["Secondary Status"]=='RS Dispute Complete - Result Changed') | (ycrm_multiple_service_credit["Secondary Status"]=='RS Dispute Complete - Result Not Changed') ]
    ws.cell(row=START_ROW, column=5, value=len(ycrm_data_credit_succ ) + len (ycrm_multiple_service_credit_succ) ) #b8

    ycrm_data_credit_changed = ycrm_data_credit[ (ycrm_data_credit["Secondary Status"]=='RS Dispute Complete - Result Changed')   ]
    ycrm_multiple_service_credit_changed= ycrm_multiple_service_credit[ (ycrm_multiple_service_credit["Secondary Status"]=='RS Dispute Complete - Result Changed')  ]
    ws.cell(row=START_ROW, column=7, value=len(ycrm_data_credit_changed ) + len (ycrm_multiple_service_credit_changed) ) #b9
    
    ws.cell(row=START_ROW,column= 3).number_format = '#,##0'
    ws.cell(row=START_ROW,column= 5).number_format = '#,##0'
    ws.cell(row=START_ROW,column= 7).number_format = '#,##0'

def generate_rental_1 (start_date, end_date, wb, data_dict):

    
    ws = wb.get_sheet_by_name('Rental History')
    
    o_SQL = ("""
    
    select num_1, num_2, num_4, num_5  from ( --num_3, 
select count(*) as num_1 from "Rent Bureau Transactions" 
where upper(trim("Rent Bureau Transactions".company_code))  not in (select upper(trim(company_code)) from "Test_Code_Master")
         and      create_stamp>='%s' and create_stamp < '%s' and upper(trim(company_code)) not in ('APP') ) aa, 

(select count(*) as num_2 from "Rent Bureau Transactions" 
where  upper(trim("Rent Bureau Transactions".company_code))  not in (select upper(trim(company_code)) from "Test_Code_Master")
             and  create_stamp>='%s' and create_stamp < '%s'
and  trim(upper(hit)) = 'HIT' and hit_type = 'RentalHistory' and   upper(trim(company_code))  not in ('APP') )  bb ,

 --(select count(*) as num_3 from "Rent Bureau Transactions"  --------- hit 
--where  create_stamp>='%s' and create_stamp < '%s'
--and  trim(upper(hit)) = 'HIT' and hit_type = 'RentalHistory' and 
--(coalesce(reason1,'') <>'' or coalesce(reason2,'')<>'' or  coalesce(reason3,'')<>'' 
--or coalesce(reason4,'') <>'' or coalesce(reason5,'')<>'' and   upper(upper(trim(company_code)))  not in ('APP'))) cc,

(select count(*) as num_4 from "Rent Bureau Transactions" 
where  upper(trim("Rent Bureau Transactions".company_code))  not in (select upper(trim(company_code)) from "Test_Code_Master")
             and create_stamp>='%s' and create_stamp < '%s'
and  trim(upper(hit)) = 'HIT' and hit_type = 'RentalHistory' and 
coalesce(posrentalhist,'') ='Meets Property Requirements' and   upper(trim(company_code))  not in ('APP') )  dd,

(select count(*) as num_5 from "Rent Bureau Transactions" 
where  upper(trim("Rent Bureau Transactions".company_code))  not in (select upper(trim(company_code)) from "Test_Code_Master")
             and create_stamp>='%s' and create_stamp < '%s'
and  trim(upper(hit)) = 'HIT' and hit_type = 'RentalHistory' and 
coalesce(negrentalhist,'')='Does Not Meet Property Requirements' and   upper(trim(company_code))  not in ('APP') )  ee 
 ; """  % (start_date, end_date ,start_date, end_date ,start_date, end_date ,start_date, end_date ,start_date, end_date  ))
    

    string_replace = "  ('" + "','".join(name.upper().strip().replace("'", r"\'") for name in list_term_code) + "'"  + ')'
   # SQL_1 =  (o_SQL).replace( regexp=True, to = "request.company_code not in ('APP') ", value = string_replace    )
    SQL_1 =  (o_SQL).replace(   "('APP')",   string_replace     ) #live
    rows_1=get_rows (SQL_1, 'reporting')

    col_num = {1: 2,
            2: 3,
            3: 5,
            4: 7}
    
    for r_idx, row in enumerate(rows_1, 1):         
            for c_idx, value_1 in enumerate(row, 1):
                ws.cell(row=START_ROW, column=col_num[c_idx] , value=value_1)
                ws.cell(row=START_ROW, column=col_num[c_idx] ).alignment = Alignment(horizontal='center', vertical = 'center')
                ws.cell(row=START_ROW,column= col_num[c_idx] ).number_format = '#,##0'

    ycrm_data_rentbureau  =  data_dict[(data_dict["Status"] == 'Closed') & ( data_dict["Issue Type"].str.contains(pat = 'rent',  regex=True, case=False)) ]
  
    ycrm_multiple_service_rentbureau  =   data_dict[ (data_dict["Status"] == 'Closed') & (data_dict["Issue Type"] == 'RS Applicant | Multiple Services') &  (data_dict['Brief Description'].str.contains('rentburea|rental',  regex=True, case=False))]    ## h.+re ? optional and plus is must
    
    ws.cell(row=START_ROW, column=9, value=len(ycrm_data_rentbureau ) + len (ycrm_multiple_service_rentbureau) ) #b7
    ws.cell(row=START_ROW, column=9).alignment = Alignment(horizontal='center', vertical = 'center')


    ycrm_data_rentbureau_succ = ycrm_data_rentbureau[ (ycrm_data_rentbureau["Secondary Status"]=='RS Dispute Complete - Result Changed') |  (ycrm_data_rentbureau["Secondary Status"]=='RS Dispute Complete - Result Not Changed')  ]
    ycrm_multiple_service_rentbureau_succ= ycrm_multiple_service_rentbureau[ (ycrm_multiple_service_rentbureau["Secondary Status"]=='RS Dispute Complete - Result Changed') | (ycrm_multiple_service_rentbureau["Secondary Status"]=='RS Dispute Complete - Result Not Changed') ]
    ws.cell(row=START_ROW, column=11, value=len(ycrm_data_rentbureau_succ ) + len (ycrm_multiple_service_rentbureau_succ) ) #b8
    ws.cell(row=START_ROW, column=11).alignment = Alignment(horizontal='center', vertical = 'center')


    ycrm_data_rentbureau_changed = ycrm_data_rentbureau[ (ycrm_data_rentbureau["Secondary Status"]=='RS Dispute Complete - Result Changed')   ]
    ycrm_multiple_service_rentbureau_changed= ycrm_multiple_service_rentbureau[ (ycrm_multiple_service_rentbureau["Secondary Status"]=='RS Dispute Complete - Result Changed')  ]
    ws.cell(row=START_ROW, column=14, value=len(ycrm_data_rentbureau_changed ) + len (ycrm_multiple_service_rentbureau_changed) ) #b9
    ws.cell(row=START_ROW, column=14).alignment = Alignment(horizontal='center', vertical = 'center')

    ws.cell(row=START_ROW,column= 9 ).number_format = '#,##0'
    ws.cell(row=START_ROW,column= 11 ).number_format = '#,##0'
    ws.cell(row=START_ROW,column= 14 ).number_format = '#,##0'
 
     

def generate_ycrm_dispute_1():
    data_dict = pd.read_excel(file_ycrm, header=0) #reading file
    #ycrm_data = ycrm_data.replace(np.nan,'', regex=True)
    data_dict = data_dict.replace('\n', ' ').replace('\r', ' ') 
    data_dict = data_dict.replace('\\n', ' ').replace('\\r', ' ') 
    data_dict.iloc[0].replace(to_replace=[r"\\t|\\n|\\r", "\t|\n|\r"], value=["",""], regex=True, inplace=True ) 
    data_dict.columns = [x.replace("\n", " ") for x in data_dict.columns.tolist()]
    #data_dict=data_dict.drop(['Property ID'], inplace=True, axis=1 ) 
    data_dict.rename(columns=lambda x: x.strip(),   inplace=True)
    data_dict = data_dict[['Case ID', 'Status', 'Issue Type', 'Brief Description', 'Secondary Status'  ]]
     

   ############crim########
    #data_dict = data_dict.dropna()
    data_dict = data_dict[ (~data_dict['Case ID'].isnull()) | (~data_dict['Case ID'] =='')]
    ''' 
    ycrm_data_crim  =  data_dict[(data_dict["Status"] == 'Closed') & 
                ((data_dict["Issue Type"] == 'RS Applicant | Criminal Dispute') | (data_dict["Issue Type"] == 'RS Applicant | Criminal Identity Dispute') |
                 (data_dict["Issue Type"] == 'RS Applicant  | Identity Theft Criminal Dispute') | (data_dict["Issue Type"] == 'RS Screening Policy  | Criminal')
                   )]
    '''

    return data_dict 
  

    ############civil########

def get_rows (sql, connect = 'rentgrow'):
    conn = None

    if (connect != 'rentgrow'):
        params = config_reporting()
    else:
        params = config_rentgrow()
    
    conn = psycopg2.connect(**params )
 
    cur = conn.cursor()

    cur.execute(sql) 
    #col = cur.description
    col = [i[0] for i in cur.description]
    print(col)
    #a = np.array(col)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return rows 

 

if __name__ == '__main__':
    manual =0

    if (manual ==0):
        last_day_month = date.today().replace(day=1) - timedelta(days=1)
       
        start_day_month = date.today().replace(day=1) - timedelta(days=last_day_month.day) # previous month start date

        #start_day_month = start_day_month.strftime("%x") # date to string 
        #last_day_month = last_day_month.strftime("%x") #current month first day
        #current_month = current_month.strftime("%x") #prev month last day
    else :
 
        start_day_month = date(2020, 9, 1) # current date to run last month
       # current_month = (start_day_month.replace(day=1) + timedelta(days=31)).replace(day=1)
       # delta = timedelta(days=1)
       # last_day_month = (start_day_month.replace(day=1) + timedelta(days=31)).replace(day=1) + timedelta(days=-1)
        print ("last day of month")
   
   
    start_date_format = start_day_month.strftime("%m/%d/%Y")  
    app = tk.Tk()
    app.eval("tk::PlaceWindow . center")
 
    app.geometry('250x250+500+500')
    app.update_idletasks()
    app.withdraw()

    USER_Date = simpledialog.askstring(title = "Monthly Metrics", prompt = "Entire Start Date in MM/DD/YYYY format           ", initialvalue=start_date_format)
    start_date_month_format = USER_Date 

     

    #file_ycrm =  filedialog.askopenfilename(initialdir = dirpath + "/",title = "Select YCRM file       ",filetypes = (("excel files","*.xlsx"),("all files","*.*")))
    file_ycrm = filedialog.askopenfilename(initialdir = "O:\ANALYTICS\Consumer Relations" + "/",title = "Select YCRM file       ",filetypes = (("excel files","*.xlsx"),("all files","*.*"))) 
    
    if  not os.path.exists(file_ycrm):
        output = easygui.msgbox("There is no ycrm file! ", "Error")
        sys.exit(0)
        quit(0)

    if USER_Date is None :
        print ("NO Code")
        sys.exit(0)
        quit()

   
    start_day_month  = datetime.strptime(start_date_month_format, "%m/%d/%Y") #date format
    
    #last_day_month =start_day_month.replace(day=1) - timedelta(days=1)
    current_month = (start_day_month.replace(day=1)  + timedelta(days=32)).replace(day=1)
    
    while year_start < start_day_month.date(): 
    
         if (year_start != start_day_month.date()):
              
             year_start =(year_start.replace(day=1) + timedelta(days=31)).replace(day=1)
             START_ROW = START_ROW +1 

   

    
     
    month_abre = start_day_month.strftime('%b') + ' ' +  start_day_month.strftime('%Y') 

         
    start_day_month_format  = start_day_month.strftime("%m/%d/%Y")    
    current_month_format = current_month.strftime("%m/%d/%Y")  
   # print (start_date.strftime("%Y-%m-%d"))
        #e_date = (start_date.replace(day=1) + timedelta(days=31)).replace(day=1) + #imedelta(days=-1)
    

    wb = load_workbook(file_template)    
    data_dict = generate_ycrm_dispute_1 ()
   
    #generate_civil_1 (start_day_month_format, current_month_format,   wb ,data_dict ) 
    
    generate_criminal_1 (start_day_month_format, current_month_format,   wb,  data_dict )  
    generate_civil_1 (start_day_month_format, current_month_format,   wb ,data_dict ) 
    generate_credit_1 (start_day_month_format, current_month_format,  wb ,data_dict )  
    generate_rental_1 (start_day_month_format, current_month_format,  wb ,data_dict  )  
    

            #month_abre = start_day_month.strftime('%b') + ' ' +  start_day_month.strftime('%Y')
    start_day_month =(start_day_month.replace(day=1) + timedelta(days=31)).replace(day=1)
    start_day_month_format  = start_day_month.strftime("%m/%d/%Y")    
 
        
    filename1 = dirpath + "\\"  + 'Consumer_Relations_Metrics_' + month_abre  + ".xlsx"
    filename1 = filename1.replace('/','.',10)
    wb.save(filename1   )
    
    filename1 = filename1.replace('/','.',10)
    filename1 = file_template
    wb.save(filename1   )
    print ("file saved")

   
    sys.exit(0)

sys.exit(0)
quit()
