 
import psycopg2
 
import datetime
from datetime import date, timedelta
#import xlsxwriter
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.cell import Cell  
import pandas as pd
import numpy as np
import os
import sys
import logging
import locale
from openpyxl.styles import Alignment
from config import config
from config2 import config2  
import dateutil.relativedelta
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Color, Fill, Font, Border,Side, PatternFill
from openpyxl.styles import colors
import openpyxl 
from openpyxl import Workbook
from datetime import date, timedelta
from win32com import client
import win32api
import pathlib
from functools import reduce
from xlutils.copy import copy #work xls
from xlrd import open_workbook
import xlwt
#import matplotlib.pyplot as plt
#import matplotlib 
import openpyxl 

dirpath = os.getcwd()
print (dirpath)
MAXCOL=68
 
last_day_month = date.today().replace(day=1) - timedelta(days=1)
start_day_month = date.today().replace(day=1) - timedelta(days=last_day_month.day)

start_day__qt =start_day_month - dateutil.relativedelta.relativedelta(months=2)
 
file_template =  dirpath +  '\\template.xlsx'

file_data =  dirpath +  '\\berk.xls'
 
  
filename1 =   dirpath +  '\\Berk_Application' + str(last_day_month) #+ '.xlsx'

 
thin_border = Border(

left=Side(border_style='thin', color='00000000'),
right=Side(border_style='thin', color='00000000'),
top=Side(border_style='thin', color='00000000'),
bottom=Side(border_style='thin', color='00000000'))

start_row = 7
  


def get_data(wb, df_applicant):

    ws =  wb.get_sheet_by_name ('All Applicant Data')
   # df_applicant.reset_index()

    df_row_state = get_db_state()

    df_all = pd.merge(df_applicant, df_row_state ,on = ['Property ID'], how='inner')


    df_row = dataframe_to_rows(df_all )  
    for index, row in enumerate(df_row, 1): 
            for c_idx, value in enumerate(row, 1):
                if (c_idx <=1 or index <=1):
                    continue 
                col = c_idx -1 
                row_num = index + start_row-1
                ws.cell(row_num, col).border = thin_border
                ws.cell(row_num, col, value) 
                ws.cell(row_num, col).alignment = Alignment(horizontal='left')  

                if (col ==6):
                    ws.cell(row_num, col).number_format = '"$"#,##0'   
                    ws.cell(row_num, col).alignment = Alignment(horizontal='center')      
                elif (col== 7):
                    ws.cell(row_num, col).number_format = '0'
                    ws.cell(row_num, col).alignment = Alignment(horizontal='center')      
                 
                #else :
                    #ws.cell(row=index+start_row,column=c_idx).number_format = '0%'
    
    df_applicant_title = pd.read_excel(file_data, header=start_row-1, sheet_name="All Applicant Data") #reading file
    df_applicant_title_date = df_applicant_title.head(1).columns[0]
    ws.cell(7, 1, df_applicant_title_date ) 


                
def get_prop (wb, df_applicant):
    ws =  wb.get_sheet_by_name ('Average Income by Property')
   
    if (len(df_applicant) <=0) :
       return; 

    ''' 
    application_data_grp  = df_applicant.groupby(by =  ['Property ID', 'Property Name']).agg ( {
       "Application Monthly Income": ['mean' ], 
       "Rent To Income Ratio (%)":  ['mean' ]
     
       #"Rent To Income Ratio (%)": lambda x :  (x <= 1 and x >=998).mean() ,
       #"Application Monthly Income": lambda x : ( x >8.3 and x < 166667).mean()  
      
     }  ).reset_index()
    '''

   # df_applicant['Application Monthly Income'] = df_applicant['Application Monthly Income'].fillna(df_applicant['Application Monthly Income'].mean())
    #df_applicant['Application Monthly Income'] = df_applicant['Application Monthly Income'].astype(object).fillna("UNKNOWN")
    #df_applicant['Rent To Income Ratio (%)'] = df_applicant['Rent To Income Ratio (%)'].astype(object).fillna("UNKNOWN")
   # df_applicant['Rent To Income Ratio (%)'] = df_applicant['Rent To Income Ratio (%)'].fillna(df_applicant['Rent To Income Ratio (%)'].mean())
    df_applicant ['new_name'] = df_applicant["Property ID"]  + ";" + df_applicant["Property Name"]  

    aggFunc = {'Application Monthly Income': np.nanmean,
           'Rent To Income Ratio (%)': np.nanmean,
           'Application ID' : np.count_nonzero
           }
    

    #'Property Name',
    application_data_grp = pd.pivot_table(df_applicant,index=[ "new_name" ] ,
                                    aggfunc=aggFunc,
                                    values=['Application Monthly Income','Rent To Income Ratio (%)', 'Application ID'],
                                    margins=True, margins_name='Grand Total',  dropna=False).reset_index()
    
     
    #application_data_grp = application_data_grp.sort_values(['Property Name', 'Property ID'])
    #prop_start_row = 2

    application_data_grp[['Property ID','Property Name']] = application_data_grp['new_name'].str.split(';',expand=True)
    application_data_grp.sort_values ("Property Name", inplace=True)

  #  application_data_grp ["Property Name"] = application_data_grp["new_name"].str.split(";") [0]
   # application_data_grp ["Property ID"] = application_data_grp["new_name"].str.split(";") [1]
    
    application_data_grp = application_data_grp.drop ('new_name', axis=1) 
 

    new_group_app =   application_data_grp.iloc[:, [4, 3, 0, 1, 2]]
     
  
    df_row = dataframe_to_rows(new_group_app )  
    for index, row in enumerate(df_row, 1): 
            for c_idx, value in enumerate(row, 1):
                if (c_idx <=1 or index <=2):
                    continue 
                col = c_idx -1
                row_num = index +1
                ws.cell(row_num, col).border = thin_border
                
                ws.cell(row_num, col, value=value) 
                ws.cell(row_num, col).alignment = Alignment(horizontal='center')  

                if (col ==4):
                    ws.cell(row_num,column=col).number_format = '"$"#,##0'       
                elif (col== 5):
                    ws.cell(row_num,column=col).number_format = '0'
                else: 
                    pass
                #else :
                    #ws.cell(row=index+start_row,column=c_idx).number_format = '0%'
    
    df_applicant_title = pd.read_excel(file_data, header=start_row-1, sheet_name="All Applicant Data") #reading file
    df_applicant_title_date = df_applicant_title.head(1).columns[0]
    ws.cell(2, 1, df_applicant_title_date ) 
 

def get_db_state ():
    conn = None
    params = config()
    
    conn = psycopg2.connect(**params )
    o_companycode=''
    o_policy =''
 
    cur = conn.cursor()
    
    
    SQL = ("""  SELECT    node_code as "Property ID",  node_state from node_tbl where company_code = 'BERK'
 """  )
  
    cur.execute(SQL) 
    col = [i[0] for i in cur.description]
            
    rows = cur.fetchall()

    cur.close 
    
    df_state = pd.DataFrame(rows, columns=col)


    return df_state
            
    #df = pd.DataFrame(rows, columns=["Node Code","Propery Name", "Average Individual Income","Average Household Income", "RVP", "PROPERTY TYPE", "applicant_uuid"])
    df_state = pd.DataFrame(rows, columns=col)

    
def get_state(wb, df_applicant):
 #try picture  
   
    ws = wb.get_sheet_by_name ('Average Income by State')
   
     
    df_row_state = get_db_state()
            
    #df = pd.DataFrame(rows, columns=["Node Code","Propery Name", "Average Individual Income","Average Household Income", "RVP", "PROPERTY TYPE", "applicant_uuid"])
    
    df_all = pd.merge(df_applicant, df_row_state ,on = ['Property ID'], how='inner')
    
    ''' 
    application_state_grp  = df_all.groupby(by =  ['node_state']).agg ( {
       "Application Monthly Income": ['mean' ], 
       "Rent To Income Ratio (%)":  ['mean' ]
     
       #"Rent To Income Ratio (%)": lambda x :  (x <= 1 and x >=998).mean() ,
       #"Application Monthly Income": lambda x : ( x >8.3 and x < 166667).mean()  
      
     }  ).reset_index()
    '''
    
   # df_applicant['Application Monthly Income'] = df_applicant['Application Monthly Income'].fillna(df_applicant['Application Monthly Income'].mean())
   # df_applicant['Rent To Income Ratio (%)'] = df_applicant['Rent To Income Ratio (%)'].fillna(df_applicant['Rent To Income Ratio (%)'].mean())
    
   # print (df_applicant['Application Monthly Income'].isna().sum() )
   # print (df_applicant['Rent To Income Ratio (%)'].isna().sum() )

    #df_applicant['Application Monthly Income'] = df_applicant['Application Monthly Income'].astype(object).fillna("UNKNOWN")
    #df_applicant['Rent To Income Ratio (%)'] = df_applicant['Rent To Income Ratio (%)'].astype(object).fillna("UNKNOWN")
    
    aggFunc = {'Application Monthly Income': np.nanmean,
           'Rent To Income Ratio (%)': np.nanmean,
           'Application ID' : np.count_nonzero
           }
    

    
    application_state_grp = pd.pivot_table(df_all,index=['node_state'] ,
                                    aggfunc=aggFunc,
                                    values=['Application Monthly Income','Rent To Income Ratio (%)', 'Application ID'],
                                    margins=True, margins_name='Grand Total' , dropna=False).reset_index()
    
    
     

    df_row = dataframe_to_rows(application_state_grp )  
    for index, row in enumerate(df_row, 1): 
            for c_idx, value in enumerate(row, 1):
                if (c_idx <=1 or index <=2):
                    continue 
                col = c_idx -1
                row_num = index  +1
                ws.cell(row_num, col).border = thin_border
                ws.cell(row_num, col, value=value) 
                ws.cell(row_num, col).alignment = Alignment(horizontal='center')  

                if (col ==3):
                    ws.cell(row_num,column=col).number_format = '"$"#,##0'   
                    ws.cell(row_num, col).alignment = Alignment(horizontal='center')          
                elif (col== 4):
                    ws.cell(row_num,column=col).number_format = '0'
                    ws.cell(row_num, col).alignment = Alignment(horizontal='center')      
                else: 
                    pass
                #else :
                    #ws.cell(row=index+start_row,column=c_idx).number_format = '0%'
    
    df_applicant_title = pd.read_excel(file_data, header=start_row-1, sheet_name="All Applicant Data") #reading file
    df_applicant_title_date = df_applicant_title.head(1).columns[0]
    ws.cell(2, 1, df_applicant_title_date ) 
 


def fix_ratio (x):
    if (x >=1 and x <= 998):
        return x
    else:
        return np.nan
 
def fix_incoming (x):
    if (x >=1000 and x <= 2000000):
        return x
    else:
        return np.nan


if __name__ == '__main__':
    
    wb = load_workbook(file_template)
    df_applicant = pd.read_excel(file_data, header=start_row, sheet_name="All Applicant Data") #reading file
    

    df_applicant = df_applicant.dropna (subset ='Property ID' ) 
    df_applicant = df_applicant [[ 'Property ID', 'Property Name' , 	'Credit Run',  'Applicant ID' ,'Application ID',   'Application Monthly Income', 'Rent To Income Ratio (%)'  ]] 
    


 
    df_applicant['Credit Run']=df_applicant['Credit Run'].apply(lambda x: str(x).upper()  )
    df_applicant_new = df_applicant.loc[df_applicant['Credit Run'] =="TRUE"]
    get_data(wb, df_applicant)
    wb.save(filename1+ ".xlsx" )
     
    df_applicant = df_applicant.replace('\n', ' ').replace('\r', ' ') 
    df_applicant = df_applicant.replace('\\n', ' ').replace('\\r', ' ') 

    df_applicant_new ['Application Monthly Income']= df_applicant_new ['Application Monthly Income']*12
    df_applicant_new.drop_duplicates( "Application ID", keep='last', inplace=True) 
    df_applicant_new["Rent To Income Ratio (%)"] = df_applicant_new["Rent To Income Ratio (%)"].apply(fix_ratio)
    df_applicant_new["Application Monthly Income"] = df_applicant_new["Application Monthly Income"].apply(fix_incoming)
    #df_applicant_new["Rent To Income Ratio (%)"] = df_applicant_new["Rent To Income Ratio (%)"].apply(lambda x: x if x>=1 and x<=998 else  np.NaN)
 
    #df_applicant_new.reset_index(inplace = True)
    
      
    get_prop(wb, df_applicant_new)
    get_state (wb,df_applicant_new )
   
    

    wb.save(filename1+ ".xlsx" )
     
    
sys.exit(0)
quit()
