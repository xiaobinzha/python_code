 
import psycopg2
from config import config
import datetime
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
 
from openpyxl.cell import Cell  
import pandas as pd
import numpy as np
import os
import sys
import logging
import functools
from pandas import DataFrame
from openpyxl.cell import Cell
from openpyxl.utils.dataframe import dataframe_to_rows
#from datetime import datetime
from openpyxl.styles import Color, Fill, Font, Border,Side, PatternFill
from openpyxl.styles import Alignment
 
from psycopg2.extensions import AsIs

from win32com import client #pypiwin32
 
import win32api
import pathlib
from functools import reduce
from xlutils.copy import copy #work xls xlutils
from xlrd import open_workbook
import xlwt
import shutil
from win32com.client import Dispatch
import tkinter as tk
from tkinter import simpledialog
from time import sleep
import easygui
import warnings
warnings.filterwarnings("ignore")
#from easygui import * 
import re
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

dirpath = os.getcwd() + "\\"
#import win32event

file_log = dirpath + "log.txt"
 
#from winerror import ERROR_ALREADY_EXISTS (change blue, grey, note, validation row numbers)
  
from easygui import passwordbox

try:
    #mutex = win32event.CreateMutex(None, False, 'name')
    #last_error = win32api.GetLastError(easygui)

    if  os.path.exists(file_log):
        output = easygui.msgbox("Process already started! ", "Error")
        sys.exit(0)
        quit(0)
		 
except:
    sys.exit(0)

password = passwordbox("PASSWORD:")

if (password == 'postgres'):
	pass 
else:
    output = easygui.msgbox("Wrong Password! ", "Error")
    sys.exit(0)
 
app = tk.Tk()
app.eval('tk::PlaceWindow . center')
app.geometry('250x250+500+500')
app.update_idletasks()
app.withdraw()
# the input dialog
USER_INP = simpledialog.askstring(title="Hi",
                                  prompt="Enter Client Code List Separated by Comma:   \t \t \t \t\t\t\t\t\t\t", parent=app)

 

if   USER_INP  is None :
    print ("NO Code")
    sys.exit(0)
    quit()
else:   
	list_code = list( [x.strip() for x in USER_INP.split(',')]) 

Node_INP = simpledialog.askstring(title="Hi",
                                  prompt="Enter Property Code List Separated by Comma:   \t \t \t \t\t\t\t\t\t\t", parent=app)

 
 
#USER_INP = USER_INP.string.strip()

if  Node_INP  is None :
	pass 
else: 
    list_node = list([x.strip() for x in Node_INP.split(',')])

 
MAXCOL = 110
#dirpath = "O:\ANALYTICS\Setup_Audit\"

  

d = datetime.datetime.today()    
END_offset = (d.weekday() - 5) % 7  
END_date =  (d - datetime.timedelta(days=END_offset)).strftime("%x")   
   
start_offset = ((d.weekday() - 6) % 7)+7
start_date = (d - datetime.timedelta(days=start_offset)).strftime("%x") 
    #END_date = ('1/9/2019');
END_date = d.strftime("%x");
    #start_date = ('1/9/2019');
 

     
file_temp =  dirpath +  '\\setup_template.xlsx'
 
file_temp = file_temp.replace('/','.',10)

 
choices = ["Default By Policy -- One Node One Policy","By Policy, Node Name -- All", "By Policy, Node Code -- All", "Cancel"]
msg = "Load application..."
title="Setup Audit"
reply = easygui.buttonbox(msg, title,  choices=choices)
  


def connect_section_1(cust_code):
    
    conn = None

    params = config()
    
    conn = psycopg2.connect(**params )
	#cust_code = str(cust_code).strip()
    
 
    cur = conn.cursor()
    o_SQL = (  """      
  
	SELECT 
		
		 --section 1
		    (trim(nt.policy)) as policy, nt.node_code, node_name template_name,   link_node, upper(trim(nt.company_code)) company_code,


		 --section 2 : 5
		
	  CASE WHEN nt.service_level ='1' THEN 'T1' 
		   WHEN nt.service_level ='2' THEN 'T2' 
		   WHEN nt.service_level ='3' THEN 'T3'
		   WHEN nt.service_level = '5' then 'TSE' 
	ELSE nt.service_level::text end AS totalscreen_level,
	 
	 CASE WHEN an.outgoing_notification=0 THEN 'None'
				 WHEN an.outgoing_notification=1 THEN 'Fax after each Service'
		   WHEN an.outgoing_notification=2 THEN 'Email after each Service'
							WHEN an.outgoing_notification=3 THEN 'Email after Screening Complete' --else   outgoing_notification 
			END AS Outgoing_response,
			
		suppress_voyager_page		as supress_details,
 case trim(node_suppress_validation_workflow) when 'Company' then 'Use Company Setting' when 'On' then 'Enabled for Property' when 'Off' then 'Disabled for Property' else node_suppress_validation_workflow end node_suppress_validation_workflow ,
 case trim(node_suppress_cancel_voi) when 'Company' then 'Use Company Setting' when 'On' then 'Enabled for Property' when 'Off' then 'Disabled for Property' else node_suppress_cancel_voi end node_suppress_cancel_voi,
			voi_manual_additional_income,
	--section 3 : 2
			
			
				CASE WHEN nt.allow_corp_applications='FALSE' THEN 'No'
				 WHEN nt.allow_corp_applications='TRUE' THEN 'Yes'
							ELSE NULL
							END AS allow_corporate_apps, 
		CASE WHEN an.biz_product=0 THEN 'None'
				 WHEN an.biz_product=1 THEN 'IntelliScore'
				 WHEN an.biz_product=2 THEN 'D&B'
							ELSE 'None'
							END as biz_product,
			
			
		--section 4 : 4
		
		CASE WHEN an.email_fcra_letter=1 THEN 'No'
				 WHEN an.email_fcra_letter=2 THEN 'Applicant and Property'

				 WHEN an.email_fcra_letter=3 THEN 'Property Only'
				 WHEN an.email_fcra_letter=4 THEN 'Applicant Only'
							ELSE NULL
							END as Email_FCRA_Letters_To, 
		an.fcra_email_addr as Adverse_Action_Email_Address,  
			
			case UPPER(CAST (fcra_batch_enabled   as text)) when 'TRUE' then 'Yes'
                   when 'FALSE' then 'No' end   fcra_batch_enabled  ,  -- 5/30/23
	case UPPER(CAST (an.allow_fcra_letter_generation as text)) when 'TRUE' then 'Yes' when 'FALSE' then 'No' end allow_fcra_letter_generation,
			
			
				--section 5 : 4
			UPPER(CAST (autoemail_fcraletter as text)) as Auto_Email_Adverse_Action_Letters, 
	 case UPPER(CAST (hold_until_active_app_processs as text)) when 'TRUE' then 'Yes' when 'FALSE' then 'No' end as Hold_Until_All_Active_Applicants_are_Processed,

      
	cast(send_fcraletter_after_hours as text) Email_FCRA_Letters_every, 
	 to_char(create_on , 'MM/DD/YYYY'::text) start_date,
	  
	 
	--section 6 :4

	case enable_fcra_template when 'comp' then 'Use Company Settings' when 'off' then 'Off' when 'system' then 'Use System Templates' else enable_fcra_template end as enable_template,  
	 UPPER(CAST (include_logo as text)) include_logo , 
	an.fcra_customized_text as Custom_Conditional_Text,  
	an.fcra_custom_reject_text as Custom_Reject_Text,  
    case when trim(an.tsr_email_delivery)='Use Company'  then  'Use Company Settings' else an.tsr_email_delivery end  sr_email_delivery,

             --- internal contact section
             account_manager,
                  
	--section 7 : 7
	
	cp.cred_prod_name AS credit_product,
		an.vax_policy AS "Credit Policy (CDS)",
		cast(an.delq_cutoff as text)  AS Delinq_Cutoff,

		UPPER(CAST (an.score_medical AS TEXT)) AS score_medical, 
		UPPER(CAST (an.score_student_loans AS TEXT)) AS score_student_loans, 
	  UPPER(CAST (an.private_owner AS TEXT)) AS Private_Owner_Passed_Res,
		
	 -----section 8 : 4

		UPPER(CAST (an.reject_apt_collect AS TEXT)) AS enable_apartment_filter, 
		case when an.reject_apt_collect = false then null else an.rac_num_months end   AS apt_filter_reject_months,
		case when an.reject_apt_collect = false then null else an.rac_max_amount end AS apt_filter_reject_$_value, 
		case when an.reject_apt_collect = false then null else an.rac_num_debts end AS apt_filter_reject_count,

	 ------section 9 :5 Utility Related
    UPPER(CAST (an.util_collect_exclude as  text)) util_exclude_from_credit_scoring,
	UPPER(CAST (an.util_collect_reject as text)) util_enable_rejection_criteria,
	
	case when util_collect_reject = false then null else an.util_collect_num_months end util_collect_num_months,
	case when util_collect_reject = false then null else an.util_collect_max_amount end util_collect_max_amount,
	case when util_collect_reject = false then null else an.util_collect_max_num end util_collect_max_num,


	 ------section 10 : 9 General Service Settings
	 
	 upper(cast(an.set_dob as text)) set_dob,
	 CASE WHEN an.show_prem_crim_res=0 THEN 'Never'
				 WHEN an.show_prem_crim_res=1 THEN 'Always'
		   WHEN an.show_prem_crim_res=2 THEN 'Does Not Meet Property Requirements'
		   WHEN an.show_prem_crim_res=3 THEN 'Inconclusive'
		   WHEN an.show_prem_crim_res=4 THEN 'Does Not Meet Property Requirements OR Inconclusive'
		   WHEN an.show_prem_crim_res=5 THEN 'Meets Property Requirements'
						ELSE NULL
						END AS "Show Criminal/Offense Results",
		CASE WHEN an.show_prem_evic_res=0 THEN 'Never'
			 WHEN an.show_prem_evic_res=1 THEN 'Always'
		   WHEN an.show_prem_evic_res=2 THEN 'Does Not Meet Property Requirements'
		   WHEN an.show_prem_evic_res=3 THEN 'Inconclusive'
		   WHEN an.show_prem_evic_res=4 THEN 'Does Not Meet Property Requirements OR Inconclusive'
		   WHEN an.show_prem_evic_res=5 THEN 'Meets Property Requirements'
						ELSE NULL
						END AS show_civil_court_results, 
		CASE WHEN an.recs_in_out_resp=0 THEN 'Never'
			 WHEN an.recs_in_out_resp=1 THEN 'Always'
		   WHEN an.recs_in_out_resp=2 THEN 'Does Not Meet Property Requirements'
		   WHEN an.recs_in_out_resp=3 THEN 'Inconclusive'
		   WHEN an.recs_in_out_resp=4 THEN 'Does Not Meet Property Requirements OR Inconclusive'
		   WHEN an.recs_in_out_resp=5 THEN 'Meets Property Requirements'
						ELSE NULL
						END AS records_in_outgoing_responses, 
		CASE WHEN an.create_linked_pcc IN (0,4) THEN 'Always'
			 WHEN an.create_linked_pcc IN (1,5) THEN 'Best Practice'
		   WHEN an.create_linked_pcc IN (2,6) THEN 'Never'
						ELSE NULL
						END AS create_supplemental_criminal_request,
		 
		UPPER(CAST(an.launch_supplemental AS TEXT)) AS enable_supplemental_vendor_criminal,
          no_suppl_crim_states as skip_supplemental_crim_search,
		UPPER(CAST((create_linked_pcc&4 > 0)AS TEXT)) as filter_only_for_criminal, 
	  UPPER(CAST (an.exclude_sex_offender AS TEXT)) AS exclude_sex_offender,
	  case  UPPER(CAST (an.rejection_reason_custom_criminal AS TEXT)) when 'TRUE' then 'Yes' when 'FALSE' then 'No' end AS rejection_reason_custom_criminal,
	 case when client_crim_pre_decision is null then 'No' 
	      when trim(client_crim_pre_decision ) =''  then 'No' 
          when trim(client_crim_pre_decision ) ='Yes'  then 'Criminal Only' 
          when trim(client_crim_pre_decision ) = 'All' then 'All Services'
            else  client_crim_pre_decision end as crim_record_assessment,

	 -----section 11 : 7 General Scoring
	 
		--UPPER(CAST((an.comp_recommendation&1>0) AS TEXT)) AS comprehensive_recommendation,
        UPPER (cast(enable_rent_to_income as TEXT))      as calculate_rti_ratio_wo_credit,  
		case when UPPER (cast(enable_rent_to_income as TEXT))  ='TRUE' then '' else an.default_recommendation end default_recommendation,
		case when UPPER (cast(enable_rent_to_income as TEXT))  ='TRUE' then '' else UPPER(CAST(an.gen_comp AS TEXT)) end AS generate_Comp_Score_wo_credit, 
		case when UPPER (cast(enable_rent_to_income as TEXT))  ='TRUE' then '' else an.score_without_credit end as default_score_wo_credit, 
             
             
		case an.hold_voyager_move_in when 'OFAC' then 'OFAC Messages' 
             when 'Checkpoint' then 'Checkpoint Messages' 
             when 'OR' then 'Checkpoint/OFAC Messages' 
             when 'Identity' then 'Identity Verification'
             else an.hold_voyager_move_in end as Review_Report_Acknowledgment,
 
	    case   coalesce(pandemic_era_civil_court_filter,'') when '0' then 'Disabled' when '1' then 'Enable Civil Court' when '2' then 'Enable Rental History' when '4' then 'Enable Both'
	 when '' then 'Disabled' when   null then 'Disabled'
	 else pandemic_era_civil_court_filter end Pandemic_Era_Filter, 
	 
	 pandemic_start_date, pandemic_end_date,
	 
	 -------section 12 : 2
	 nova_credit_scoring, nova_risk_score,
	 
	 ----section 13 :6 AAL Workflows
	 /*aal setting */
	 
	 case when crim_email='t' then 'Yes' else 'No' end   as  Email_Conditional_Offer_of_Pending_Housing, 
	case when crim_pre_aal = 'Yes' then 'Enabled' else 'Disabled' end as  PreAdverse_Action_Letter, 
	case   preaal_template_id when 36 then 'Pre-AAL Cook County' when 38 then 'Test_PreAAL' when 40 then 'Generic Pre-AAL' else null end  as  PreAdverse_Action_Letter_Template, 
	reconsider_request_period || ' ' ||  reconsider_period_in as Reconsideration_Request_Period, 
	reconsider_review_period || ' ' || reconsider_period_in as Reconsideration_Review_Period,
	reconsider_period_in,

	 ----section 14 : 12
	 (select max(vendor_name) from services  where ns.cred = service_id) credit_report,
	  case when coalesce(cred_tier,0)=0 then null else ns.cred_tier end cred_tier,

	   (select max(vendor_name) from services  where ns.civilcrt = service_id) civil_court,
	    case when coalesce(civilcrt_tier,0)=0 then null else ns.civilcrt_tier end civilcrt_tier,

		   (select max(vendor_name) from services  where ns.rhist = service_id) rental_history,
		   case when coalesce(rhist_tier,0)=0 then null else ns.rhist_tier end rhist_tier,

			  (select max(vendor_name) from services  where ns.ofac = service_id) as "OFAC",  
			  case when coalesce(ofac_tier,0)=0 then null else ns.ofac_tier end ofac_tier,

	   (select max(vendor_name) from services  where ns.crim = service_id) criminal,
	   case when coalesce(crim_tier,0)=0 then null else ns.crim_tier end crim_tier,


		 (select max(vendor_name) from services  where ns.cust_crim = service_id) custom_criminal, 
		      case when coalesce(cust_crim_tier,0)=0 then null else cust_crim_tier end cust_crim_tier, 
             
          (select max(vendor_name) from services  where ns.offense = service_id) offensealert, 
		      case when coalesce(offense_tier,0)=0 then null else offense_tier end Offensealert_tier, 

              (select max(vendor_name) from services  where ns.voi = service_id)  voi, 
		      case when coalesce(voi_tier,0)=0 then null else voi_tier end voi_tier,     
             
	-----section 15 :6 ondemand services

	   (select max(vendor_name) from services  where ns.cred_ondmd = service_id)   on_demand_credit,
	(select max(vendor_name) from services  where ns.civilcrt_ondmd = service_id)   on_demand_civil_court,
	  (select max(vendor_name) from services  where ns.rhist_ondmd = service_id)   on_demand_rental,
	(select max(vendor_name) from services  where ns.ofac_ondmd = service_id)   "On Demand OFAC",
	  (select max(vendor_name) from services  where ns.crim_ondmd = service_id)   on_demand_crim,
	(select max(vendor_name) from services  where ns.cust_crim_ondmd = service_id)   on_demand_cust_crim,	 
    (select max(vendor_name) from services  where ns.offense_ondmd = service_id)   on_demand_offensealert,	 
            
       (select max(vendor_name) from services  where ns.voi_ondmd = service_id)  on_demand_voi,	 

	----section 16 :6 On Demand - Additional Functions

	 case cast(ns.reeval_func  as text)  when '19' then 'Yes' when  '0' then 'No'     end  on_demand_reevaluation_request,
	  case cast(ns.sbond_func  as text)  when '20' then 'Yes' when  '0' then 'No'     end  on_demand_sure_deposit,
   case  an.enable_prequal_screening when 'On' then 'Enabled For Property' when 'Company' then 'Use Company Setting' else an.enable_prequal_screening end enable_prequal_screening,
      
             
             -------new section income verification
 case when trim(lower(an.enable_work_number)) = 'ssv' then  ssv_num_months   end ssv_num_of_month, 

  
  case coalesce(enable_dollar_validation_2,'')
           when 'On' then 'Enable with Rejection'
           when 'Rejection Pending Workflow' then 'Enable with Pending Workflow'
           when 'Off' then 'Off' when '' then 'Off' end --end 
         as enable_dollar_valication,   ---3
             
      
 case when  enable_voi_notification_emails = true then 'Yes' else 'No' end 
             as send_notification_emails   ,  
	  
             case when coalesce(skip_voi,'f')= 't' then 'On' else 'Off' end skip_voi ,  skip_voi_for_riskscore,
             
             an.enable_voi_embedded, --6
     
	
      ----------------   service verification serivce -------------------------
      
     case when ts_service_tier = 'On' then 'Enable'  when ts_service_tier = 'Off' then 'Disable'     end ts_selected_tier, ts_selected_tier, ts_on_demand ,
             
             case when coalesce( vs_ssv_num_months ,0) =3 then '3 Months'
   when coalesce( vs_ssv_num_months ,0) =6 then '6 Months'
when coalesce( vs_ssv_num_months ,0) =12 then '1 Year'
when coalesce( vs_ssv_num_months ,0) =36 then '3 Year'
when coalesce( vs_ssv_num_months ,0) =1000 then 'All Records' 
else null end as vs_ssv_num_months,
     case when py_service_tier = 'On' then 'Enable'  when py_service_tier = 'Off' then 'Disable'    end py_service_tier , py_selected_tier,  py_on_demand ,
      case when vs_service_tier = 'On' then 'Enable'  when vs_service_tier = 'Off' then 'Disable'    end vs_service_tier , vs_selected_tier,  vs_on_demand ,   
         
     case when upper(asset_verify_additional_data)='YES' then 'Yes' 
             when upper(asset_verify_additional_data)='NO'  then 'No' else asset_verify_additional_data end,
     vs_num_of_paystubs,
      
      ---------------end of notification service 
		
             --------service configration,
             
             enable_vs_service_link as show_service_link, enable_manual_sync,
	---section 17 :2

	/*Group Scoring */
		CASE 	WHEN gsm.method_description ISNULL THEN gsm2.method_description
										ELSE gsm.method_description 
										END AS group_scoring, 
		CASE 	WHEN gsm.method_description ISNULL THEN gsuco.method_parameter_1
										ELSE gsu.method_parameter_1
										END AS scoring_table, 
	------case 18 :2

		UPPER(CAST((credit_report&1 > 0) AS TEXT)) AS suppress_details, 
	UPPER(CAST((credit_report&4 > 0) AS TEXT)) AS suppress_Items_for_Review,

	------section 19 :23

	UPPER(CAST((an.rental_history&4>0) AS TEXT)) AS show_tenant_information,
	  UPPER(CAST((an.rental_history&8>0) AS TEXT)) AS show_collections,
			UPPER(CAST((an.rental_history&16>0) AS TEXT)) AS show_statement,
 				 

	Case when rental_history=159 then 'TRUE' when rental_history=157 then 'TRUE'  when rental_history=158  then 'FALSE'  when rental_history=156 then 'FALSE' else 'FALSE' end AS  retrieve_rentPredict_score,  
	Case when rental_history=159 then 'TRUE' when rental_history=158 then 'TRUE'  when rental_history=157  then 'FALSE'  when rental_history=156 then 'FALSE' else 'FALSE' end AS display_rentPredict_score,  
       
	case     coalesce(an.rental_history,0)   
	when 0 then 'Off'
   when 4 then 'Off'
   when 8 then 'Off'
   when 12 then 'Off'
   when 16 then 'Off'
   when 20 then 'Off'
   when 24 then 'Off'
   when 28 then 'Off'
   when 31 then 'Off'
   when 64 then 'Off'
   when 68 then 'Off'
   when 76 then 'Off'
   when 80 then 'Off'
   when 84 then 'Off'
   when 92 then 'Off'
   when 128 then 'On'
   when 132 then 'On'
   when 136 then 'On'
   when 140 then 'On'
   when 144 then 'On'
   when 148 then 'On'
   when 156 then 'On'
   when 157 then 'On'
   when 159 then 'On'
   when 192 then 'On'
   when 196 then 'On'
   when 200 then 'On'
   when 204 then 'On'
   when 208 then 'On'
   when 212 then 'On'
   when 216 then 'On'
   when 220 then 'On'
   when 222 then 'On'
   when 223 then 'On'
   when null then 'Off' end as rental_history,
 
						 UPPER(CAST((an.rental_history&64>0) AS TEXT)) AS show_reason, 
           
		UPPER(CAST((rhs.late_payments_scoring) AS TEXT)) AS late_payments_scoring, 
		
		rhs.late_payments_max, 
		rhs.late_payments_period, 

	    UPPER(CAST(rhs.nsf_scoring AS TEXT)) AS nsf_scoring, 
		rhs.nsf_max, 
		rhs.nsf_period, 
		UPPER(CAST(rhs.outstanding_balances_scoring AS TEXT)) AS outstanding_balances_scoring, 
		rhs.outstanding_balances_max, 
		rhs.outstanding_balances_period, 
		UPPER(CAST(rhs.write_offs_scoring AS TEXT)) AS write_offs_scoring, 
		rhs.write_offs_amount, 
		rhs.write_offs_period, 
		UPPER(CAST(rhs.collections_scoring AS TEXT)) AS collections_scoring, 
		rhs.collections_amount, 
		rhs.collections_period,
		
		--------section 20 :14
		case yardi_interface when 1 then 'One-Way Interface' when 2 then 'Two-Way Interface' end as interface_functionality,
	(select display_value  from list_values where list_id =1 and "index" = pms_if limit 1) "PMS Interface",
	 (select display_value  from list_values where list_id =2 and "index" = oli_if limit 1) Online_Leasing_Interface,
		CASE	WHEN an.disable_edit_applicant=0 THEN 'Do Not Disable'
					WHEN an.disable_edit_applicant=1 THEN 'Edit Applicant'
					WHEN an.disable_edit_applicant=2 THEN 'Enter New Applicant'
					WHEN an.disable_edit_applicant=3 THEN 'Edit and Enter New Applicant'
						ELSE NULL
						END AS disable_edit_applicant,
			UPPER(CAST(an.enable_group_apps AS TEXT)) AS enable_group_apps,
		UPPER(CAST(include_checkpoint_msg as text)) include_checkpoint_msg,

		ylp.interface_username, ylp.password as psd,
		ylp.server_name AS database_server_name, 
		ylp.database_name, 
		ylp.platform, 
		ylp.interface_entity, 
		ylp.web_service_uri AS "URL", 
		case when coalesce(sso_enabled,'false') = 'false' then 'No' else 'Yes' end sso_enabled,

		--------section 21							

		 rs.a_active,  rs.a_startbreakpoint, rs.a_endbreakpoint,    rs.b_active, rs.b_startbreakpoint, rs.b_endbreakpoint,
	     rs.c_active, rs.c_startbreakpoint, rs.c_endbreakpoint,    rs.d_active, rs.d_startbreakpoint, rs.d_endbreakpoint, 
	    rs.g_active, rs.g_startbreakpoint, rs.g_endbreakpoint, 
		
		CASE WHEN  rs.r_active is not null THEN rs.r_active else   rs.f_active END r_active,
        CAST(CASE WHEN rs.r_startbreakpoint is not null  THEN   rs.r_startbreakpoint  else rs.f_startbreakpoint END  AS TEXT) as r_startbreakpoint, 
	    CAST( CASE WHEN rs.r_endbreakpoint is not null  THEN   rs.r_endbreakpoint  else rs.f_endbreakpoint END AS TEXT) as r_endbreakpoint, 
        
		def_credit_score as no_risk,
        
		 rsp.a_active  ap_active, CAST(rsp.a_startbreakpoint AS TEXT) ap_startpoint,  CAST(rsp.a_endbreakpoint AS TEXT) ap_endpoint,
	     rsp.b_active  bp_active, CAST(rsp.b_startbreakpoint AS TEXT) bp_startpoint,  CAST(rsp.b_endbreakpoint AS TEXT) bp_endpoint,
	     rsp.c_active  cp_active, CAST(rsp.c_startbreakpoint AS TEXT) cp_startpoint,  CAST(rsp.c_endbreakpoint AS TEXT) cp_endpoint,
         rsp.d_active  dp_active, CAST(rsp.d_startbreakpoint AS TEXT)  dp_startpoint,  CAST(rsp.d_endbreakpoint AS TEXT) dp_endpoint,
 	     rsp.g_active  gp_active, CAST(rsp.g_startbreakpoint AS TEXT) gp_startpoint,  CAST(rsp.g_endbreakpoint AS TEXT) gp_endpoint,
	    
		CASE WHEN  rsp.r_active  is not null THEN rsp.r_active else rsp.f_active END rp_active, 
        CAST( CASE WHEN  rsp.r_startbreakpoint  is not null THEN rsp.r_startbreakpoint else rsp.f_startbreakpoint END AS TEXT) rp_startpoint,
	    CAST( CASE WHEN  rsp.r_endbreakpoint is not null THEN    rsp.r_endbreakpoint else rsp.f_endbreakpoint END AS TEXT) rp_endpoint ,


		--------section 22
		case  when enable_new_suredep_v =false then null else case surety_bond_vENDor when 0 THEN 'Not Set' when 1 THEN 'SureDeposit' when 2 THEN 'DepositIQ' END end Surety_Bond_Vendor ,
		case  when enable_new_suredep_v =false then null else case sure_deposit_layout when 0 THEN 'Normal' when 1 THEN 'Property Configured for Overrides' END end Surety_Bond_Layout ,
		sure_deposit_settings as company_default_scoring,  -------??
 
		sdm.sd_risk_level as a_score, b_score, c_score, d_score,e_score,   f_score, g_score, r_score, p_score, i_score,
			  
		--section 23
		an.notes, nt.create_stamp as node_create_date  --concat('"', an.notes, '"') notes, 
 
	FROM node_tbl nt
	LEFT JOIN account_tbl as BA ON nt.node_code = BA.account_code
	LEFT JOIN auxiliary_node an ON nt.node_code=an.node_code
	LEFT JOIN credit_products cp ON an.cred_prod_id=cp.cred_prod_id
	LEFT JOIN group_score_users gsu ON gsu.node_or_company_code=nt.node_code
	LEFT JOIN group_score_methods gsm ON gsu.method=gsm.method
	LEFT JOIN group_score_users gsuco ON gsuco.node_or_company_code=nt.company_code /* We join the group scoring tables again by company for those nodes that take on the company score*/
	LEFT JOIN group_score_methods gsm2 ON gsuco.method=gsm2.method
	LEFT JOIN auxiliary_company ac ON nt.company_code=ac.company_code
	LEFT JOIN rhs_policies rhs ON nt.node_code=rhs.policy_code
	LEFT JOIN yardi_login_parameters ylp ON nt.node_code=ylp.node_code
	LEFT JOIN (select primary_node, array_agg(link_node)::text link_node from b2b_node  where active<>'N' and primary_node <> link_node  group by primary_node) Link_Node on primary_node = nt.node_code 
	LEFT JOIN ( SELECT comp_code, node_code, autoemail_fcraletter, hold_until_active_app_processs, send_fcraletter_after_hours, fcraletter_start_date::date as create_on    FROM autogenerate_fcraletter_config --where node_code = 'K7956'
	) fc on fc.node_code = nt.node_code and fc.comp_code = nt.company_code
	LEFT JOIN node_services ns on    trim(ns.pr_code) = trim(nt.node_code) and trim(ns.co_code) = trim(nt.company_code)
    LEFT JOIN(
							SELECT 
								rpd.node_code,
								CASE 	WHEN start_break_point ISNULL THEN ''
											ELSE 'TRUE' 
		END AS a_active, 	
								rpd.start_break_point AS a_startbreakpoint,
								rpd.end_break_point AS a_endbreakpoint, 
								CASE    WHEN brpd.b_startbreakpoint ISNULL THEN ''
											ELSE 'TRUE' 
		END AS b_active, 	
								brpd.b_startbreakpoint,
								brpd.b_endbreakpoint,
								CASE 	WHEN crpd.c_startbreakpoint ISNULL THEN ''
											ELSE 'TRUE' 
		END AS c_active, 	
								crpd.c_startbreakpoint,
								crpd.c_endbreakpoint,
								CASE 	WHEN drpd.d_startbreakpoint ISNULL THEN ''
											ELSE 'TRUE' 
		END AS d_active,
								drpd.d_startbreakpoint,
								drpd.d_endbreakpoint,
								CASE 	WHEN frpd.f_startbreakpoint ISNULL THEN ''
											ELSE 'TRUE' 
		END AS f_active,
								frpd.f_startbreakpoint,
								frpd.f_endbreakpoint,
								CASE 	WHEN grpd.g_startbreakpoint ISNULL THEN ''
											ELSE 'TRUE' 
		END AS g_active,
								grpd.g_startbreakpoint,
								grpd.g_endbreakpoint,
								CASE 	WHEN rrpd.r_startbreakpoint ISNULL THEN ''
											ELSE 'TRUE' 
		END AS r_active,
								rrpd.r_startbreakpoint,
								rrpd.r_endbreakpoint
							FROM 	risk_score_policy_details rpd
							LEFT JOIN(
													SELECT 
														node_code,
														CASE 	WHEN start_break_point ISNULL THEN ''
																	ELSE 'TRUE' 
		END AS b_active,
														start_break_point AS b_startbreakpoint,
														end_break_point AS b_endbreakpoint	
													FROM 	risk_score_policy_details
													WHERE score = 'B'
													) brpd ON brpd.node_code=rpd.node_code
							LEFT JOIN(
													SELECT 
														node_code,
														CASE 	WHEN start_break_point ISNULL THEN ''
																	ELSE 'TRUE' 
		END AS c_active,
														start_break_point AS c_startbreakpoint,
														end_break_point AS c_endbreakpoint	
													FROM 	risk_score_policy_details
													WHERE score = 'C'
													) crpd ON crpd.node_code=rpd.node_code
							LEFT JOIN(
													SELECT 
														node_code,
														CASE 	WHEN start_break_point ISNULL THEN ''
																	ELSE 'TRUE' 
		END AS d_active,
														start_break_point AS d_startbreakpoint,
														end_break_point AS d_endbreakpoint	
													FROM 	risk_score_policy_details
													WHERE score = 'D'
													) drpd ON drpd.node_code=rpd.node_code
							LEFT JOIN(
													SELECT 
														node_code,
														CASE 	WHEN start_break_point ISNULL THEN ''
																	ELSE 'TRUE' 
		END AS f_active,
														start_break_point AS f_startbreakpoint,
														end_break_point AS f_endbreakpoint	
													FROM 	risk_score_policy_details
													WHERE score = 'F'
													) frpd ON frpd.node_code=rpd.node_code
							LEFT JOIN(
													SELECT 
														node_code,
														CASE 	WHEN start_break_point ISNULL THEN ''
																	ELSE 'TRUE' 
		END AS g_active,
														start_break_point AS g_startbreakpoint,
														end_break_point AS g_endbreakpoint	
													FROM 	risk_score_policy_details
													WHERE score = 'G'

													) grpd ON grpd.node_code=rpd.node_code
							LEFT JOIN(
													SELECT 
														node_code,
														CASE 	WHEN start_break_point ISNULL THEN ''
																	ELSE 'TRUE' 
		END AS r_active,
														start_break_point AS r_startbreakpoint,
														end_break_point AS r_endbreakpoint	
													FROM 	risk_score_policy_details
													WHERE score = 'R'
													) rrpd ON rrpd.node_code=rpd.node_code

							WHERE score = 'A'
							) AS rs ON rs.node_code=nt.node_code
		left join (

		SELECT 
								rpd.node_code,
								CASE 	WHEN start_break_point ISNULL THEN ''
											ELSE 'TRUE' 
		END AS a_active, 	
								rpd.start_break_point AS a_startbreakpoint,
								rpd.end_break_point AS a_endbreakpoint, 
								CASE    WHEN brpd.b_startbreakpoint ISNULL THEN ''
											ELSE 'TRUE' 
		END AS b_active, 	
								brpd.b_startbreakpoint,
								brpd.b_endbreakpoint,
								CASE 	WHEN crpd.c_startbreakpoint ISNULL THEN ''
											ELSE 'TRUE' 
		END AS c_active, 	
								crpd.c_startbreakpoint,
								crpd.c_endbreakpoint,
								CASE 	WHEN drpd.d_startbreakpoint ISNULL THEN ''
											ELSE 'TRUE' 
		END AS d_active,
								drpd.d_startbreakpoint,
								drpd.d_endbreakpoint,
								CASE 	WHEN frpd.f_startbreakpoint ISNULL THEN ''
											ELSE 'TRUE' 
		END AS f_active,
								frpd.f_startbreakpoint,
								frpd.f_endbreakpoint,
								CASE 	WHEN grpd.g_startbreakpoint ISNULL THEN ''
											ELSE 'TRUE' 
		END AS g_active,
								grpd.g_startbreakpoint,
								grpd.g_endbreakpoint,
								CASE 	WHEN rrpd.r_startbreakpoint ISNULL THEN ''
											ELSE 'TRUE' 
		END AS r_active,
								rrpd.r_startbreakpoint,
								rrpd.r_endbreakpoint
							FROM 	group_risk_score_policy_details rpd
							LEFT JOIN(
													SELECT 
														node_code,
														CASE 	WHEN start_break_point ISNULL THEN ''
																	ELSE 'TRUE' 
		END AS b_active,
														start_break_point AS b_startbreakpoint,
														end_break_point AS b_endbreakpoint	
													FROM 	group_risk_score_policy_details
													WHERE score = 'B'
													) brpd ON brpd.node_code=rpd.node_code
							LEFT JOIN(
													SELECT 
														node_code,
														CASE 	WHEN start_break_point ISNULL THEN ''
																	ELSE 'TRUE' 
		END AS c_active,
														start_break_point AS c_startbreakpoint,
														end_break_point AS c_endbreakpoint	
													FROM 	group_risk_score_policy_details
													WHERE score = 'C'
													) crpd ON crpd.node_code=rpd.node_code
							LEFT JOIN(
													SELECT 
														node_code,
														CASE 	WHEN start_break_point ISNULL THEN ''
																	ELSE 'TRUE' 
		END AS d_active,
														start_break_point AS d_startbreakpoint,
														end_break_point AS d_endbreakpoint	
													FROM 	group_risk_score_policy_details
													WHERE score = 'D'
													) drpd ON drpd.node_code=rpd.node_code
							LEFT JOIN(
													SELECT 
														node_code,
														CASE 	WHEN start_break_point ISNULL THEN ''
																	ELSE 'TRUE' 
		END AS f_active,
														start_break_point AS f_startbreakpoint,
														end_break_point AS f_endbreakpoint	
													FROM 	group_risk_score_policy_details
													WHERE score = 'F'
													) frpd ON frpd.node_code=rpd.node_code
							LEFT JOIN(
													SELECT 
														node_code,
														CASE 	WHEN start_break_point ISNULL THEN ''
																	ELSE 'TRUE' 
		END AS g_active,
														start_break_point AS g_startbreakpoint,
														end_break_point AS g_endbreakpoint	
													FROM 	group_risk_score_policy_details
													WHERE score = 'G'

													) grpd ON grpd.node_code=rpd.node_code
							LEFT JOIN(
													SELECT 
														node_code,
														CASE 	WHEN start_break_point ISNULL THEN ''
																	ELSE 'TRUE' 
		END AS r_active,
														start_break_point AS r_startbreakpoint,
														end_break_point AS r_endbreakpoint	
													FROM 	group_risk_score_policy_details
													WHERE score = 'R'
													) rrpd ON rrpd.node_code=rpd.node_code

							WHERE score = 'A' ) rsp on rsp.node_code = nt.node_code 

		left join (select * from sure_deposit_score_map where yrs_score = 'A') sdm   on nt.node_code = sdm.node_code	 	
		LEFT JOIN ( SELECT 	node_code, sd_risk_level as b_score	FROM 	sure_deposit_score_map WHERE yrs_score = 'B' and COALESCE(node_code,'')<>''   ) bsdm ON bsdm.node_code=sdm.node_code 
		LEFT JOIN  ( SELECT 	node_code, sd_risk_level as c_score	FROM 	sure_deposit_score_map WHERE yrs_score = 'C' and COALESCE(node_code,'')<>''   ) csdm ON csdm.node_code=sdm.node_code
		LEFT JOIN ( SELECT 	node_code, sd_risk_level as d_score	FROM 	sure_deposit_score_map WHERE yrs_score = 'D' and COALESCE(node_code,'')<>''   ) dsdm ON dsdm.node_code=sdm.node_code 
		LEFT JOIN  ( SELECT 	node_code, sd_risk_level as e_score	FROM 	sure_deposit_score_map WHERE yrs_score = 'E' and COALESCE(node_code,'')<>''   ) esdm ON esdm.node_code=sdm.node_code

		LEFT JOIN ( SELECT 	node_code, sd_risk_level as f_score	FROM 	sure_deposit_score_map WHERE yrs_score = 'F' and COALESCE(node_code,'')<>''   ) fsdm ON fsdm.node_code=sdm.node_code 
		LEFT JOIN  ( SELECT 	node_code, sd_risk_level as g_score	FROM 	sure_deposit_score_map WHERE yrs_score = 'G' and COALESCE(node_code,'')<>''   ) gsdm ON gsdm.node_code=sdm.node_code
		LEFT JOIN ( SELECT 	node_code, sd_risk_level as r_score	FROM 	sure_deposit_score_map WHERE yrs_score = 'R' and COALESCE(node_code,'')<>''   ) rsdm ON rsdm.node_code=sdm.node_code 
		LEFT JOIN  ( SELECT 	node_code, sd_risk_level as p_score	FROM 	sure_deposit_score_map WHERE yrs_score = 'P' and COALESCE(node_code,'')<>''   ) psdm ON psdm.node_code=sdm.node_code
		LEFT JOIN  ( SELECT 	node_code, sd_risk_level as i_score	FROM 	sure_deposit_score_map WHERE yrs_score = 'I' and COALESCE(node_code,'')<>''   ) isdm ON isdm.node_code=sdm.node_code
		LEFT JOIN (select node_code, service_tier as ts_service_tier,  selected_tier ts_selected_tier, on_demand as ts_on_demand
            from vs_node_services where  service_code = 'VS_TWNSSV') ts_service on nt.node_code = ts_service.node_code
          
        LEFT JOIN (select node_code, service_tier as py_service_tier,  selected_tier py_selected_tier, on_demand as py_on_demand
            from vs_node_services where  service_code = 'VS_REQ_PSTB') py_service on nt.node_code = py_service.node_code

        LEFT JOIN (select node_code, service_tier as vs_service_tier,  selected_tier vs_selected_tier, on_demand as vs_on_demand
            from vs_node_services where  service_code = 'VS_ASSET_VER') vs_service on nt.node_code = vs_service.node_code

             
	WHERE  nt.canceled <> 'T'  and   --and nt.node_code in ('B6768' ) --'B6768'X5716, B6768
     nt.company_code  =   '%s'   """ % (cust_code))
 
    #print (SQL)
     
    if Node_INP  is None:
        SQL = o_SQL
    elif (Node_INP == '*' or Node_INP == ''):
        SQL = o_SQL  
    elif (len(list_node) <=0):
        SQL = o_SQL  
    else:
        #SQL = str(o_SQL).replace("WHERE nt.canceled <> 'T'", "WHERE nt.canceled <> 'T' and nt.node_code in (" + ','.join( x for x in  list_node)  + ')' )
        string_replace = "WHERE nt.canceled <> 'T' and upper(nt.node_code) in ('" + "','".join(name.upper().strip().replace("'", r"\'") for name in list_node) + "'"  + ')'
        SQL =  (o_SQL).replace("WHERE nt.canceled <> 'T'", string_replace  )
     
  

    cur.execute(SQL   )
	#col = cur.description 
    # 
    # 
    rows = cur.fetchall()

	
    col = [i[0] for i in cur.description]
    
    df = pd.DataFrame(rows, columns=col)
    
    df = df.replace(np.nan, '', regex=True)
   
    df['default_recommendation'] = df['default_recommendation'].str.replace( '+',   ' '  )  

   
    df['default_recommendation'] = df['default_recommendation'].str.replace('%2B' , ' ') 
    df['default_recommendation'] = df['default_recommendation'].str.replace('%252B',' ') 
    df['default_recommendation'] = df['default_recommendation'].str.replace('%09',  ' ') 
    df['default_recommendation'] = df['default_recommendation'].str.replace('%27',  "'" )  
    df['default_recommendation'] = df['default_recommendation'].str.replace('%2F', '/' ) 
 
    df['default_recommendation'] = df['default_recommendation'].str.replace('%28' , '(' ) 
    df['default_recommendation'] = df['default_recommendation'].str.replace('%29' , ')' ) 
    df['default_recommendation'] = df['default_recommendation'].str.replace('%24' , '$' ) 
    df['default_recommendation'] = df['default_recommendation'].str.replace('%26' , '&' ) 
    df['default_recommendation'] = df['default_recommendation'].str.replace('%2C' , ',' ) 
    df['default_recommendation'] = df['default_recommendation'].str.replace('%21' , '!' ) 
    df['default_recommendation'] = df['default_recommendation'].str.replace('%3D' , '=' ) 
    df['default_recommendation'] = df['default_recommendation'].str.replace('%3A' , ':' ) 
    df['default_recommendation'] = df['default_recommendation'].str.replace('%40' , '@' ) 
    df['default_recommendation'] = df['default_recommendation'].str.replace('%3F' , '?' ) 
    df['default_recommendation'] = df['default_recommendation'].str.replace('%7E' , '~' ) 
    

    df['voi'] = df['voi'].str.replace(r'^VOI$' , 'VOI (TWN Disabled)' , regex=True)   
    df['on_demand_voi'] = df['on_demand_voi'].str.replace(r'^VOI$', 'VOI (TWN Disabled)',regex=True ) 
    df['voi'] = df['voi'].str.replace(r'^WDocumentVerification$' , 'Document Verification' , regex=True)   
    df['on_demand_voi'] = df['on_demand_voi'].str.replace(r'^WDocumentVerification$', 'Document Verification',regex=True )   
     
    df['send_notification_emails'] = np.where((df.voi == '') & (df.on_demand_voi == '' ), '',df['send_notification_emails'] )
    df['send_notification_emails'] = np.where((df.voi == 'The Work Number SSV') & (df.on_demand_voi == 'The Work Number SSV' ), '',df['send_notification_emails']  )
    df['send_notification_emails'] = np.where((df.voi == 'The Work Number SSV') & (df.on_demand_voi == '' ), '',df['send_notification_emails']  )
    df['send_notification_emails'] = np.where((df.voi == '') & (df.on_demand_voi == 'The Work Number SSV' ), '',df['send_notification_emails']  )
    #df['rental_history'] = df['rental_history'].replace([128,132,136,140,144,148,156,157,159,192,196,200,204,208,212,216,220,222,223],'On')
    #df['rental_history'] = df['rental_history'].replace([0,4,8,12,16,20,24,28,31,64,68,76,80,84,92,''],'Off')
    
	#df['notes'] = df['notes'].replace(regex=True, to_replace=r'[^0-9A-z \-\n\t\r,:.\\;^~%$#@&\*\'\(\)<>\?\/]', value=r'' ) 
    #print ('a')
 
		 
	#print (df.sec_15_credit_report) 
    #choices = ["Default By Policy -- One Node One Policy","By Policy, Node Name -- All ", "By Policy, Node Code -- All", "Cancel"]
    df.replace( np.nan, 'n/a',inplace = True)
    df.replace( '', 'n/a', inplace=True)
    #df_1 =df   #inplace = True,'company_code',

    if (reply == "Default By Policy -- One Node One Policy"): 
        df.sort_values(by=[ 'company_code', 'policy', 'node_create_date'],   ascending=True, inplace=True ) #, 'node_create_date'
        df = df.drop_duplicates ([ 'company_code', 'policy',], keep = 'first' )
  
    elif (reply == "By Policy, Node Code -- All"):
        df.sort_values(by=['company_code', 'policy', 'node_code'] ,  ascending = True,   inplace=True  )
	  
    elif (reply == "By Policy, Node Name -- All"):
        df.sort_values(by=[ 'company_code', 'policy', 'template_name'] ,  ascending = True, inplace=True) #'node_name',
    elif (reply == "Cancel"):
        sys.exit(0)
        quit()
    
    df = df.drop (['node_create_date'], axis = 1)
    
    return (df)        
    cur.close()
    conn.close()

def column_num_to_string(n):
    n, rem = divmod(n - 1, 26)
    next_char = chr(65 + rem)
    if n:
        return column_num_to_string(n) + next_char
    else:
        return next_char  

def format_file(df, sheet):
    try:
        sbuf1 = '=group_scoring'
        dv = DataValidation(type="list", formula1= '=' + sbuf1 , allow_blank=True, showDropDown=False)
        dv.add('C117:XFD117') ######## change 1
        sheet.add_data_validation(dv)
        sbuf2 = '=servicelink'
        dv_2 = DataValidation(type="list", formula1= '=' + sbuf1 , allow_blank=True, showDropDown=False)
        dv_2.add('C115:XFD115') ######## change 1
        sheet.add_data_validation(dv_2)
        sbuf3 = '=manualsync'
        dv_3 = DataValidation(type="list", formula1= '=' + sbuf1 , allow_blank=True, showDropDown=False)
        dv_3.add('C116:XFD116') ######## change 1
        sheet.add_data_validation(dv_3)
        

    
        rows = dataframe_to_rows(df) 
        thin_border = Border(left=Side(style='medium'), 
                        right=Side(style='medium'), 
                        top=Side(style='medium'), 
                        bottom=Side(style='medium'))
        

		######## change 2
        grey_numbers = [6,7,8,9,10,11,14,15,16,17,22,23,24,25,27,34,35,36,37,43,44,45,46,47,48,50,51,52,53,62,63,70,71,72,73,74,75,76,77,78,79,80,81,82,83,84,85,94,95,96,103,104,105,106,107,108,109,110,111,112,113,114,117,118,143,144,145,146,147,148,149,150,151,152,153,154,155,156,194,195,196,197,198,199,200,201,202,203,204,205,206]
        greyFill =PatternFill(fill_type='solid',start_color='D9D9D9',end_color='D9D9D9') 
    
        blue_numbers = [128,129,130,134,135,136,140,141,142,160,161,162,166,167,168,172,173,174,176,177,178,182,183,184,188,189,190]
        blueFill =  PatternFill(fill_type='solid',start_color='D9E1F2',end_color='D9E1F2') 
        node_start_row = 207 ######## change 3

        for r_idx, row in enumerate(rows, 1):         
            for c_idx, value_1 in enumerate(row, 1):
                # print (c_idx)
                if r_idx >=2 and c_idx>=2 :
                    try:
                        sheet.cell(row=r_idx-2, column=c_idx-1+2, value=value_1)
                    except Exception as e:
                       # value_1 = value_1.replace(regex=True, to_replace=r'[^0-9A-z \-\n\t\r,:.\\;^~%$#@&\*\'\(\)<>\?\/]', value=r'' ) 
                        #value_1 = re.sub(r"[^0-9A-z \-\n\t\r,:.\\;\^~%\$#@&\+\*\(\)<>\?\[\]\{\}]\|]", '', value_1) #.[{ () \^$| ? *+
                        value_1 = re.sub(r"[^0-9A-z \-\n\t\r,:.\\;\^~%\$#@&\+\*\(\<>\?\[\]) \{\}\|\]( ]", '', value_1)  
                        sheet.cell(row=r_idx-2, column=c_idx-1+2, value=value_1)
                    finally:
                        sheet.cell(row=r_idx-2, column=c_idx-1+2).border = thin_border
                        sheet.cell(row=r_idx-2, column=c_idx-1+1).border = thin_border 
                        sheet.cell(row=r_idx-2, column=c_idx-1+2).font = Font(bold=False,  name="Calibri", size=12) #color="FFFFFF",
                        sheet.cell(row=r_idx-2, column=c_idx-1+2).alignment = Alignment(horizontal='left', vertical = 'center')
                if r_idx-2 in grey_numbers:
                    sheet.cell(row=r_idx-2, column=c_idx-1+2).fill = greyFill
                elif r_idx -2 in blue_numbers:
                    sheet.cell(row=r_idx-2, column=c_idx-1+2).fill = blueFill
                 
                elif r_idx -2 == node_start_row:
                    sheet.cell(row=r_idx-2, column=c_idx-1+2).alignment = Alignment(wrapText=True)
                    sheet.merge_cells(column_num_to_string(c_idx-1+2) + str(node_start_row) + ':' + column_num_to_string(c_idx-1+2)  + str(node_start_row+3))
                    for i in range (c_idx-1 , c_idx +2) :
                        if (i>0):
                            sheet.cell(row=node_start_row+1, column=i).border = thin_border
                            sheet.cell(row=node_start_row+2, column=i).border = thin_border
                            sheet.cell(row=node_start_row+3, column=i).border = thin_border

       # sheet.set_column(3, c_idx-1+2, 50)
        for i in range  (3, c_idx-1+3 ):
            sheet.column_dimensions[column_num_to_string(i)].width = 50                	 
    except Exception as e:
            print (e )
    finally:
            print ('')
				      
                
def save_workbook (sname):
    excel = Dispatch('Excel.Application')
    wb = excel.Workbooks.Open(s_name)
     
    getNumSheet = wb.Worksheets.count+1

    for i in range  (1, getNumSheet ):
        excel.Worksheets(i).Activate() 
    excel.ActiveSheet.Columns.AutoFit()
        
#Save changes in a new file
    excel.DisplayAlerts = False    
    wb.SaveAs (s_name)
    wb.Close()

##sheet1.title = s_name 
try: 
    f = open(file_log, "x")
    f.flush()
    f.close()
    sleep(1)
    for cust_code in list_code: 
        cust_code = cust_code.strip().upper()
        s_name =   dirpath + cust_code + ' Setup Audit ' +  str(END_date)  + '.xlsx'
        s_name = s_name.replace('/','.',10)
        df1 = connect_section_1(cust_code)
        df1 = df1.T # or df1.transpose() 

        wb = load_workbook(file_temp)
        sheet1 = wb.get_sheet_by_name('Sheet1')
        sheet1.title = cust_code
        format_file (df1, sheet1) 
    
        wb.save(s_name)
       # sleep(1)
      #  save_workbook(s_name)
		
 
except Exception as e:
		print("There were  errors-" , e)
finally:
   # win32event.ReleaseMutex(mutex)
   if  os.path.exists(file_log):
       os.remove(file_log)
   # mutex.close()
   sys.exit(0)
   quit()
