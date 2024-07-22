 
from tkinter import FIRST
import psycopg2
 
import datetime
from datetime import date, timedelta,datetime
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.cell import Cell  
import pandas as pd
import numpy as np
import os
import sys
import logging
import locale
from openpyxl.styles import Alignment
import dateutil.relativedelta
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Color, Fill, Font, Border,Side, PatternFill
from openpyxl.styles import colors
import openpyxl 
from openpyxl import Workbook
 
from win32com import client
 
import pathlib
from functools import reduce
import win32api
from shutil import copyfile
from config import config
from config2 import config2
 
 
from dateutil.relativedelta import relativedelta
from win32com.client import Dispatch
import win32com.client as win32
import codecs

 
 
 
from openpyxl.styles import Color, Fill, Font, Border,Side 
from openpyxl.cell import Cell  
import pandas as pd
import pyexcelerate 
 
 


 
#dirpath = "O:\ANALYTICS\Setup_Audit\"
dirpath = os.getcwd()

file_template = dirpath +    '\\QA_Template.xlsx' 

dirpath='O:\\ANALYTICS\\Setup_Audit\\Setup Audit Weekly Reports\\' 
#dirpath='C:\\Temp\\' 
log = "Setup Audit.log"
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# create a file handler
handler = logging.FileHandler(log,mode='w')
handler.setLevel(logging.INFO)

# create a logging format
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)

# add the file handler to the logger
logger.addHandler(handler)
logger.info(' Log Entry Start Here.'  )
 

today =  date.today()
idx =  (today.weekday() + 1) % 7
w_start_date = sys.argv[1] if len(sys.argv) > 1 else today -  timedelta(7+idx)
w_end_date =  sys.argv[2] if len(sys.argv) > 2 else w_start_date + timedelta(days=6)

w_start_date = w_start_date.strftime("%m/%d/%Y")  
w_end_date=w_end_date.strftime("%m/%d/%Y")  
filename1 =   dirpath +  'Setup Audit Weekly ' + str(w_start_date).replace('/','.',10) + '-' + w_end_date.replace('/','.',10)  +  '.xlsx'


s_name_compare =  'Compare ' +  str(w_start_date).replace('/','.')[0:5]   + '-' +  str(w_end_date).replace('/','.')[0:5]
 
 

    #filename1 =  'template.xlsx '

print (filename1)

def compare_sql():
    SQL =  ("""    set client_encoding to 'SQL_ASCII' ; 
SELECT '' Master,  
            nt.company_code,
	nt.policy AS screening_policy,
nt.node_code,
 username as "Property Added By",

  CASE WHEN nt.service_level ='1' THEN 'T1' 
       WHEN nt.service_level ='2' THEN 'T2' 
       WHEN nt.service_level ='3' THEN 'T3' 
       WHEN nt.service_level = '5' then 'TSE' 
ELSE nt.service_level::text end AS totalscreen_level,

CASE WHEN an.outgoing_notification=0 THEN 'None'
			 WHEN an.outgoing_notification=1 THEN 'Fax'
       WHEN an.outgoing_notification=2 THEN 'Email'
						ELSE NULL       
		END AS Outgoing_response,
	CASE WHEN nt.allow_corp_applications='FALSE' THEN 'No'
			 WHEN nt.allow_corp_applications='TRUE' THEN 'Yes'
						ELSE NULL
						END AS allow_corporate_apps, 
	CASE WHEN an.biz_product=0 THEN 'None'
			 WHEN an.biz_product=1 THEN 'IntelliScore'
			 WHEN an.biz_product=2 THEN 'D&B'
						ELSE 'None'
						END as biz_product,
	CASE WHEN an.email_fcra_letter=1 THEN 'No'
			 WHEN an.email_fcra_letter=2 THEN 'Applicant and Property'

			 WHEN an.email_fcra_letter=3 THEN 'Property Only'
			 WHEN an.email_fcra_letter=4 THEN 'Applicant Only'
						ELSE NULL
						END as Email_FCRA_Letters_To, 
  	 
UPPER(CAST (fcra_batch_enabled   as text)) fcra_batch_enabled  , 
UPPER(CAST (show_late_payment_letter as text)) show_late_payment_letter,
UPPER(CAST (allow_custom_fcra_reasons as text)) allow_custom_fcra_reasons,
UPPER(CAST (an.allow_fcra_letter_generation as text)) allow_fcra_letter_generation,

--UPPER(CAST (autoemail_fcraletter as text)) autoemail_fcraletter, 
--UPPER(CAST (hold_until_active_app_processs as text)) hold_until_active_app_processs, 
--send_fcraletter_after_hours Email_FCRA_Letters_every, 

cp.cred_prod_name AS credit_product,
	an.vax_policy AS CDS_policy,
	an.delq_cutoff AS Delinq_Cutoff,

	UPPER(CAST (an.score_medical AS TEXT)) AS score_medical, 
	UPPER(CAST (an.score_student_loans AS TEXT)) AS score_student_loans, 
  --UPPER(CAST (an.private_owner AS TEXT)) AS Private_Owner_Passed_Res, --4/7/2023
	UPPER(CAST (an.reject_apt_collect AS TEXT)) AS enable_apartment_filter, 
	CASE WHEN reject_apt_collect ='TRUE' AND an.rac_num_months ISNULL THEN 	(	SELECT CAST (resource_value AS INT)
																																						FROM resources 
																																						WHERE resource_group = 'AptCollectionFilter' AND resource_key='DefaultMonths4Reject'
																																					) 	/*Must account for system defaults */
					ELSE an.rac_num_months 
					END AS apt_filter_reject_months,
	CASE WHEN reject_apt_collect ='TRUE' AND an.rac_max_amount ISNULL THEN (	SELECT CAST (resource_value AS INT)
																																						FROM resources 
																																						WHERE resource_group = 'AptCollectionFilter' AND resource_key='DefaultAmount4Reject'
																																					) 	/*Must account for system defaults */
					ELSE an.rac_max_amount
					END AS apt_filter_reject_$_value, 
	CASE WHEN reject_apt_collect ='TRUE' AND an.rac_num_debts ISNULL THEN  (	SELECT CAST (resource_value AS INT)
																																						FROM resources 
																																						WHERE resource_group = 'AptCollectionFilter' AND resource_key='DefaultAlwaysRejectQuantity'
																																					)		/*Must account for system defaults */
					ELSE an.rac_num_debts
					END AS apt_filter_reject_count,
UPPER(CAST (an.util_collect_reject as text)) util_enable_rejection_criteria,
UPPER(CAST (an.util_collect_exclude as  text)) util_exclude_from_credit_scoring,
an.util_collect_num_months,
an.util_collect_max_amount,
an.util_collect_max_num,
CASE WHEN an.show_prem_crim_res=0 THEN 'Never'
			 WHEN an.show_prem_crim_res=1 THEN 'Always'
       WHEN an.show_prem_crim_res=2 THEN 'Does Not Meet Property Requirements'
       WHEN an.show_prem_crim_res=3 THEN 'Inconclusive'
       WHEN an.show_prem_crim_res=4 THEN 'Does Not Meet Property Requirements OR Inconclusive'
       WHEN an.show_prem_crim_res=5 THEN 'Meets Property Requirements'
					ELSE NULL
					END AS show_criminaloffense_alert_results,
	CASE WHEN an.show_prem_evic_res=0 THEN 'Never'
	     WHEN an.show_prem_evic_res=1 THEN 'Always'
       WHEN an.show_prem_evic_res=2 THEN 'Does Not Meet Property Requirements'
       WHEN an.show_prem_evic_res=3 THEN 'Inconclusive'
       WHEN an.show_prem_evic_res=4 THEN 'Does Not Meet Property Requirements OR Inconclusive'
       WHEN an.show_prem_evic_res=5 THEN 'Meets Property Requirements'
					ELSE NULL
					END AS show_civil_court_results, 
	CASE WHEN an.recs_in_out_resp=0 THEN 'Never'
	     WHEN an.recs_in_out_resp=1 THEN 'Always'
       WHEN an.recs_in_out_resp=2 THEN 'Does Not Meet Property Requirements'
       WHEN an.recs_in_out_resp=3 THEN 'Inconclusive'
       WHEN an.recs_in_out_resp=4 THEN 'Does Not Meet Property Requirements OR Inconclusive'
       WHEN an.recs_in_out_resp=5 THEN 'Meets Property Requirements'
					ELSE NULL
					END AS records_in_outgoing_responses, 
	CASE WHEN an.create_linked_pcc IN (0,4) THEN 'Always'
	     WHEN an.create_linked_pcc IN (1,5) THEN 'Best Practice'
       WHEN an.create_linked_pcc IN (2,6) THEN 'Never'
					ELSE NULL
					END AS create_supplemental_criminal_request,
	 
	UPPER(CAST(an.launch_supplemental AS TEXT)) AS enable_supplemental_criminal,
   no_suppl_crim_states as Skip_Supplemental_Searches_for_States, -- added 4/7

	UPPER(CAST((create_linked_pcc&4 > 0)AS TEXT)) as filter_only_for_criminal, 
  UPPER(CAST (an.exclude_sex_offender AS TEXT)) AS exclude_sex_offender,
 UPPER(CAST (an.rejection_reason_custom_criminal AS TEXT)) AS rejection_reason_custom_criminal,

client_crim_pre_decision as crim_record_assessment, --4/13/2023
UPPER(CAST (an.crim_email AS TEXT)) AS crim_email,
	UPPER(CAST((an.comp_recommendation&1>0) AS TEXT)) AS comprehensive_recommendation,
--	an.default_recommendation,
replace( replace( 
replace(  
replace(replace( replace( replace( 
replace( replace( replace(replace(
replace(replace(replace(replace(
replace( replace(replace(an.default_recommendation,
      '+',   ' '), 
     '%%2B' , ' '),
     '%%252B',' '), 
     '%%09',  ' '),
     '%%27',  '''' ), 
     '%%2F', '/' ),
     '%%2F' , '/' ),
  '%%28' , '(' ),
  '%%29' , ')' ),
  '%%24' , '$' ),
  '%%26' , '&' ),
  '%%2C' , ',' ),
  '%%21' , '!' ),
  '%%3D' , '=' ),
  '%%3A' , ':' ),
  '%%40' , '@' ),
  '%%3F' , '?' ),
  '%%7E' , '~' ) default_recommendation,

	UPPER(CAST(an.gen_comp AS TEXT)) AS generate_comp_score_wo_credit, 
	an.fa_credit_score AS no_ssn_score, 
	an.score_without_credit, 
/* Automatic Service */
	--tsc.ts_credit_report, 6/20/18
 case when coalesce(cred_tier,0)=0 then null else ns.cred_tier end cred_tier, 
(select max(vendor_name) from services  where ns.cred = service_id) credit_report,


 case when coalesce(civilcrt_tier,0)=0 then null else ns.civilcrt_tier end civilcrt_tier,
   (select max(vendor_name) from services  where ns.civilcrt = service_id) civil_court,


 case when coalesce(rhist_tier,0)=0 then null else ns.rhist_tier end rhist_tier,
  (select max(vendor_name) from services  where ns.rhist = service_id) rental_history,

case when coalesce(ofac_tier,0)=0 then null else ns.ofac_tier end ofac_tier,
 (select max(vendor_name) from services  where ns.ofac = service_id) ofac,  

 
case when coalesce(crim_tier,0)=0 then null else ns.crim_tier end crim_tier,
   (select max(vendor_name) from services  where ns.crim = service_id) criminal,

	--cc.custom_criminal, 
       case when coalesce(cust_crim_tier,0)=0 then null else cust_crim_tier end cust_crim_tier,     
(select max(vendor_name) from services  where ns.cust_crim = service_id) custom_criminal, 

--	ccsa.custom_criminal_submitafter,

 case when coalesce(offense_tier,0)=0 then null else offense_tier end offense_tier,    
  (select max(vendor_name) from services  where ns.offense = service_id) offense_alert, 

 -- voi
 (select max(vendor_name) from services  where ns.voi = service_id) "VOI", 
		      case when coalesce(voi_tier,0)=0 then null else voi_tier end "VOI Tier",     
 (select max(vendor_name) from services  where ns.voi_ondmd = service_id)   "On Demand VOI",
  case coalesce(trim(enable_dollar_validation_2),'')
           when 'On' then 'Enable with Rejection'
           when 'Rejection Pending Workflow' then 'Enable with Pending Workflow'
           when 'Off' then 'Off' when '' then 'Off' end 
  as enable_dollar_valication,   ---3

 
 case when  enable_voi_notification_emails = true then 'Yes' else 'No' end   as send_notification_emails   ,  ---4

             an.enable_voi_embedded as "Enable VOI Embedded", 


  case when trim(lower(an.enable_work_number)) = 'ssv' then  ssv_num_months   end ssv_num_of_month, 
-- voi

   (select max(vendor_name) from services  where ns.cred_ondmd = service_id)   od_credit,
(select max(vendor_name) from services  where ns.civilcrt_ondmd = service_id)   od_civil_court,
  (select max(vendor_name) from services  where ns.rhist_ondmd = service_id)   od_rental,
(select max(vendor_name) from services  where ns.ofac_ondmd = service_id)   od_ofac,
  (select max(vendor_name) from services  where ns.crim_ondmd = service_id)   od_crim,
(select max(vendor_name) from services  where ns.cust_crim_ondmd = service_id)   od_cust_crim,
(select max(vendor_name) from services  where ns.offense_ondmd = service_id)   od_Offense,

	(case when ns.reeval_func=19 then 'On' else null end) od_reevaluation_request,
	(case when ns.sbond_func=20 then 'On' else null end) od_sure_deposit,

  
/*Group Scoring */
	CASE 	WHEN gsm.method_description ISNULL THEN gsm2.method_description
									ELSE gsm.method_description 
									END AS group_scoring, 
	CASE 	WHEN gsm.method_description ISNULL THEN gsuco.method_parameter_1
									ELSE gsu.method_parameter_1
									END AS scoring_table, 

/*Credit Report Options*/
	UPPER(CAST((credit_report&1 > 0) AS TEXT)) AS suppress_details, 
UPPER(CAST((credit_report&4 > 0) AS TEXT)) AS Suppress_Items_for_Review,
 
 --UPPER(CAST (include_logo as text)) include_logo , --4/13/2023
 
        UPPER(CAST((an.rental_history&4>0) AS TEXT)) AS show_tenant_information,
        UPPER(CAST((an.rental_history&8>0) AS TEXT)) AS show_collections,
        UPPER(CAST((an.rental_history&16>0) AS TEXT)) AS show_statement,
UPPER(CAST((an.rental_history&128>0) AS TEXT)) AS  rental_history_scoring, --RHscoring from NodeInfo where node_code = 'RGDE1';
	UPPER(CAST((an.rental_history&64>0) AS TEXT)) AS show_reasons,
	UPPER(CAST((rhs.late_payments_scoring) AS TEXT)) AS late_payments_scoring, 
	rhs.late_payments_max, 
	rhs.late_payments_period, 
	UPPER(CAST(rhs.nsf_scoring AS TEXT)) AS nsf_scoring, 
	rhs.nsf_max, 
	rhs.nsf_period, 
	UPPER(CAST(rhs.outstanding_balances_scoring AS TEXT)) AS outstanding_balances_scoring, 
	rhs.outstanding_balances_max, 
	rhs.outstanding_balances_period, 
	UPPER(CAST(rhs.write_offs_scoring AS TEXT)) AS write_offs_scoring, 
	rhs.write_offs_amount, 
	rhs.write_offs_period, 
	UPPER(CAST(rhs.collections_scoring AS TEXT)) AS collections_scoring, 
	rhs.collections_amount, 
	rhs.collections_period,

/*Interface Options */
/*new addition */
case yardi_interface when 1 then 'One-Way Interface' when 2 then 'Two-Way Interface' end as interface_functionality,
--(select display_value  from list_values where list_id =1 and "index" = pms_if limit 1) PMS_Interface,
 --(select display_value  from list_values where list_id =2 And "index" = oli_if limit 1) Online_Leasing_Interface,
 	--CASE	WHEN an.disable_edit_applicant=0 THEN 'Do Not Disable'
		--		WHEN an.disable_edit_applicant=1 THEN 'Edit Applicant'
			--	WHEN an.disable_edit_applicant=2 THEN 'Enter New Applicant'
			--	WHEN an.disable_edit_applicant=3 THEN 'Edit and Enter New Applicant'
				--	ELSE NULL
				--	END AS disable_edit_applicant,

  	UPPER(CAST(include_checkpoint_msg as text)) include_checkpoint_msg,
	--UPPER(CAST(an.enable_group_apps AS TEXT)) AS enable_group_apps,
	--ylp.interface_username, 
	--ylp.server_name AS database_server_name, 
	--ylp.database_name, 
	--ylp.platform, 
	--ylp.interface_entity, 
	--ylp.web_service_uri AS url, 
/* Risk Score Options */
	rs.a_active, 
	rs.a_startbreakpoint, 
	rs.a_endbreakpoint,
	rs.b_active, 
	rs.b_startbreakpoint,
	rs.b_endbreakpoint,
	rs.c_active, 
	rs.c_startbreakpoint,
	rs.c_endbreakpoint, 
	rs.d_active, 
	rs.d_startbreakpoint, 
	rs.d_endbreakpoint, 
	rs.f_active,
	rs.f_startbreakpoint, 
	rs.f_endbreakpoint, 
	rs.g_active, 
	rs.g_startbreakpoint, 
	rs.g_endbreakpoint, 
	rs.r_active, 
	rs.r_startbreakpoint, 
	rs.r_endbreakpoint,
	an.def_credit_score  AS no_risk_score_available,
nova_credit_scoring, nova_risk_score,
 case when crim_email='t' then 'Yes' else 'No' end   as  "Email Conditional Offer of Pending Housing", ---4/7/2023
case when crim_pre_aal = 'Yes' then 'Enabled' else 'Disabled' end as  PreAdverse_Action_Letter, 
case   preaal_template_id when 36 then 'Pre-AAL Cook County' when 38 then 'Test_PreAAL'  when 40 then 'Generic Pre-AAL' else '' end  as  PreAdverse_Action_Letter_Template, 

reconsider_request_period || ' ' ||  reconsider_period_in as Reconsideration_Request_Period, 
reconsider_review_period || ' ' || reconsider_period_in as Reconsideration_Review_Period,
 case when coalesce(an.b2b_screening_workflow,false)= false then 'Off' else 'On' end as Alternate_Screening_Criteria_Rules--, 
--to_char(nt.create_stamp , 'MM/DD/YYYY'::text)   || ' ' || to_char(nt.create_stamp, 'HH24:MI:SS'::text) AS setup_date

	 
FROM node_tbl nt
LEFT JOIN account_tbl as BA ON nt.node_code = BA.account_code
LEFT JOIN auxiliary_node an ON nt.node_code=an.node_code
LEFT JOIN credit_products cp ON an.cred_prod_id=cp.cred_prod_id
LEFT JOIN group_score_users gsu ON gsu.node_or_company_code=nt.node_code
LEFT JOIN group_score_methods gsm ON gsu.method=gsm.method
LEFT JOIN group_score_users gsuco ON gsuco.node_or_company_code=nt.company_code /* We join the group scoring tables again by company for those nodes that take on the company score*/
LEFT JOIN group_score_methods gsm2 ON gsuco.method=gsm2.method
LEFT JOIN auxiliary_company ac ON nt.company_code=ac.company_code
LEFT JOIN rhs_policies rhs ON nt.node_code=rhs.policy_code
LEFT JOIN yardi_login_parameters ylp ON nt.node_code=ylp.node_code
 

 
LEFT JOIN (
					SELECT 
						rpd.node_code, 
						CASE 	WHEN start_break_point ISNULL THEN 'FALSE'
									ELSE 'TRUE' 
									END AS a_active, 	
						rpd.start_break_point AS a_startbreakpoint, 
						rpd.end_break_point AS a_endbreakpoint, 
						CASE 	WHEN brpd.b_startbreakpoint ISNULL THEN 'FALSE'
									ELSE 'TRUE' 
									END AS b_active, 	
						brpd.b_startbreakpoint,
						brpd.b_endbreakpoint, 
						CASE 	WHEN crpd.c_startbreakpoint ISNULL THEN 'FALSE'
									ELSE 'TRUE' 
									END AS c_active, 	
						crpd.c_startbreakpoint,
						crpd.c_endbreakpoint, 
						CASE 	WHEN drpd.d_startbreakpoint ISNULL THEN 'FALSE'
									ELSE 'TRUE' 
									END AS d_active,
						drpd.d_startbreakpoint,
						drpd.d_endbreakpoint, 
						CASE 	WHEN frpd.f_startbreakpoint ISNULL THEN 'FALSE'
									ELSE 'TRUE' 
									END AS f_active,
						frpd.f_startbreakpoint,
						frpd.f_endbreakpoint, 
						CASE 	WHEN grpd.g_startbreakpoint ISNULL THEN 'FALSE'
									ELSE 'TRUE' 
									END AS g_active,
						grpd.g_startbreakpoint,
						grpd.g_endbreakpoint, 
						CASE 	WHEN rrpd.r_startbreakpoint ISNULL THEN 'FALSE'
									ELSE 'TRUE' 
									END AS r_active,
						rrpd.r_startbreakpoint,
						rrpd.r_endbreakpoint
					FROM 	risk_score_policy_details rpd
					LEFT JOIN 	(
											SELECT 
												node_code, 
												CASE 	WHEN start_break_point ISNULL THEN 'FALSE'
															ELSE 'TRUE' 
															END AS b_active,
												start_break_point AS b_startbreakpoint, 
												end_break_point AS b_endbreakpoint	
											FROM 	risk_score_policy_details
											WHERE score = 'B'
											) brpd ON brpd.node_code=rpd.node_code
					LEFT JOIN 	(
											SELECT 
												node_code,  
												CASE 	WHEN start_break_point ISNULL THEN 'FALSE'
															ELSE 'TRUE' 
															END AS c_active,
												start_break_point AS c_startbreakpoint, 
												end_break_point AS c_endbreakpoint	
											FROM 	risk_score_policy_details
											WHERE score = 'C'
											) crpd ON crpd.node_code=rpd.node_code
					LEFT JOIN 	(
											SELECT 
												node_code,  
												CASE 	WHEN start_break_point ISNULL THEN 'FALSE'
															ELSE 'TRUE' 
															END AS d_active,
												start_break_point AS d_startbreakpoint, 
												end_break_point AS d_endbreakpoint	
											FROM 	risk_score_policy_details
											WHERE score = 'D'
											) drpd ON drpd.node_code=rpd.node_code
					LEFT JOIN 	(
											SELECT 
												node_code, 
												CASE 	WHEN start_break_point ISNULL THEN 'FALSE'
															ELSE 'TRUE' 
															END AS f_active,
												start_break_point AS f_startbreakpoint, 
												end_break_point AS f_endbreakpoint	
											FROM 	risk_score_policy_details
											WHERE score = 'F'
											) frpd ON frpd.node_code=rpd.node_code
					LEFT JOIN 	(
											SELECT 
												node_code, 
												CASE 	WHEN start_break_point ISNULL THEN 'FALSE'
															ELSE 'TRUE' 
															END AS g_active,
												start_break_point AS g_startbreakpoint, 
												end_break_point AS g_endbreakpoint	
											FROM 	risk_score_policy_details
											WHERE score = 'G'
											 
											) grpd ON grpd.node_code=rpd.node_code
					LEFT JOIN 	(
											SELECT 
												node_code, 
												CASE 	WHEN start_break_point ISNULL THEN 'FALSE'
															ELSE 'TRUE' 
															END AS r_active,
												start_break_point AS r_startbreakpoint, 
												end_break_point AS r_endbreakpoint	
											FROM 	risk_score_policy_details
											WHERE score = 'R'
											) rrpd ON rrpd.node_code=rpd.node_code
      
					WHERE score = 'A'
					) AS rs ON rs.node_code=nt.node_code
         LEFT JOIN (
                      SELECT comp_code, node_code, autoemail_fcraletter, hold_until_active_app_processs, send_fcraletter_after_hours, created_on::date 
                        FROM autogenerate_fcraletter_config --where node_code = 'K7956'

             ) fc on fc.node_code = nt.node_code and fc.comp_code = nt.company_code
             LEFT JOIN node_services ns on    trim(ns.pr_code) = trim(nt.node_code) and trim(ns.co_code) = trim(nt.company_code)
   LEFT Join (select who, string_agg(username, ',') username from audit_tbl  where lower(change_details) like '%%added new property%%' group by who ) audit on audit.who = nt.node_code 
WHERE nt.canceled <> 'T' and nt.create_stamp::DATE >= '%s' AND  nt.create_stamp::DATE <= '%s'
        -- and nt.node_code in ('EH826',	'EH827',	'EH829',	'EH746',	'EH754',	'EH756')
                  order by nt.company_code,  nt.policy, nt.create_stamp,nt.node_code   ;    """ % (w_start_date,  w_end_date))

    return SQL
    
 
def setup_sql():
     sql = ( """

SELECT 
	/* General */
       nt.company_code,
   nt.policy AS screening_policy,
  CASE WHEN nt.service_level ='1' THEN 'T1' 
       WHEN nt.service_level ='2' THEN 'T2' 
       WHEN nt.service_level ='3' THEN 'T3' 
       WHEN nt.service_level = '5' then 'TSE' 
ELSE nt.service_level::text end AS totalscreen_level,
 
CASE WHEN an.outgoing_notification=0 THEN 'None'
			 WHEN an.outgoing_notification=1 THEN 'Fax'
       WHEN an.outgoing_notification=2 THEN 'Email'
						ELSE NULL       
		END AS Outgoing_response,
	CASE WHEN nt.allow_corp_applications='FALSE' THEN 'No'
			 WHEN nt.allow_corp_applications='TRUE' THEN 'Yes'
						ELSE NULL
						END AS allow_corporate_apps, 
	CASE WHEN an.biz_product=0 THEN 'None'
			 WHEN an.biz_product=1 THEN 'IntelliScore'
			 WHEN an.biz_product=2 THEN 'D&B'
						ELSE 'None'
						END as biz_product,
	CASE WHEN an.email_fcra_letter=1 THEN 'No'
			 WHEN an.email_fcra_letter=2 THEN 'Applicant and Property'

			 WHEN an.email_fcra_letter=3 THEN 'Property Only'
			 WHEN an.email_fcra_letter=4 THEN 'Applicant Only'
						ELSE NULL
						END as Email_FCRA_Letters_To, 
  	an.fcra_email_addr, 
UPPER(CAST (fcra_batch_enabled   as text)) fcra_batch_enabled  , 
UPPER(CAST (show_late_payment_letter as text)) show_late_payment_letter,
UPPER(CAST (allow_custom_fcra_reasons as text)) allow_custom_fcra_reasons,
UPPER(CAST (an.allow_fcra_letter_generation as text)) allow_fcra_letter_generation,

UPPER(CAST (autoemail_fcraletter as text)) autoemail_fcraletter, 
UPPER(CAST (hold_until_active_app_processs as text)) hold_until_active_app_processs, 
send_fcraletter_after_hours Email_FCRA_Letters_every, 
created_on start_date,

cp.cred_prod_name AS credit_product,
	an.vax_policy AS CDS_policy,
	an.delq_cutoff AS Delinq_Cutoff,

	UPPER(CAST (an.score_medical AS TEXT)) AS score_medical, 
	UPPER(CAST (an.score_student_loans AS TEXT)) AS score_student_loans, 
  --UPPER(CAST (an.private_owner AS TEXT)) AS Private_Owner_Passed_Res, --4/7/23
	UPPER(CAST (an.reject_apt_collect AS TEXT)) AS enable_apartment_filter, 
	CASE WHEN reject_apt_collect ='TRUE' AND an.rac_num_months ISNULL THEN 	(	SELECT CAST (resource_value AS INT)
																																						FROM resources 
																																						WHERE resource_group = 'AptCollectionFilter' AND resource_key='DefaultMonths4Reject'
																																					) 	/*Must account for system defaults */
					ELSE an.rac_num_months 
					END AS apt_filter_reject_months,
	CASE WHEN reject_apt_collect ='TRUE' AND an.rac_max_amount ISNULL THEN (	SELECT CAST (resource_value AS INT)
																																						FROM resources 
																																						WHERE resource_group = 'AptCollectionFilter' AND resource_key='DefaultAmount4Reject'
																																					) 	/*Must account for system defaults */
					ELSE an.rac_max_amount
					END AS apt_filter_reject_$_value, 
	CASE WHEN reject_apt_collect ='TRUE' AND an.rac_num_debts ISNULL THEN  (	SELECT CAST (resource_value AS INT)
																																						FROM resources 
																																						WHERE resource_group = 'AptCollectionFilter' AND resource_key='DefaultAlwaysRejectQuantity'
																																					)		/*Must account for system defaults */
					ELSE an.rac_num_debts
					END AS apt_filter_reject_count,
UPPER(CAST (an.util_collect_reject as text)) util_enable_rejection_criteria,
UPPER(CAST (an.util_collect_exclude as  text)) util_exclude_from_credit_scoring,
an.util_collect_num_months,
an.util_collect_max_amount,
an.util_collect_max_num,
CASE WHEN an.show_prem_crim_res=0 THEN 'Never'
			 WHEN an.show_prem_crim_res=1 THEN 'Always'
       WHEN an.show_prem_crim_res=2 THEN 'Does Not Meet Property Requirements'
       WHEN an.show_prem_crim_res=3 THEN 'Inconclusive'
       WHEN an.show_prem_crim_res=4 THEN 'Does Not Meet Property Requirements OR Inconclusive'
       WHEN an.show_prem_crim_res=5 THEN 'Meets Property Requirements'
					ELSE NULL
					END AS show_criminaloffense_alert_results,
	CASE WHEN an.show_prem_evic_res=0 THEN 'Never'
	     WHEN an.show_prem_evic_res=1 THEN 'Always'
       WHEN an.show_prem_evic_res=2 THEN 'Does Not Meet Property Requirements'
       WHEN an.show_prem_evic_res=3 THEN 'Inconclusive'
       WHEN an.show_prem_evic_res=4 THEN 'Does Not Meet Property Requirements OR Inconclusive'
       WHEN an.show_prem_evic_res=5 THEN 'Meets Property Requirements'
					ELSE NULL
					END AS show_civil_court_results, 
	CASE WHEN an.recs_in_out_resp=0 THEN 'Never'
	     WHEN an.recs_in_out_resp=1 THEN 'Always'
       WHEN an.recs_in_out_resp=2 THEN 'Does Not Meet Property Requirements'
       WHEN an.recs_in_out_resp=3 THEN 'Inconclusive'
       WHEN an.recs_in_out_resp=4 THEN 'Does Not Meet Property Requirements OR Inconclusive'
       WHEN an.recs_in_out_resp=5 THEN 'Meets Property Requirements'
					ELSE NULL
					END AS records_in_outgoing_responses, 
	CASE WHEN an.create_linked_pcc IN (0,4) THEN 'Always'
	     WHEN an.create_linked_pcc IN (1,5) THEN 'Best Practice'
       WHEN an.create_linked_pcc IN (2,6) THEN 'Never'
					ELSE NULL
					END AS create_supplemental_criminal_request,
	 
	UPPER(CAST(an.launch_supplemental AS TEXT)) AS enable_supplemental_criminal,
   no_suppl_crim_states as Skip_Supplemental_Searches_for_States, -- added 4/7
	UPPER(CAST((create_linked_pcc&4 > 0)AS TEXT)) as filter_only_for_criminal, 
  UPPER(CAST (an.exclude_sex_offender AS TEXT)) AS exclude_sex_offender,
 UPPER(CAST (an.rejection_reason_custom_criminal AS TEXT)) AS rejection_reason_custom_criminal,
client_crim_pre_decision as crim_record_assessment, -- 4/13/2023
UPPER(CAST (an.crim_email AS TEXT)) AS crim_email,
	UPPER(CAST((an.comp_recommendation&1>0) AS TEXT)) AS comprehensive_recommendation,
 	an.default_recommendation,
 	UPPER(CAST(an.gen_comp AS TEXT)) AS generate_comp_score_wo_credit, 
	an.fa_credit_score AS no_ssn_score, 
	an.score_without_credit, 
/* Automatic Service */
	--tsc.ts_credit_report, 6/20/18
 case when coalesce(cred_tier,0)=0 then null else ns.cred_tier end cred_tier, 
(select max(vendor_name) from services  where ns.cred = service_id) credit_report,


 case when coalesce(civilcrt_tier,0)=0 then null else ns.civilcrt_tier end civilcrt_tier,
   (select max(vendor_name) from services  where ns.civilcrt = service_id) civil_court,


 case when coalesce(rhist_tier,0)=0 then null else ns.rhist_tier end rhist_tier,
  (select max(vendor_name) from services  where ns.rhist = service_id) rental_history,

case when coalesce(ofac_tier,0)=0 then null else ns.ofac_tier end ofac_tier,
 (select max(vendor_name) from services  where ns.ofac = service_id) ofac,  

 
case when coalesce(crim_tier,0)=0 then null else ns.crim_tier end crim_tier,
   (select max(vendor_name) from services  where ns.crim = service_id) criminal,

	--cc.custom_criminal, 
       case when coalesce(cust_crim_tier,0)=0 then null else cust_crim_tier end cust_crim_tier,     
(select max(vendor_name) from services  where ns.cust_crim = service_id) custom_criminal, 

--	ccsa.custom_criminal_submitafter,

 case when coalesce(offense_tier,0)=0 then null else offense_tier end offense_tier,    
  (select max(vendor_name) from services  where ns.offense = service_id) offense_alert, 

-- voi
 (select max(vendor_name) from services  where ns.voi = service_id) "VOI", 
		      case when coalesce(voi_tier,0)=0 then null else voi_tier end "VOI Tier",     
 (select max(vendor_name) from services  where ns.voi_ondmd = service_id)   "On Demand VOI",
  case coalesce(trim(enable_dollar_validation_2),'')
           when 'On' then 'Enable with Rejection'
           when 'Rejection Pending Workflow' then 'Enable with Pending Workflow'
           when 'Off' then 'Off' when '' then 'Off' end 
  as enable_dollar_valication,   ---3

 
 case when  enable_voi_notification_emails = true then 'Yes' else 'No' end   as send_notification_emails   ,  ---4

             an.enable_voi_embedded as "Enable VOI Embedded", 


  case when trim(lower(an.enable_work_number)) = 'ssv' then  ssv_num_months   end ssv_num_of_month, 
-- voi


   (select max(vendor_name) from services  where ns.cred_ondmd = service_id)   od_credit,
(select max(vendor_name) from services  where ns.civilcrt_ondmd = service_id)   od_civil_court,
  (select max(vendor_name) from services  where ns.rhist_ondmd = service_id)   od_rental,
(select max(vendor_name) from services  where ns.ofac_ondmd = service_id)   od_ofac,
  (select max(vendor_name) from services  where ns.crim_ondmd = service_id)   od_crim,
(select max(vendor_name) from services  where ns.cust_crim_ondmd = service_id)   od_cust_crim,
(select max(vendor_name) from services  where ns.offense_ondmd = service_id)   od_Offense,

	(case when ns.reeval_func=19 then 'On' else null end) od_reevaluation_request,
	(case when ns.sbond_func=20 then 'On' else null end) od_sure_deposit,

  
/*Group Scoring */
	CASE 	WHEN gsm.method_description ISNULL THEN gsm2.method_description
									ELSE gsm.method_description 
									END AS group_scoring, 
	CASE 	WHEN gsm.method_description ISNULL THEN gsuco.method_parameter_1
									ELSE gsu.method_parameter_1
									END AS scoring_table, 

/*Credit Report Options*/
	UPPER(CAST((credit_report&1 > 0) AS TEXT)) AS suppress_details, 
UPPER(CAST((credit_report&4 > 0) AS TEXT)) AS Suppress_Items_for_Review,
 
 --UPPER(CAST (include_logo as text)) include_logo , --4/13/2023
 
        UPPER(CAST((an.rental_history&4>0) AS TEXT)) AS show_tenant_information,
        UPPER(CAST((an.rental_history&8>0) AS TEXT)) AS show_collections,
        UPPER(CAST((an.rental_history&16>0) AS TEXT)) AS show_statement,
UPPER(CAST((an.rental_history&128>0) AS TEXT)) AS  rental_history_scoring, --RHscoring from NodeInfo where node_code = 'RGDE1';
	UPPER(CAST((an.rental_history&64>0) AS TEXT)) AS show_reasons,
	UPPER(CAST((rhs.late_payments_scoring) AS TEXT)) AS late_payments_scoring, 
	rhs.late_payments_max, 
	rhs.late_payments_period, 
	UPPER(CAST(rhs.nsf_scoring AS TEXT)) AS nsf_scoring, 
	rhs.nsf_max, 
	rhs.nsf_period, 
	UPPER(CAST(rhs.outstanding_balances_scoring AS TEXT)) AS outstanding_balances_scoring, 
	rhs.outstanding_balances_max, 
	rhs.outstanding_balances_period, 
	UPPER(CAST(rhs.write_offs_scoring AS TEXT)) AS write_offs_scoring, 
	rhs.write_offs_amount, 
	rhs.write_offs_period, 
	UPPER(CAST(rhs.collections_scoring AS TEXT)) AS collections_scoring, 
	rhs.collections_amount, 
	rhs.collections_period,

/*Interface Options */
/*new addition */
case yardi_interface when 1 then 'One-Way Interface' when 2 then 'Two-Way Interface' end as interface_functionality,
--(select display_value  from list_values where list_id =1 and "index" = pms_if limit 1) PMS_Interface,
 --(select display_value  from list_values where list_id =2 and "index" = oli_if limit 1) Online_Leasing_Interface,
 	CASE	WHEN an.disable_edit_applicant=0 THEN 'Do Not Disable'
				WHEN an.disable_edit_applicant=1 THEN 'Edit Applicant'
				WHEN an.disable_edit_applicant=2 THEN 'Enter New Applicant'
				WHEN an.disable_edit_applicant=3 THEN 'Edit and Enter New Applicant'
					ELSE NULL
					END AS disable_edit_applicant,

  	UPPER(CAST(include_checkpoint_msg as text)) include_checkpoint_msg,
	UPPER(CAST(an.enable_group_apps AS TEXT)) AS enable_group_apps,
	ylp.interface_username, 
	ylp.server_name AS database_server_name, 
	ylp.database_name, 
	ylp.platform, 
	ylp.interface_entity, 
	ylp.web_service_uri AS url, 
/* Risk Score Options */
	rs.a_active, 
	rs.a_startbreakpoint, 
	rs.a_endbreakpoint,
	rs.b_active, 
	rs.b_startbreakpoint,
	rs.b_endbreakpoint,
	rs.c_active, 
	rs.c_startbreakpoint,
	rs.c_endbreakpoint, 
	rs.d_active, 
	rs.d_startbreakpoint, 
	rs.d_endbreakpoint, 
	rs.f_active,
	rs.f_startbreakpoint, 
	rs.f_endbreakpoint, 
	rs.g_active, 
	rs.g_startbreakpoint, 
	rs.g_endbreakpoint, 
	rs.r_active, 
	rs.r_startbreakpoint, 
	rs.r_endbreakpoint,
	an.def_credit_score  AS no_risk_score_available,

to_char(nt.create_stamp , 'MM/DD/YYYY'::text)   || ' ' || to_char(nt.create_stamp, 'HH24:MI:SS'::text) AS setup_date,

	nt.node_code,
	nt.node_name AS property_name,

	/* NODE ADDRESS AND BILLING INFO*/
	nt.node_street_1, 
	nt.node_street_2, 
	nt.node_city, 
	nt.node_state, 
	nt.node_zip,
	--BA.billto_name, 
	--BA.billto_street_1, 
	--BA.billto_street_2, 
    --BA.billto_city
    regexp_replace(BA.billto_name::text, '([^[:ascii:]])', ' ','g') billto_name, 
    regexp_replace(BA.billto_street_1::text, '([^[:ascii:]])', ' ','g') billto_street_1, 
    regexp_replace(BA.billto_street_2::text, '([^[:ascii:]])', ' ','g') billto_street_2, 
	regexp_replace(BA.billto_city::text, '([^[:ascii:]])', ' ','g') billto_city, 
	BA.billto_state, 
	BA.billto_zip ,

  /*General Continued*/
 
	nt.contact_name AS contact,
	nt.contact_phone AS phone,
	nt.fax_phone AS fax,
	an.email_address, 
	an.email_address_man AS manager_email,
  email_for_invoices,
 
 /* Property Information*/
an.region, 
	an.manager, 
	an.manager2, 
	an.manager3  , nova_credit_scoring, nova_risk_score, nt.market_rate_units,yardi_prop_code,

/*aal setting */

 case when crim_email='t' then 'Yes' else 'No' end   as  "Email Conditional Offer of Pending Housing", ---4/7/2023
case when crim_pre_aal = 'Yes' then 'Enabled' else 'Disabled' end as  PreAdverse_Action_Letter, 
 case   preaal_template_id when 36 then 'Pre-AAL Cook County' when 38 then 'Test_PreAAL'  when 40 then 'Generic Pre-AAL' else '' end  as  PreAdverse_Action_Letter_Template, 

reconsider_request_period || ' ' ||  reconsider_period_in as Reconsideration_Request_Period, 
reconsider_review_period || ' ' || reconsider_period_in as Reconsideration_Review_Period,
 case when coalesce(an.b2b_screening_workflow,false)= false then 'Off' else 'On' end as Alternate_Screening_Criteria_Rules , 
case when icraa = 'No'  then 'Disable' when icraa = 'Yes' then 'Enabled' else  'Not Applicable' end "ICRAA",
username as "Property Added By"
FROM node_tbl nt
LEFT JOIN account_tbl as BA ON nt.node_code = BA.account_code
LEFT JOIN  auxiliary_node an ON nt.node_code=an.node_code    -----changed
LEFT JOIN credit_products cp ON an.cred_prod_id=cp.cred_prod_id
LEFT JOIN group_score_users gsu ON gsu.node_or_company_code=nt.node_code
LEFT JOIN group_score_methods gsm ON gsu.method=gsm.method
LEFT JOIN group_score_users gsuco ON gsuco.node_or_company_code=nt.company_code /* We join the group scoring tables again by company for those nodes that take on the company score*/
LEFT JOIN group_score_methods gsm2 ON gsuco.method=gsm2.method
LEFT JOIN auxiliary_company ac ON nt.company_code=ac.company_code
LEFT JOIN rhs_policies rhs ON nt.node_code=rhs.policy_code
LEFT JOIN yardi_login_parameters ylp ON nt.node_code=ylp.node_code
  
LEFT JOIN (
					SELECT 
						rpd.node_code, 
						CASE 	WHEN start_break_point ISNULL THEN 'FALSE'
									ELSE 'TRUE' 
									END AS a_active, 	
						rpd.start_break_point AS a_startbreakpoint, 
						rpd.end_break_point AS a_endbreakpoint, 
						CASE 	WHEN brpd.b_startbreakpoint ISNULL THEN 'FALSE'
									ELSE 'TRUE' 
									END AS b_active, 	
						brpd.b_startbreakpoint,
						brpd.b_endbreakpoint, 
						CASE 	WHEN crpd.c_startbreakpoint ISNULL THEN 'FALSE'
									ELSE 'TRUE' 
									END AS c_active, 	
						crpd.c_startbreakpoint,
						crpd.c_endbreakpoint, 
						CASE 	WHEN drpd.d_startbreakpoint ISNULL THEN 'FALSE'
									ELSE 'TRUE' 
									END AS d_active,
						drpd.d_startbreakpoint,
						drpd.d_endbreakpoint, 
						CASE 	WHEN frpd.f_startbreakpoint ISNULL THEN 'FALSE'
									ELSE 'TRUE' 
									END AS f_active,
						frpd.f_startbreakpoint,
						frpd.f_endbreakpoint, 
						CASE 	WHEN grpd.g_startbreakpoint ISNULL THEN 'FALSE'
									ELSE 'TRUE' 
									END AS g_active,
						grpd.g_startbreakpoint,
						grpd.g_endbreakpoint, 
						CASE 	WHEN rrpd.r_startbreakpoint ISNULL THEN 'FALSE'
									ELSE 'TRUE' 
									END AS r_active,
						rrpd.r_startbreakpoint,
						rrpd.r_endbreakpoint
					FROM 	risk_score_policy_details rpd
					LEFT JOIN 	(
											SELECT 
												node_code, 
												CASE 	WHEN start_break_point ISNULL THEN 'FALSE'
															ELSE 'TRUE' 
															END AS b_active,
												start_break_point AS b_startbreakpoint, 
												end_break_point AS b_endbreakpoint	
											FROM 	risk_score_policy_details
											WHERE score = 'B'
											) brpd ON brpd.node_code=rpd.node_code
					LEFT JOIN 	(
											SELECT 
												node_code,  
												CASE 	WHEN start_break_point ISNULL THEN 'FALSE'
															ELSE 'TRUE' 
															END AS c_active,
												start_break_point AS c_startbreakpoint, 
												end_break_point AS c_endbreakpoint	
											FROM 	risk_score_policy_details
											WHERE score = 'C'
											) crpd ON crpd.node_code=rpd.node_code
					LEFT JOIN 	(
											SELECT 
												node_code,  
												CASE 	WHEN start_break_point ISNULL THEN 'FALSE'
															ELSE 'TRUE' 
															END AS d_active,
												start_break_point AS d_startbreakpoint, 
												end_break_point AS d_endbreakpoint	
											FROM 	risk_score_policy_details
											WHERE score = 'D'
											) drpd ON drpd.node_code=rpd.node_code
					LEFT JOIN 	(
											SELECT 
												node_code, 
												CASE 	WHEN start_break_point ISNULL THEN 'FALSE'
															ELSE 'TRUE' 
															END AS f_active,
												start_break_point AS f_startbreakpoint, 
												end_break_point AS f_endbreakpoint	
											FROM 	risk_score_policy_details
											WHERE score = 'F'
											) frpd ON frpd.node_code=rpd.node_code
					LEFT JOIN 	(
											SELECT 
												node_code, 
												CASE 	WHEN start_break_point ISNULL THEN 'FALSE'
															ELSE 'TRUE' 
															END AS g_active,
												start_break_point AS g_startbreakpoint, 
												end_break_point AS g_endbreakpoint	
											FROM 	risk_score_policy_details
											WHERE score = 'G'
											 
											) grpd ON grpd.node_code=rpd.node_code
					LEFT JOIN 	(
											SELECT 
												node_code, 
												CASE 	WHEN start_break_point ISNULL THEN 'FALSE'
															ELSE 'TRUE' 
															END AS r_active,
												start_break_point AS r_startbreakpoint, 
												end_break_point AS r_endbreakpoint	
											FROM 	risk_score_policy_details
											WHERE score = 'R'
											) rrpd ON rrpd.node_code=rpd.node_code
      
					WHERE score = 'A'
					) AS rs ON rs.node_code=nt.node_code
         LEFT JOIN (
                      SELECT comp_code, node_code, autoemail_fcraletter, hold_until_active_app_processs, send_fcraletter_after_hours, created_on::date 
                        FROM autogenerate_fcraletter_config --where node_code = 'K7956'

             ) fc on fc.node_code = nt.node_code and fc.comp_code = nt.company_code
             LEFT JOIN node_services ns on    trim(ns.pr_code) = trim(nt.node_code) and trim(ns.co_code) = trim(nt.company_code)
             LEFT Join (select who, string_agg(username, ',') username from audit_tbl  where lower(change_details) like '%%added new property%%' group by who ) audit on audit.who = nt.node_code 
WHERE nt.canceled <> 'T'  
  and nt.create_stamp::DATE BETWEEN '%s' AND  '%s'
ORDER BY nt.company_code,
	nt.policy , nt.create_stamp ; --  limit 11; 
               """ % (w_start_date,  w_end_date))

 
     return sql


def connect_setup_audit(writer, wb):
    """ Connect to the PostgreSQL database server """
     
    o_policy ="" 
 
    conn = None
    params = config()
    
    conn = psycopg2.connect(**params )
  
    cur = conn.cursor()
  
    cur.execute(setup_sql()) 
    col = cur.description 
            
    rows = cur.fetchall()
            
    df = pd.DataFrame(rows, columns = col)

    sheet_name = "Setup Audit"   
   
         
    df.to_excel(writer,sheet_name=sheet_name, index=False )
    df = df.replace(np.nan, '', regex=True)
    
    ws = writer.sheets[sheet_name]
            #format
    for i in range(0, len(col)):
        ws.write(0, i, col[i][0])
        ws.set_column(i, i, len(col[i][0])+2)
    
            
    master_row  = []
    header_fmt = wb.add_format({'bold': True})
   

    row_format_top = wb.add_format({'top': 5})
	# write the first row - heading
 

 
    ws.set_row(0, None, header_fmt)
    ws.autofilter('A1:FL1')
    #mycell = ws['B3']
    ws.freeze_panes(1,2)
    #ws.freeze_panes(1,0)
   
    cur.close()
    conn.close()
 
def new_template(writer, wb ):
    df_template = pd.read_excel(file_template)
    df_template.to_excel(writer,sheet_name="What's New", index=False )
    ws = writer.sheets["What's New"]
    col = df_template.columns
    for i in range(0, len(col)):
        #ws.write(0, i, col[i])
        ws.set_column(i, i, len(col[i])+20)
        

    ws.freeze_panes(1,0)
   
     
def connect_compare(writer, wb):
    """ Connect to the PostgreSQL database server """
    
  
    o_policy ="" 
 
    conn = None
    params = config()
    
    conn = psycopg2.connect(**params )
    #com_list = ['AIY']
    #com_list = ['AIY','ANTER','ASI','ASLI','AVAN','CHAH','COMBU','DECRO','GHAI','HORN','LAR','NRES','PINI','PRMAI','REDW','RHM','SENTI','WALTC','WPMC','WRH']
    o_companycode=''
    o_policy =''
 
    cur = conn.cursor()
    #writer = pd.ExcelWriter(filename1, engine='xlsxwriter', options={'nan_inf_to_errors': False} )  
     
     
    cur.execute(compare_sql()) 
    col = [i[0] for i in cur.description]
            
    rows = cur.fetchall()
            
    df = pd.DataFrame(rows )
            
    df.to_excel(writer,sheet_name=s_name_compare, index=False )
    df = df.replace(np.nan, '', regex=True)
            #format
   
    ws = writer.sheets[s_name_compare]
    master_row=[]
             
	# write the first row - heading
	# only compare, not write
    for i in range(0, len(col)):
        ws.write(0, i, col[i])
        ws.set_column(i, i, len(col[i])+2)
    

    start_row = 0
    df['new_company'] = False
    df['all_color'] = False
    df['diff'] = 0
    mismatch = False 
    start_row = 0
    for index, row in df.iterrows():
       
            
        if (o_companycode != str(row[1]) or o_policy != str(row[2])):

            
            df.iloc[index, df.columns.get_loc('new_company')] = True
			# msater row : with new company code (col 1) and policy (col 2) - mark M 
            #row_num = index + 1 # plus header row - 1st row
            master_row = row
         
            o_companycode = str(row[1])
            o_policy = str(row[2])
            if  (mismatch == True):                    
                    for i in range(start_row, index ):
                        df.iloc[i, df.columns.get_loc('all_color')] = True
            mismatch = False
            start_row = index
        else: # not master row need compar
            diff = 0
            for col_num  in range(5, len(df.columns)-3):  # addtional new company and highlight
                if ( master_row[col_num]  != row[col_num]  and stupid_col(col_num)):
            #        ws.write(index+1, col_num, row[col_num],cell_format_hightlight)
                    diff = diff + 1
                    if (diff > 0) :
                        #ws.write(index+1, 0, "+", cell_format_red)
                        mismatch   = True
            df.iloc[index, df.columns.get_loc('diff')] = diff
    if  (mismatch == True):       # for last instance              
                    for i in range(start_row,  index + 1 ):
                        df.iloc[i, df.columns.get_loc('all_color')] = True        
        
            
    test_sheet(ws,df, col )
    cur.close()
    conn.close()
    logger.info("Last Company - %s", o_companycode)
    logger.info("Last polocy -  %s", o_policy)
    logger.info('Log Entry End Here.'  )
    
def test_sheet(ws, df, col) :
    
     
    header_fmt = wb.add_format({'bold': True})
    cell_format_red = wb.add_format({ 'font_color': 'red' })
    
    cell_format_hightlight = wb.add_format({  'bg_color': '#FFC7CE'})
    master_diff_hightlight= wb.add_format({  'bg_color': '#FFC7CE','font_color': 'red'})
    
    master_first_hightlight= wb.add_format({  'bg_color': '#FFC7CE',  'top':5})
    row_format_top = wb.add_format({'top': 5})
     
     #df = df.applymap(lambda x: x.upper() if type(x) == str else x)
    '''
    for i in range(0, len(col)-3):
        ws.write(0, i, col[i][1] )
        ws.set_column(i, i, len(col  )+2)
    '''
	# process data and rest of row
    mismatch = False
    master_row = []

    start_row =0
 
    for index, row in df.iterrows():
        
        if (index==0):
          master_row = row
    
        if (  df.iloc[index, df.columns.get_loc('new_company')]):
 		# msater row : with new company code (col 1) and policy (col 2) - mark M 
            row_num = index + 1 # plus header row - 1st row
            start_row = row_num
            master_row = row
            if df.iloc[index, df.columns.get_loc('all_color')]==True:
                ws.write(row_num, 0, 'M',  master_first_hightlight) 
            else:
                ws.write(row_num, 0, 'M')
            ws.set_row(row_num, 15, row_format_top)
            
              
        else:

            #diff = 0
            #mismatch = False
            #for col_num  in range(5, len(df.columns)-3):
            for col_num  in range(5, len(col)):
                
                if ( master_row[col_num]  != row[col_num]  and stupid_col(col_num)):
                    ws.write(index+1, col_num, row[col_num],cell_format_hightlight)
                   
		
        if (df.iloc[index, df.columns.get_loc('diff')]) > 0:
            ws.write (index+1, 0 , "*" + str(df.iloc[index, df.columns.get_loc('diff')]), master_diff_hightlight)
        else:
            if df.iloc[index, df.columns.get_loc('all_color')] and (not df.iloc[index, df.columns.get_loc('new_company')]):
                ws.write(index + 1, 0, "",cell_format_hightlight) 
        
 
   
    ws.set_row(0, None, header_fmt)
    ws.autofilter('A1:DZ1')
    #mycell = ws['B3']
    ws.freeze_panes(1,5)
    #ws.freeze_panes(1,0)
 

def stupid_col(col):
	col_not_to_comp = [0,1,2,3,4]
	for i in col_not_to_comp:
		if col == i:
			return False
	return True




def send_to_qa(filename1 , filename2="" ):


    signature_file = "C:\App\Data\Sig\sig.htm"
    #signature_path = os.path.join((os.environ['USERPROFILE']), sig_files_path) # Finds the path to Outlook signature files with signature name "Work"
    #html_doc = os.path.join((os.environ['USERPROFILE']),sig_html_path)     #Specifies the name of the HTML version of the stored signature
    #html_doc = html_doc.replace('\\\\', '\\')


    html_file = codecs.open(signature_file, 'r', 'utf-8', errors='ignore') #Opens HTML file and converts to UTF-8, ignoring errors
    signature_code = html_file.read()               #Writes contents of HTML signature file to a string

    #signature_code = signature_code.replace((signature_name + '_files/'), signature_path)      #Replaces local directory with full directory path
    html_file.close()

    
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = "xiaobin.zhang@yardi.com"

    
    mail.Subject = "Weekly QA Setup Audit - RS_QA@Yardi.Com ;   "  
    
    #filename2 =  filename1.replace('O:\\ANALYTICS\\New Property Lists\\', '\\\ysifwfs07\\Vol2\ANALYTICS\\New Property Lists\\') # + filename.replace("/", ".") + ".xlsx"
        #path  = "\"\\\\windows_Server\\golobal_directory\\the folder\\file yyymm.xlsx\""
    path = '"' + filename1 + '"'
    string  = """<a href=""" +  path + ' style=text-decoration: none>' + filename1 +  '<' +  r'\a'  + '>'
    string =  string.replace('O:\\ANALYTICS\\Setup_Audit\\Setup Audit Weekly Reports\\', '\\\ysifwfs07\\Vol2\ANALYTICS\\Setup_Audit\\Setup Audit Weekly Reports\\')

    path_2 = '"' + filename2 + '"'
    string_2  = """<a href=""" +  path_2 + ' style=text-decoration: none>' + filename2 +  '<' +  r'\a'  + '>'
    string_2 =  string_2.replace('O:\\ANALYTICS\\Setup_Audit\\', '\\\ysifwfs07\\Vol2\ANALYTICS\\Setup_Audit\\')
    
      #  string.replace('\\a>', '\a>')
    #mail.body = string
     
    mail.HTMLbody =   string + " <BR> " + string_2 +" <BR><BR><BR> "  +signature_code + " <BR><BR><BR> "
    
    mail.send
    

def Emailer(message, subject, recipient):
    import win32com.client as win32   

    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = recipient
    mail.Subject = subject
    mail.GetInspector 

    index = mail.HTMLbody.find('>', mail.HTMLbody.find('<body')) 
    mail.HTMLbody = mail.HTMLbody[:index + 1] + message + mail.HTMLbody[index + 1:] 
    #mail.Attachments.Add(attachment)
    mail.Display(True)
    #mail.Send()

if __name__ == '__main__':
#try:
    writer = pd.ExcelWriter(filename1, engine='xlsxwriter'  )  
    wb  = writer.book
    connect_compare(writer, wb)
    connect_setup_audit(writer,wb)
    new_template (writer, wb)
    writer.close()  
    
#except Exception as e:
 #logger.info(e)

#attachment  = '"' +  (dirpath + '\\' + filename1) + '"'
    #attachment  = '"' + filename1 + '"'
#print (attachment)
    #filelink = """<p>  <a href=%s>%s</a>  </p>"""%(attachment, attachment)
    
    send_to_qa(filename1  ) 
    

    #print (filelink)
 


sys.exit(0)


quit()
